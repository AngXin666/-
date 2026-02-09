"""
GUI 界面模块
GUI Interface Module
"""

import asyncio
import logging
import os
import threading
import time
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from pathlib import Path
from typing import Optional, List
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

from .config import ConfigLoader
from .emulator_controller import EmulatorController, EmulatorType
from .adb_bridge import ADBBridge
from .account_manager import AccountManager
from .screen_capture import ScreenCapture
from .ui_automation import UIAutomation
from .auto_login import AutoLogin
from .ximeng_automation import XimengAutomation
from .logger import get_logger
from .models.models import AccountResult
from .models.error_types import ErrorType
from .local_db import LocalDatabase
from .selection_manager import SelectionManager

try:
    import yaml
except ImportError:
    yaml = None


class AutomationGUI:
    """自动化脚本 GUI 界面"""
    
    def __init__(self, adb_bridge=None):
        """初始化GUI
        
        Args:
            adb_bridge: ADBBridge实例（可选），用于后台加载模型
        """
        self.adb_bridge = adb_bridge  # 保存ADB实例
        self.models_loaded = False  # 模型加载状态标志
        
        self.root = tk.Tk()
        self.root.title("自动签到助手 v2.0.6")
        self.root.geometry("1000x950")
        self.root.resizable(True, True)
        
        # 设置默认字体，确保中文正确显示
        import tkinter.font as tkfont
        default_font = tkfont.nametofont("TkDefaultFont")
        default_font.configure(family="Microsoft YaHei UI", size=9)
        text_font = tkfont.nametofont("TkTextFont")
        text_font.configure(family="Microsoft YaHei UI", size=9)
        
        # 许可证信息
        self.license_info = None
        # 延迟加载许可证信息，避免阻塞GUI初始化
        self.root.after(100, self._load_license_info_async)
        
        # 配置
        self.config_loader = ConfigLoader()
        self.config = self.config_loader.load()
        self.is_running = False
        self.is_paused = False  # 暂停状态
        self.emulator_controller: Optional[EmulatorController] = None
        
        # 初始化勾选状态管理器
        self.selection_manager = SelectionManager()
        
        # 提前初始化界面变量（避免在 _create_widgets 完成前被访问）
        self.auto_launch_var = tk.BooleanVar(value=True)
        self.enable_cache_var = tk.BooleanVar(value=True)
        self.parallel_var = tk.BooleanVar(value=True)
        self.instance_count_var = tk.IntVar(value=1)
        self.max_retries_var = tk.IntVar(value=3)
        self.stuck_timeout_var = tk.IntVar(value=15)
        self.max_wait_time_var = tk.IntVar(value=60)
        self.launch_timeout_var = tk.IntVar(value=120)
        self.switch_delay_var = tk.IntVar(value=3)
        
        # 定时运行配置
        self.scheduled_run_enabled = tk.BooleanVar(value=False)
        self.scheduled_run_time = tk.StringVar(value="08:00")
        self.scheduled_hour_var = tk.StringVar(value="08")
        self.scheduled_minute_var = tk.StringVar(value="00")
        self.last_scheduled_run_date = None  # 记录上次定时运行的日期
        self.schedule_check_thread = None  # 定时检查线程
        
        # 线程控制事件(改进：使用Event对象，线程可以更快响应)
        self.stop_event = threading.Event()  # 停止事件
        self.pause_event = threading.Event()  # 暂停事件
        
        # 线程池(用于并发处理账号任务)
        # 最大线程数设置为 10，支持后期多线程并发
        # 当前通过 instance_count_var 控制实际并发数
        self.executor = None  # 延迟初始化，在运行时创建
        self.pending_futures = []  # 存储待处理的 Future 对象
        
        # 统计锁(用于多线程安全地更新统计数据)
        self.stats_lock = threading.Lock()
        
        # 已保存到数据库的账号集合（防止重复保存）
        # 格式：{phone: timestamp}，每次运行开始时清空
        self.saved_accounts = {}
        self.saved_accounts_lock = threading.Lock()
        
        # 模拟器实例池管理
        self.instance_pool = []  # 可用的实例编号池 [0, 1, 2, ...]
        self.instance_lock = threading.Lock()  # 实例池访问锁
        
        # 设置窗口关闭协议
        self.root.protocol("WM_DELETE_WINDOW", self._on_closing)
        
        # 创建界面
        self._create_widgets()
        self._load_config_to_ui()
        
        # 窗口居中显示
        self._center_window()
        
        # 如果提供了ADB实例，在GUI显示后后台加载模型
        if self.adb_bridge:
            self.root.after(500, self._load_models_in_background)
        else:
            # 没有提供ADB实例，检查模型是否已加载（兼容旧的启动方式）
            self._display_model_loading_status()
        
        # 自动检测模拟器
        self._auto_detect_emulator()
        
        # 自动检测运行中的实例并调整并行数（延迟2秒执行，确保GUI完全加载）
        self.root.after(2000, lambda: self._detect_running_instances(auto_adjust=True))
        
        # 自动检测并注册新模型（延迟1秒执行，确保GUI完全加载）
        self.root.after(1000, self._auto_check_new_models)
    
    def _center_window(self):
        """将主窗口居中显示在屏幕中间"""
        self.root.update_idletasks()
        
        # 获取窗口实际大小
        width = 1000
        height = 950
        
        # 相对于屏幕居中
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        x = (screen_width // 2) - (width // 2)
        y = (screen_height // 2) - (height // 2)
        
        self.root.geometry(f'{width}x{height}+{x}+{y}')
    
    def _load_models_in_background(self):
        """在后台线程中加载模型"""
        self._log("正在后台加载模型，请稍候...")
        
        def load_models_thread():
            """后台加载模型的线程函数"""
            try:
                from .model_manager import ModelManager
                model_manager = ModelManager.get_instance()
                
                # 定义进度回调
                def progress_callback(message, current, total):
                    self.root.after(0, lambda m=message, c=current, t=total: 
                                   self._log(f"[{c}/{t}] {m}"))
                
                # 加载所有模型
                stats = model_manager.initialize_all_models(
                    adb_bridge=self.adb_bridge,
                    log_callback=lambda msg: self.root.after(0, lambda m=msg: self._log(m)),
                    progress_callback=progress_callback
                )
                
                # 标记模型已加载
                self.models_loaded = True
                
                # 在主线程中显示简洁的加载完成信息
                self.root.after(0, lambda: self._log(f"✅ 程序已准备就绪"))
                
                if stats['errors']:
                    self.root.after(0, lambda: self._log(f"⚠️ 警告: {len(stats['errors'])} 个模型加载失败"))
                
            except Exception as e:
                self.root.after(0, lambda err=str(e): self._log(f"❌ 模型加载失败: {err}"))
                import traceback
                traceback.print_exc()
        
        # 启动后台线程
        thread = threading.Thread(target=load_models_thread, daemon=True)
        thread.start()
    
    def _display_model_loading_status(self):
        """显示模型加载状态信息"""
        try:
            from .model_manager import ModelManager
            model_manager = ModelManager.get_instance()
            
            # 获取加载统计信息
            stats = model_manager.get_loading_stats()
            
            print("\n" + "=" * 60)
            print("模型加载状态")
            print("=" * 60)
            print(f"已加载模型: {stats['loaded_models']}/{stats['total_models']}")
            print(f"加载时间: {stats['total_time']:.2f}秒")
            print(f"内存占用: {stats['memory_after'] / 1024 / 1024:.1f}MB")
            
            if stats['errors']:
                print(f"警告: {len(stats['errors'])} 个模型加载失败")
            
            print("=" * 60)
            
        except Exception as e:
            print(f"获取模型加载状态失败: {e}")
    
    def _create_widgets(self):
        """创建界面组件"""
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # === 配置区域 ===
        config_frame = ttk.LabelFrame(main_frame, text="配置", padding="5")
        config_frame.pack(fill=tk.X, pady=(0, 10))
        
        # 许可证信息标签（放在配置标题的右侧）
        self.license_label = tk.Label(
            main_frame,
            text="",
            font=("Microsoft YaHei UI", 9),
            fg="#00AA00"  # 绿色
        )
        # 定位到配置框架标题栏的右侧，与"配置"文字垂直对齐
        self.license_label.place(in_=config_frame, relx=1.0, y=-12, anchor=tk.E, x=-10)
        
        # 更新许可证显示
        self._update_license_display()
        
        # 更新许可证显示
        self._update_license_display()
        
        # 模拟器路径
        row1 = ttk.Frame(config_frame)
        row1.pack(fill=tk.X, pady=2)
        ttk.Label(row1, text="模拟器路径:", width=12).pack(side=tk.LEFT)
        self.emulator_path_var = tk.StringVar()
        # 路径输入框设置为只读，用户不能手动输入
        path_entry = ttk.Entry(row1, textvariable=self.emulator_path_var, state='readonly')
        path_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        ttk.Button(row1, text="选择盘符", command=self._select_drive_and_search, width=10).pack(side=tk.LEFT)
        ttk.Button(row1, text="自动检测", command=self._auto_detect_emulator, width=10).pack(side=tk.LEFT, padx=(5, 0))
        
        # 模拟器类型选择
        row1b = ttk.Frame(config_frame)
        row1b.pack(fill=tk.X, pady=2)
        ttk.Label(row1b, text="模拟器类型:", width=12).pack(side=tk.LEFT)
        self.emulator_type_var = tk.StringVar(value="未检测")
        
        # 创建下拉选择框
        self.emulator_type_combo = ttk.Combobox(row1b, textvariable=self.emulator_type_var, 
                                                 state='readonly', width=20)
        self.emulator_type_combo['values'] = ('自动检测', 'MuMu模拟器')
        self.emulator_type_combo.current(0)  # 默认选择"自动检测"
        self.emulator_type_combo.pack(side=tk.LEFT, padx=(0, 10))
        self.emulator_type_combo.bind('<<ComboboxSelected>>', self._on_emulator_type_changed)
        
        # 添加应用按钮
        ttk.Button(row1b, text="🔄 应用选择", command=self._apply_emulator_selection, width=12).pack(side=tk.LEFT)
        
        # 状态标签
        self.emulator_status_var = tk.StringVar(value="")
        ttk.Label(row1b, textvariable=self.emulator_status_var, foreground="blue").pack(side=tk.LEFT, padx=(10, 0))
        
        # 账号文件路径（隐藏，只在内部使用）
        self.accounts_file_var = tk.StringVar()
        
        # 并发数量和实例检测
        # 多开窗口数量和重试次数配置（同一行）
        row4 = ttk.Frame(config_frame)
        row4.pack(fill=tk.X, pady=2)
        ttk.Label(row4, text="多开窗口数量:", width=12).pack(side=tk.LEFT)
        self.instance_count_var = tk.IntVar(value=1)
        # 添加实时保存回调
        self.instance_count_var.trace_add('write', lambda *args: self._auto_save_config())
        ttk.Spinbox(row4, from_=1, to=10, textvariable=self.instance_count_var, width=10).pack(side=tk.LEFT)
        
        # 重试次数（同一行）
        ttk.Label(row4, text="重试次数:", width=12).pack(side=tk.LEFT, padx=(15, 0))
        self.max_retries_var = tk.IntVar(value=3)
        # 添加实时保存回调
        self.max_retries_var.trace_add('write', lambda *args: self._auto_save_config())
        ttk.Spinbox(row4, from_=1, to=10, textvariable=self.max_retries_var, width=6).pack(side=tk.LEFT)
        
        ttk.Button(row4, text="🔍 检测运行中的实例", command=self._detect_running_instances, width=18).pack(side=tk.LEFT, padx=(10, 0))
        
        # 运行中的实例显示
        self.running_instances_var = tk.StringVar(value="未检测")
        ttk.Label(row4, textvariable=self.running_instances_var, foreground="green").pack(side=tk.LEFT, padx=(10, 0))
        
        # 自动转账开关(放在检测实例按钮右侧)
        from .toggle_switch import ToggleSwitch
        
        # 创建一个Frame来容纳开关和标签
        transfer_switch_frame = ttk.Frame(row4)
        transfer_switch_frame.pack(side=tk.LEFT, padx=(20, 10))
        
        ttk.Label(transfer_switch_frame, text="自动转账:").pack(side=tk.LEFT, padx=(0, 5))
        
        self.auto_transfer_switch = ToggleSwitch(
            transfer_switch_frame, 
            width=60, 
            height=28,
            command=self._on_auto_transfer_changed
        )
        self.auto_transfer_switch.pack(side=tk.LEFT)
        
        # 定时运行配置（新增一行）
        row5 = ttk.Frame(config_frame)
        row5.pack(fill=tk.X, pady=2)
        
        # 定时运行开关
        scheduled_switch_frame = ttk.Frame(row5)
        scheduled_switch_frame.pack(side=tk.LEFT, padx=(0, 10))
        
        ttk.Label(scheduled_switch_frame, text="定时运行:").pack(side=tk.LEFT, padx=(0, 5))
        
        self.scheduled_run_switch = ToggleSwitch(
            scheduled_switch_frame,
            width=60,
            height=28,
            command=self._on_scheduled_run_changed
        )
        self.scheduled_run_switch.pack(side=tk.LEFT)
        
        # 定时运行时间设置（使用Spinbox滚轮选择）
        ttk.Label(row5, text="运行时间:").pack(side=tk.LEFT, padx=(10, 5))
        
        # 小时选择（00-23）
        self.scheduled_hour_var = tk.StringVar(value="08")
        hour_spinbox = ttk.Spinbox(
            row5, 
            from_=0, 
            to=23, 
            textvariable=self.scheduled_hour_var, 
            width=4,
            format="%02.0f",
            command=self._on_scheduled_time_changed
        )
        hour_spinbox.pack(side=tk.LEFT)
        hour_spinbox.bind('<FocusOut>', lambda e: self._on_scheduled_time_changed())
        hour_spinbox.bind('<Return>', lambda e: self._on_scheduled_time_changed())
        
        ttk.Label(row5, text=":").pack(side=tk.LEFT)
        
        # 分钟选择（00-59）
        self.scheduled_minute_var = tk.StringVar(value="00")
        minute_spinbox = ttk.Spinbox(
            row5, 
            from_=0, 
            to=59, 
            textvariable=self.scheduled_minute_var, 
            width=4,
            format="%02.0f",
            command=self._on_scheduled_time_changed
        )
        minute_spinbox.pack(side=tk.LEFT)
        minute_spinbox.bind('<FocusOut>', lambda e: self._on_scheduled_time_changed())
        minute_spinbox.bind('<Return>', lambda e: self._on_scheduled_time_changed())
        
        # 定时运行状态显示
        self.scheduled_status_var = tk.StringVar(value="")
        ttk.Label(row5, textvariable=self.scheduled_status_var, foreground="blue").pack(side=tk.LEFT, padx=(20, 0))
        
        # 添加账号统计信息(在自动转账开关右侧)
        self.account_total_var = tk.StringVar(value="账号总计: 0")
        self.account_pending_var = tk.StringVar(value="待处理: 0")
        self.status_var = tk.StringVar(value="就绪")
        
        ttk.Label(row5, textvariable=self.account_total_var, foreground="blue").pack(side=tk.LEFT, padx=(10, 0))
        ttk.Label(row5, textvariable=self.account_pending_var, foreground="orange").pack(side=tk.LEFT, padx=(10, 0))
        ttk.Label(row5, textvariable=self.status_var, foreground="green").pack(side=tk.LEFT, padx=(10, 0))

        # === 控制按钮区域 ===
        control_frame = ttk.Frame(main_frame)
        control_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.start_btn = ttk.Button(control_frame, text="▶ 开始运行", command=self._start_automation, width=12)
        self.start_btn.pack(side=tk.LEFT, padx=(0, 5))
        
        self.pause_btn = ttk.Button(control_frame, text="⏸ 暂停", command=self._pause_automation, width=8, state=tk.DISABLED)
        self.pause_btn.pack(side=tk.LEFT, padx=(0, 5))
        
        self.stop_btn = ttk.Button(control_frame, text="■ 停止", command=self._stop_automation, width=8, state=tk.DISABLED)
        self.stop_btn.pack(side=tk.LEFT, padx=(0, 5))
        
        ttk.Button(control_frame, text="💰 转账配置", command=self._open_transfer_config, width=10).pack(side=tk.LEFT, padx=(0, 5))
        
        ttk.Button(control_frame, text="📜 转账历史", command=self._open_transfer_history, width=10).pack(side=tk.LEFT, padx=(0, 5))
        
        ttk.Button(control_frame, text="📊 历史结果", command=self._open_history_results, width=10).pack(side=tk.LEFT, padx=(0, 5))
        
        ttk.Button(control_frame, text="👥 用户管理", command=self._open_user_management, width=10).pack(side=tk.LEFT, padx=(0, 5))
        
        ttk.Button(control_frame, text="⚙️ 流程控制", command=self._open_workflow_control, width=10).pack(side=tk.LEFT, padx=(0, 5))
        
        # 注册新模型按钮
        ttk.Button(control_frame, text="🔄 注册新模型", command=self._register_new_models, width=11).pack(side=tk.LEFT, padx=(0, 5))
        
        # 窗口排列按钮
        ttk.Button(control_frame, text="📐 窗口排列", command=self._open_window_arranger, width=10).pack(side=tk.LEFT, padx=(0, 5))
        
        # === 进度区域 ===
        progress_frame = ttk.LabelFrame(main_frame, text="进度", padding="5")
        progress_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.progress_var = tk.DoubleVar(value=0)
        self.progress_bar = ttk.Progressbar(progress_frame, variable=self.progress_var, maximum=100)
        self.progress_bar.pack(fill=tk.X, pady=(0, 5))
        
        self.progress_label_var = tk.StringVar(value="等待开始...")
        ttk.Label(progress_frame, textvariable=self.progress_label_var).pack(anchor=tk.W)
        
        # === 日志区域 ===
        log_frame = ttk.LabelFrame(main_frame, text="运行日志 (双击清空)", padding="5")
        log_frame.pack(fill=tk.X, expand=False, pady=(0, 10))
        
        # 添加实例过滤器(在日志框上方)
        filter_row = ttk.Frame(log_frame)
        filter_row.pack(fill=tk.X, pady=(0, 5))
        ttk.Label(filter_row, text="查看日志:", width=10).pack(side=tk.LEFT)
        self.log_filter_var = tk.StringVar(value="全部")
        log_filter_combo = ttk.Combobox(filter_row, textvariable=self.log_filter_var, 
                                        state='readonly', width=15)
        log_filter_combo['values'] = ('全部', '实例0', '实例1', '实例2', '实例3', '实例4')
        log_filter_combo.pack(side=tk.LEFT, padx=(0, 10))
        log_filter_combo.bind('<<ComboboxSelected>>', self._on_log_filter_changed)
        
        # 存储所有日志(用于过滤)
        self.all_logs = []  # 存储所有日志消息
        
        self.log_text = scrolledtext.ScrolledText(log_frame, height=12, state=tk.DISABLED)
        self.log_text.pack(fill=tk.BOTH, expand=False)
        self.log_text.bind("<Double-Button-1>", lambda e: self._clear_log())
        
        # 日志自动滚动控制
        self.log_auto_scroll = True  # 默认开启自动滚动
        
        # 绑定鼠标滚轮和滚动条事件，用户滑动时停止自动滚动
        self.log_text.bind("<MouseWheel>", self._on_log_scroll)
        self.log_text.bind("<Button-4>", self._on_log_scroll)  # Linux
        self.log_text.bind("<Button-5>", self._on_log_scroll)  # Linux
        
        # 绑定滚动条拖动事件
        log_scrollbar = self.log_text.vbar
        if log_scrollbar:
            log_scrollbar.bind("<B1-Motion>", self._on_log_scrollbar_drag)
        
        # === 错误日志区域 ===
        error_log_frame = ttk.LabelFrame(main_frame, text="错误日志", padding="5")
        error_log_frame.pack(fill=tk.X, expand=False, pady=(0, 10))
        
        # 存储所有错误日志
        self.all_error_logs = []  # 存储所有错误日志消息
        
        self.error_log_text = scrolledtext.ScrolledText(error_log_frame, height=6, state=tk.DISABLED)
        self.error_log_text.pack(fill=tk.BOTH, expand=False)
        
        # 配置错误日志文本颜色
        self.error_log_text.tag_configure("error", foreground="red")
        
        
        # === 统计区域 ===
        stats_frame = ttk.LabelFrame(main_frame, text="统计", padding="5")
        stats_frame.pack(fill=tk.X, pady=(0, 10))
        
        # 单行统计：总计、成功、失败、总余额、总签到奖励
        stats_row = ttk.Frame(stats_frame)
        stats_row.pack(fill=tk.X, pady=2)
        
        self.total_var = tk.StringVar(value="总计: 0")
        self.success_var = tk.StringVar(value="成功: 0")
        self.failed_var = tk.StringVar(value="失败: 0")
        self.total_balance_var = tk.StringVar(value="总余额: 0.00 元")
        self.total_checkin_reward_var = tk.StringVar(value="总签到奖励: 0.00 元")
        
        ttk.Label(stats_row, textvariable=self.total_var, width=12, foreground="blue").pack(side=tk.LEFT, padx=(0, 5))
        ttk.Label(stats_row, textvariable=self.success_var, width=12, foreground="green").pack(side=tk.LEFT, padx=(0, 5))
        ttk.Label(stats_row, textvariable=self.failed_var, width=12, foreground="red").pack(side=tk.LEFT, padx=(0, 5))
        ttk.Label(stats_row, textvariable=self.total_balance_var, width=18, foreground="purple").pack(side=tk.LEFT, padx=(0, 5))
        ttk.Label(stats_row, textvariable=self.total_checkin_reward_var, width=20, foreground="darkgreen").pack(side=tk.LEFT)
        
        # === 结果表格区域 ===
        results_frame = ttk.LabelFrame(main_frame, text="账号处理结果汇总", padding="5")
        results_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 0))
        
        # 添加操作按钮行
        button_row = ttk.Frame(results_frame)
        button_row.grid(row=0, column=0, sticky="ew", pady=(0, 5))
        
        ttk.Button(button_row, text="全选", command=self._select_all_results, width=8).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(button_row, text="反选", command=self._invert_selection, width=8).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(button_row, text="撤回", command=self._undo_selection, width=8).pack(side=tk.LEFT, padx=(0, 5))
        
        # 快速筛选按钮
        ttk.Button(button_row, text="🔍 执行失败", command=self._filter_failed, width=10).pack(side=tk.LEFT, padx=(10, 5))
        ttk.Button(button_row, text="💰 有余额", command=self._filter_has_balance, width=10).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(button_row, text="📭 无余额", command=self._filter_no_balance, width=10).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(button_row, text="🔄 显示全部", command=self._show_all, width=10).pack(side=tk.LEFT, padx=(0, 5))
        
        # 搜索框
        ttk.Label(button_row, text="搜索:", width=6).pack(side=tk.LEFT, padx=(10, 5))
        self.main_search_var = tk.StringVar()
        self.main_search_entry = ttk.Entry(button_row, textvariable=self.main_search_var, width=15)
        self.main_search_entry.pack(side=tk.LEFT, padx=(0, 5))
        self.main_search_entry.bind('<Return>', lambda e: self._search_main_table())
        ttk.Button(button_row, text="🔍 搜索", command=self._search_main_table, width=8).pack(side=tk.LEFT, padx=(0, 5))
        
        # 创建Treeview表格 (带勾选框)
        columns = (
            "phone", "nickname", "user_id", "balance_before", "points", "vouchers", "coupons",
            "checkin_reward", "checkin_total_times", 
            "balance_after", "transfer_amount", "transfer_recipient", "duration", "status", "login_method", "owner"
        )
        
        # show="tree headings" 显示勾选框列和数据列
        self.results_tree = ttk.Treeview(results_frame, columns=columns, show="tree headings", height=15)
        
        # 配置勾选框列(第一列) - 使用自定义绘制
        self.results_tree.heading("#0", text="", anchor=tk.CENTER)
        self.results_tree.column("#0", width=40, anchor=tk.CENTER, stretch=False, minwidth=40)
        
        # 创建勾选框图标(使用 PhotoImage)
        self._create_checkbox_images()
        
        # 定义列标题和宽度(手机号移到最前面，管理员列在最后)
        column_config = {
            "phone": ("手机号", 100),
            "nickname": ("昵称", 80),
            "user_id": ("ID", 80),
            "balance_before": ("余额前", 70),
            "points": ("积分", 60),
            "vouchers": ("抵扣券", 60),
            "coupons": ("优惠券", 60),
            "checkin_reward": ("签到奖励", 80),
            "checkin_total_times": ("签到次数", 70),
            "balance_after": ("余额", 70),
            "transfer_amount": ("转账金额", 80),
            "transfer_recipient": ("收款人ID", 80),
            "duration": ("耗时(秒)", 70),
            "status": ("状态", 60),
            "login_method": ("登录方式", 80),
            "owner": ("管理员", 80)
        }
        
        for col, (heading, width) in column_config.items():
            self.results_tree.heading(col, text=heading)
            self.results_tree.column(col, width=width, anchor=tk.CENTER)
        
        # 绑定点击事件(用于切换勾选状态)
        self.results_tree.bind("<Button-1>", self._on_tree_click)
        
        # 绑定双击事件(用于快速勾选/取消勾选)
        self.results_tree.bind("<Double-Button-1>", self._on_tree_double_click)
        
        # 添加滚动条
        results_scrollbar_y = ttk.Scrollbar(results_frame, orient=tk.VERTICAL, command=self.results_tree.yview)
        results_scrollbar_x = ttk.Scrollbar(results_frame, orient=tk.HORIZONTAL, command=self.results_tree.xview)
        self.results_tree.configure(yscrollcommand=results_scrollbar_y.set, xscrollcommand=results_scrollbar_x.set)
        
        # 布局
        self.results_tree.grid(row=1, column=0, sticky="nsew")
        results_scrollbar_y.grid(row=1, column=1, sticky="ns")
        results_scrollbar_x.grid(row=2, column=0, sticky="ew")
        
        results_frame.grid_rowconfigure(1, weight=1)
        results_frame.grid_columnconfigure(0, weight=1)
        
        # 配置标签颜色(用于高亮正值变化)
        self.results_tree.tag_configure("positive", foreground="green")
        self.results_tree.tag_configure("negative", foreground="purple")  # 转账成功余额减少用紫色
        self.results_tree.tag_configure("neutral", foreground="black")
        self.results_tree.tag_configure("checked", foreground="blue")  # 已完成的行用蓝色
        
        # 初始化勾选状态字典和历史记录
        self.checked_items = {}  # {item_id: True/False}
        self.selection_history = []  # 最多保存5个历史状态
        self.all_tree_items = []  # 存储所有表格项目ID（用于筛选恢复）
    
    def _create_checkbox_images(self):
        """创建勾选框图标(使用对称的Unicode字符)"""
        # 使用完全对称的字符组合
        # □ (U+25A1) 大方框 - 未选中
        # ☑ (U+2611) 带勾方框 - 选中
        # 为了确保对称，我们在未选中状态使用相同的字符但用空格填充
        self.checkbox_unchecked_text = "□"  # 大方框(空心)
        self.checkbox_checked_text = "☑"    # 带勾方框
        
        # 配置字体以确保对称显示
        try:
            # 尝试使用等宽字体
            import tkinter.font as tkfont
            checkbox_font = tkfont.Font(family="Segoe UI Symbol", size=12)
            self.results_tree.tag_configure("checkbox", font=checkbox_font)
        except:
            pass  # 如果字体配置失败，使用默认字体
    
    def _is_account_checked(self, phone: str) -> bool:
        """检查指定手机号的账号是否被选中(线程安全)
        
        Args:
            phone: 手机号
            
        Returns:
            True 如果账号被选中，False 如果未选中或不存在
        """
        # 使用锁保护，确保线程安全
        with self.stats_lock:
            # 遍历表格中的所有项目
            for item_id in self.results_tree.get_children():
                # 获取该行的数据
                values = self.results_tree.item(item_id, 'values')
                if values and len(values) > 0:
                    # 第一列是手机号
                    item_phone = values[0]
                    if item_phone == phone:
                        # 找到了对应的账号，返回其勾选状态
                        return self.checked_items.get(item_id, False)
            
            # 如果没找到，返回 False(未选中)
            return False
    
    def _acquire_instance(self) -> Optional[int]:
        """从实例池获取一个可用的模拟器实例编号（线程安全）
        
        Returns:
            实例编号，如果没有可用实例则返回 None
        """
        with self.instance_lock:
            if self.instance_pool:
                return self.instance_pool.pop(0)
            return None
    
    def _release_instance(self, instance_id: int):
        """释放模拟器实例编号回池中(线程安全)
        
        Args:
            instance_id: 要释放的实例编号
        """
        with self.instance_lock:
            if instance_id not in self.instance_pool:
                self.instance_pool.append(instance_id)
                self.instance_pool.sort()  # 保持有序
    
    def _load_config_to_ui(self):
        """加载配置到界面"""
        self.emulator_path_var.set(self.config.nox_path)
        self.accounts_file_var.set(self.config.accounts_file)
        self.instance_count_var.set(self.config.max_concurrent_instances)
        
        # 加载模拟器类型选择
        saved_selection = self.config.emulator_type_selection
        if saved_selection in ('自动检测', 'MuMu模拟器'):
            self.emulator_type_var.set(saved_selection)
        else:
            self.emulator_type_var.set("自动检测")
        
        # 加载自动转账开关状态
        from .transfer_config import get_transfer_config
        transfer_config = get_transfer_config()
        self.auto_transfer_switch.set_state(transfer_config.enabled)
        
        # 加载定时运行配置
        scheduled_enabled = getattr(self.config, 'scheduled_run_enabled', False)
        scheduled_time = getattr(self.config, 'scheduled_run_time', '08:00')
        self.scheduled_run_enabled.set(scheduled_enabled)
        self.scheduled_run_time.set(scheduled_time)
        
        # 解析时间并设置到小时和分钟变量
        try:
            hour, minute = scheduled_time.split(':')
            self.scheduled_hour_var.set(f"{int(hour):02d}")
            self.scheduled_minute_var.set(f"{int(minute):02d}")
        except:
            self.scheduled_hour_var.set("08")
            self.scheduled_minute_var.set("00")
        
        self.scheduled_run_switch.set_state(scheduled_enabled)
        
        # 如果启用了定时运行，启动定时检查线程
        if scheduled_enabled:
            self._start_schedule_check_thread()
        
        # 自动加载账号列表(如果配置了账号文件)
        self._auto_load_accounts()
    
    def _browse_emulator_path(self):
        """选择盘符并自动搜索模拟器"""
        # 让用户选择盘符
        drive_path = filedialog.askdirectory(
            title="选择 MuMu 模拟器安装的盘符（如 C:\\ 或 D:\\）",
            initialdir="C:\\"
        )
        
        if not drive_path:
            return
        
        # 提取盘符
        drive_letter = os.path.splitdrive(drive_path)[0]
        if not drive_letter:
            self._log("❌ 无效的盘符")
            return
        
        self._log(f"正在 {drive_letter} 中搜索 MuMu 模拟器，请稍候...")
        self._log("提示：搜索可能需要 1-3 分钟，请耐心等待")
        
        # 在后台线程中执行搜索
        def search_task():
            try:
                found_paths = EmulatorController.search_in_drive(drive_letter)
                
                # 在主线程中更新 UI
                self.root.after(0, lambda: self._handle_search_results(found_paths, drive_letter))
            except Exception as e:
                self.root.after(0, lambda: self._log(f"❌ 搜索失败: {e}"))
        
        # 启动搜索线程
        search_thread = threading.Thread(target=search_task, daemon=True)
        search_thread.start()
    
    def _select_drive_and_search(self):
        """选择盘符并自动搜索模拟器（新方法名）"""
        self._browse_emulator_path()
    
    def _handle_search_results(self, found_paths: List[str], drive_letter: str):
        """处理搜索结果"""
        if not found_paths:
            self._log(f"❌ 在 {drive_letter} 中未找到 MuMu 模拟器")
            self._log("提示：请确认 MuMu 模拟器已安装在该盘符，或尝试其他盘符")
            return
        
        if len(found_paths) == 1:
            # 只找到一个，直接使用
            path = found_paths[0]
            self.emulator_path_var.set(path)
            self._update_emulator_type(path)
            self._log(f"✅ 找到 MuMu 模拟器: {path}")
        else:
            # 找到多个，显示所有路径
            self._log(f"✅ 找到 {len(found_paths)} 个 MuMu 模拟器安装位置:")
            for i, path in enumerate(found_paths, 1):
                self._log(f"  {i}. {path}")
            
            # 使用第一个作为默认值
            path = found_paths[0]
            self.emulator_path_var.set(path)
            self._update_emulator_type(path)
            self._log(f"已自动选择: {path}")
            self._log("如需使用其他路径，请重新选择盘符搜索")
    
    def _reload_accounts(self):
        """重新加载账号列表"""
        # 先重新加载配置（确保账号文件路径是最新的）
        self.config = self.config_loader.load()
        self.accounts_file_var.set(self.config.accounts_file)
        
        # 确认是否重新加载(会清空当前结果)
        if self.results_tree.get_children():
            result = messagebox.askyesno(
                "确认", 
                "重新加载将清空当前表格中的所有数据(包括处理结果)\n\n是否继续？"
            )
            if not result:
                return
        
        self._auto_load_accounts()
    
    def _silent_reload_accounts(self):
        """静默重新加载账号列表（不弹确认框，用于删除账号后自动刷新）"""
        # 先重新加载配置（确保账号文件路径是最新的）
        self.config = self.config_loader.load()
        self.accounts_file_var.set(self.config.accounts_file)
        
        # 直接加载，不弹确认框
        self._auto_load_accounts()
    
    def _auto_load_accounts(self):
        """自动加载账号列表到表格(支持从历史结果恢复数据)
        
        策略：
        1. 加载历史记录中的所有账号（永久保存）
        2. 加载账号文件中的账号
        3. 合并两者，账号文件中的账号优先（可以更新历史数据）
        4. 即使账号从文件中删除，历史记录仍然保留
        """
        try:
            accounts_file = self.accounts_file_var.get()
            
            # ===== 步骤1: 加载历史记录（永久数据源）=====
            history_data = self._load_from_history()
            history_dict = {}
            if history_data:
                # 转换为字典格式,以手机号为key
                history_dict = {record['手机号']: record for record in history_data}
            
            # ===== 步骤2: 加载账号文件（当前活跃账号）=====
            current_accounts = []
            current_phones = set()
            if accounts_file:
                # 检查明文文件或加密文件是否存在
                plain_file_exists = Path(accounts_file).exists()
                enc_file_exists = Path(f"{accounts_file}.enc").exists()
                
                if plain_file_exists or enc_file_exists:
                    try:
                        account_manager = AccountManager(accounts_file)
                        current_accounts = account_manager.load_accounts()
                        current_phones = {acc.phone for acc in current_accounts}
                    except Exception as e:
                        self._log(f"❌ 加载账号文件失败: {e}")
                        import traceback
                        traceback.print_exc()
            
            # ===== 步骤3: 只显示账号文件中的账号（不显示历史记录中的其他账号）=====
            if not current_accounts:
                return
            
            # ===== 步骤4: 清空表格并重新填充 =====
            
            # 清空现有表格
            all_items = self.results_tree.get_children()
            for item in all_items:
                self.results_tree.delete(item)
            
            # 清空勾选状态和历史记录
            self.checked_items.clear()
            self.selection_history.clear()
            
            # 清空筛选缓存
            self.all_tree_items = []
            
            # 统计变量
            success_count = 0
            restored_count = 0  # 从历史恢复的账号数
            total_balance = 0.0
            total_points = 0
            total_vouchers = 0.0
            total_coupons = 0
            total_checkin_reward = 0.0
            
            # ===== 步骤5: 填充表格（只显示账号文件中的账号）=====
            # 从数据库批量加载所有账号的管理员信息（一次查询，避免循环查询）
            from .local_db import LocalDatabase
            db = LocalDatabase()
            
            # 批量获取所有账号的最新记录（包含管理员信息）
            owner_map = {}  # {phone: owner_name}
            try:
                all_phones = [acc.phone for acc in current_accounts]
                # 使用一次查询获取所有账号的最新记录
                conn = db._get_connection()
                cursor = conn.cursor()
                placeholders = ','.join(['?' for _ in all_phones])
                query = f"""
                    SELECT phone, owner 
                    FROM history_records 
                    WHERE phone IN ({placeholders})
                    AND id IN (
                        SELECT MAX(id) 
                        FROM history_records 
                        WHERE phone IN ({placeholders})
                        GROUP BY phone
                    )
                """
                cursor.execute(query, all_phones + all_phones)
                rows = cursor.fetchall()
                for phone, owner in rows:
                    if owner:
                        owner_map[phone] = owner
                conn.close()
            except Exception as e:
                self._log(f"⚠️ 批量加载管理员信息失败: {e}")
            
            for account in current_accounts:
                phone = account.phone
                hist = None
                
                # 从批量查询结果中获取管理员
                owner_name = owner_map.get(phone, "-")
                
                # 尝试从历史记录中匹配数据（按手机号）
                if phone in history_dict:
                    hist = history_dict[phone]
                    restored_count += 1
                
                if hist:
                    # 有历史数据，使用历史数据
                    # 处理空值：None 显示为 0 或 -
                    def format_value(value, default='-', is_number=False):
                        """格式化显示值，None显示为默认值"""
                        # 处理 None 和空字符串
                        if value is None or value == '' or str(value).lower() == 'none':
                            return '0' if is_number else default
                        # 如果是数值类型，格式化显示
                        if is_number and isinstance(value, (int, float)):
                            if isinstance(value, float):
                                return f"{value:.2f}" if value != int(value) else str(int(value))
                            return str(value)
                        # 返回字符串，但要确保不是 "None"
                        str_value = str(value)
                        if str_value.lower() == 'none':
                            return '0' if is_number else default
                        return str_value
                    
                    values = (
                        phone,
                        format_value(hist.get('昵称'), '待处理'),
                        format_value(hist.get('用户ID'), '待处理'),
                        format_value(hist.get('余额前(元)'), '0.0', True),
                        format_value(hist.get('积分'), '0', True),
                        format_value(hist.get('抵扣券(张)'), '0', True),
                        format_value(hist.get('优惠券(张)'), '0', True),
                        format_value(hist.get('签到奖励(元)'), '0.0', True),
                        format_value(hist.get('签到总次数'), '0', True),
                        format_value(hist.get('余额(元)'), '0.0', True),
                        format_value(hist.get('转账金额(元)'), '0.0', True),
                        format_value(hist.get('转账收款人'), '-'),
                        format_value(hist.get('耗时(秒)'), '-'),
                        format_value(hist.get('状态'), '待处理'),
                        format_value(hist.get('登录方式'), '-'),
                        owner_name  # 管理员（从批量查询结果中获取，放在最后）
                    )
                    
                    # 累积统计(只统计成功的)
                    if hist.get('状态') == '成功':
                        success_count += 1
                        try:
                            if hist.get('余额(元)') and hist.get('余额(元)') != '-':
                                total_balance += float(hist.get('余额(元)', 0))
                            if hist.get('积分') and hist.get('积分') != '-':
                                total_points += int(hist.get('积分', 0))
                            if hist.get('抵扣券(张)') and hist.get('抵扣券(张)') != '-':
                                total_vouchers += float(hist.get('抵扣券(张)', 0))
                            if hist.get('优惠券(张)') and hist.get('优惠券(张)') != '-':
                                total_coupons += int(hist.get('优惠券(张)', 0))
                            if hist.get('签到奖励(元)') and hist.get('签到奖励(元)') != '-':
                                total_checkin_reward += float(hist.get('签到奖励(元)', 0))
                        except (ValueError, TypeError):
                            pass
                else:
                    # 没有历史数据，显示为待处理
                    values = (
                        phone,
                        "待处理",
                        "待处理",
                        "-", "-", "-", "-", "-", "-", "-", "-", "-", "-",
                        "待处理",
                        "-",
                        owner_name  # 管理员（从数据库读取，放在最后）
                    )
                
                # 插入行，默认不勾选
                item_id = self.results_tree.insert("", tk.END, text=self.checkbox_unchecked_text, values=values)
                self.checked_items[item_id] = False
            
            # ===== 步骤6: 更新统计信息 =====
            total = len(current_accounts)
            # 启动时，所有账号都是待处理状态，不应该显示历史失败数
            # 只有在实际运行后才更新成功/失败统计
            pending_count = total - success_count
            
            # 从表格统计数据（这样会统计所有账号，包括历史数据）
            self._update_stats_from_table()
            
            # 更新账号统计信息
            self.account_total_var.set(f"账号总计: {total}")
            self.account_pending_var.set(f"待处理: {pending_count}")
            
            # ===== 步骤7: 显示加载总结 =====
            # 只在有账号时显示简洁的加载信息
            if total > 0:
                self._log(f"✅ 已加载 {total} 个账号")
            
            # ===== 步骤8: 恢复勾选状态 =====
            saved_selections = self.selection_manager.load_selections()
            if saved_selections:
                restored_selection_count = 0
                for item_id in self.results_tree.get_children():
                    values = self.results_tree.item(item_id, 'values')
                    if values and len(values) > 0:
                        phone = values[0]
                        # 如果配置文件中有该账号的状态，恢复它
                        if phone in saved_selections:
                            checked = saved_selections[phone]
                            self.checked_items[item_id] = checked
                            # 更新显示
                            if checked:
                                self.results_tree.item(item_id, text=self.checkbox_checked_text)
                                restored_selection_count += 1
                            else:
                                self.results_tree.item(item_id, text=self.checkbox_unchecked_text)
                
                if restored_selection_count > 0:
                    self._log(f"✅ 已恢复 {restored_selection_count} 个账号的勾选状态")
            
            # ===== 步骤9: 更新筛选缓存 =====
            # 保存所有项目ID，用于筛选功能
            self.all_tree_items = list(self.results_tree.get_children())
        
        except Exception as e:
            import traceback
            self._log(f"加载账号失败: {e}")
            self._log(f"详细错误: {traceback.format_exc()}")
    
    def _load_history_data(self):
        """从数据库加载历史数据
        
        Returns:
            dict: {手机号: {字段: 值}} 或 None
        """
        try:
            db = LocalDatabase()
            
            records = db.get_all_history_records()
            if not records:
                return None
            
            history_data = {}
            for record in records:
                phone = record.get('phone')
                if phone:
                    history_data[phone] = record
            
            return history_data if history_data else None
            
        except Exception as e:
            print(f"加载历史记录失败: {e}")
            return None
    
    def _load_license_info_async(self):
        """异步加载许可证信息（避免阻塞GUI）"""
        def load_in_thread():
            try:
                # 优先使用简化版许可证管理器
                try:
                    from .license_manager_simple import SimpleLicenseManager
                    license_manager = SimpleLicenseManager()
                    self.license_info = license_manager.get_license_info()
                except ImportError:
                    # 如果简化版不可用，使用完整版
                    from .license_manager import LicenseManager
                    license_manager = LicenseManager()
                    self.license_info = license_manager.get_license_info()
                
                # 在主线程更新UI
                self.root.after(0, self._update_license_display)
            except Exception as e:
                print(f"加载许可证信息失败: {e}")
                self.license_info = None
        
        # 在后台线程加载
        thread = threading.Thread(target=load_in_thread, daemon=True)
        thread.start()
    
    def _update_license_display(self):
        """更新许可证显示"""
        if self.license_info:
            days_left = self.license_info.get('days_left', 0)
            expires_at = self.license_info.get('expires_at', '')
            
            # 检查是否是永久授权
            if days_left >= 999999 or expires_at == '9999-12-31T23:59:59':
                self.license_label.config(text="许可证状态: 永久授权（本机）", fg='green')
            # 格式化到期日期
            elif expires_at:
                try:
                    from datetime import datetime
                    expires_date = datetime.fromisoformat(expires_at.replace('Z', '+00:00'))
                    date_str = expires_date.strftime('%Y-%m-%d')
                    
                    if days_left > 0:
                        self.license_label.config(text=f"许可证到期: {date_str} (剩余 {days_left} 天)")
                    elif days_left == 0:
                        self.license_label.config(text=f"许可证到期: {date_str} (今天到期)")
                    else:
                        self.license_label.config(text=f"许可证已过期: {date_str} (已过期 {abs(days_left)} 天)")
                except:
                    if days_left > 0:
                        self.license_label.config(text=f"许可证剩余: {days_left} 天")
                    else:
                        self.license_label.config(text=f"许可证已过期")
            else:
                # 没有到期时间信息
                if days_left > 0:
                    self.license_label.config(text=f"许可证剩余: {days_left} 天")
                else:
                    self.license_label.config(text="许可证状态未知")
        else:
            self.license_label.config(text="")
    
    def _auto_detect_emulator(self):
        """自动检测模拟器"""
        found = EmulatorController.detect_all_emulators()
        
        if found:
            emulator_type, path = found[0]
            self.emulator_path_var.set(path)
            self._update_emulator_type(path)
            
            type_names = {
                EmulatorType.MUMU: "MuMu模拟器"
            }
            type_name = type_names.get(emulator_type, "未知模拟器")
            self._log(f"自动检测到 {type_name}: {path}")
            
            if len(found) > 1:
                self._log(f"还检测到其他模拟器: {[f[1] for f in found[1:]]}")
        else:
            self._log("未能自动检测到模拟器，请手动选择路径")
            self.emulator_type_var.set("未检测到")
    
    def _update_emulator_type(self, path: str):
        """更新模拟器类型显示"""
        controller = EmulatorController(path)
        if controller.emulator_type == EmulatorType.MUMU:
            self.emulator_type_var.set("MuMu模拟器")
            self.emulator_status_var.set("✅ MuMu")
        else:
            self.emulator_type_var.set("未知类型")
            self.emulator_status_var.set("❌ 未知")
        
        self.emulator_controller = controller
    
    def _detect_running_instances(self, auto_adjust=False):
        """检测运行中的模拟器实例
        
        Args:
            auto_adjust: 是否自动调整并行数
        """
        # 在后台线程中执行检测
        def detect_task():
            try:
                # 检查是否已配置模拟器
                emulator_path = self.emulator_path_var.get()
                if not emulator_path:
                    self.root.after(0, lambda: self._log("❌ 请先配置模拟器路径"))
                    self.root.after(0, lambda: self.running_instances_var.set("未配置模拟器"))
                    return
                
                # 创建控制器
                controller = EmulatorController(emulator_path)
                
                # 简洁日志：正在检测
                self.root.after(0, lambda: self._log("正在检测运行中的实例..."))
                self.root.after(0, lambda: self.running_instances_var.set("检测中..."))
                
                # 异步检测实例
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                running_instances = loop.run_until_complete(controller.get_running_instances())
                loop.close()
                
                if not running_instances:
                    self.root.after(0, lambda: self._log("❌ 未检测到运行中的实例"))
                    self.root.after(0, lambda: self.running_instances_var.set("未检测到"))
                else:
                    count = len(running_instances)
                    instances_str = ", ".join([f"#{i}" for i in running_instances])
                    
                    # 获取当前设置的并发实例数
                    concurrent_count = self.instance_count_var.get()
                    
                    # 简洁日志：检测结果
                    self.root.after(0, lambda: self._log(f"✅ 检测到 {count} 个运行中的实例"))
                    
                    # 自动调整并行数
                    if auto_adjust:
                        self.root.after(0, lambda c=count: self.instance_count_var.set(c))
                        self.root.after(0, lambda c=count: self._log(f"   已自动调整并行数为: {c}"))
                    
                    # 更新显示：显示并发实例数/总实例数
                    final_concurrent = count if auto_adjust else concurrent_count
                    self.root.after(0, lambda c=final_concurrent, t=count: 
                                  self.running_instances_var.set(f"✅ 并发{c}/{t}个实例"))
                
            except Exception as e:
                self.root.after(0, lambda: self._log(f"❌ 检测失败: {e}"))
                self.root.after(0, lambda: self.running_instances_var.set("检测失败"))
        
        # 启动后台线程
        thread = threading.Thread(target=detect_task, daemon=True)
        thread.start()
    
    def _on_emulator_type_changed(self, event=None):
        """模拟器类型选择变化时的处理"""
        selected = self.emulator_type_var.get()
        if selected == "自动检测":
            self.emulator_status_var.set("请点击'应用选择'")
        else:
            self.emulator_status_var.set(f"已选择: {selected}")
    
    def _apply_emulator_selection(self):
        """应用模拟器选择"""
        selected = self.emulator_type_var.get()
        
        if selected == "自动检测":
            # 自动检测所有模拟器
            self._auto_detect_emulator()
            # 保存选择
            self._save_config()
            return
        
        # 根据选择查找对应的模拟器
        self._log(f"正在查找 {selected}...")
        found_emulators = EmulatorController.detect_all_emulators()
        
        if not found_emulators:
            error_msg = (
                f"未检测到任何模拟器\n\n"
                f"请确保已安装MuMu模拟器\n\n"
                f"如果已安装,请尝试:\n"
                f"  1. 手动点击'浏览'选择模拟器路径\n"
                f"  2. 确认模拟器已正确安装"
            )
            messagebox.showerror("错误", error_msg)
            self.emulator_status_var.set("❌ 未找到")
            self._log("❌ 未检测到任何模拟器")
            return
        
        # 类型映射
        type_map = {
            'MuMu模拟器': EmulatorType.MUMU
        }
        
        target_type = type_map.get(selected)
        
        # 查找匹配的模拟器
        found_path = None
        for emu_type, emu_path in found_emulators:
            if emu_type == target_type:
                found_path = emu_path
                break
        
        if found_path:
            # 找到了对应的模拟器
            self.emulator_path_var.set(found_path)
            self._update_emulator_type(found_path)
            self._log(f"✅ 找到 {selected}: {found_path}")
            
            # 验证路径是否有效
            controller = EmulatorController(found_path)
            if not controller.is_available():
                error_msg = (
                    f"模拟器路径无效\n\n"
                    f"路径: {found_path}\n\n"
                    f"可能的原因:\n"
                    f"  • 模拟器未正确安装\n"
                    f"  • 缺少必要的可执行文件\n\n"
                    f"请尝试重新安装模拟器"
                )
                messagebox.showerror("错误", error_msg)
                self.emulator_status_var.set("❌ 路径无效")
                self._log(f"❌ 模拟器路径无效: {found_path}")
                return
            
            # 自动保存配置
            self._save_config()
            
            # 显示成功信息
            exe_name = "未知"
            if controller.emulator_type == EmulatorType.MUMU:
                exe_name = "MuMuNxMain.exe"
            
            messagebox.showinfo("成功", 
                              f"✅ 已成功配置 {selected}\n\n"
                              f"路径: {found_path}\n"
                              f"可执行文件: {exe_name}\n\n"
                              f"配置已自动保存")
        else:
            # 未找到对应的模拟器
            type_names = {
                EmulatorType.MUMU: "MuMu模拟器"
            }
            available = [type_names.get(t, "未知") for t, _ in found_emulators]
            available_paths = [f"  • {type_names.get(t, '未知')}: {p}" for t, p in found_emulators]
            
            error_msg = (
                f"未找到 {selected}\n\n"
                f"检测到的模拟器:\n" + "\n".join(available_paths) + "\n\n"
                f"建议:\n"
                f"  1. 选择已检测到的模拟器类型\n"
                f"  2. 或安装 {selected}"
            )
            messagebox.showerror("错误", error_msg)
            self.emulator_status_var.set(f"❌ 未找到{selected}")
            self._log(f"❌ 未找到 {selected}")
            self._log(f"   可用的模拟器: {', '.join(available)}")

    def _save_config(self):
        """保存配置"""
        self.config.nox_path = self.emulator_path_var.get()
        self.config.accounts_file = self.accounts_file_var.get()
        self.config.max_concurrent_instances = self.instance_count_var.get()
        
        # 保存模拟器类型选择(新增)
        self.config.emulator_type_selection = self.emulator_type_var.get()
        
        # 保存定时运行配置
        self.config.scheduled_run_enabled = self.scheduled_run_enabled.get()
        self.config.scheduled_run_time = self.scheduled_run_time.get()
        
        self.config_loader.save(self.config)
        self._log("配置已保存")
    
    def _auto_save_config(self):
        """自动保存配置（静默，不显示日志）"""
        try:
            self.config.nox_path = self.emulator_path_var.get()
            self.config.accounts_file = self.accounts_file_var.get()
            self.config.max_concurrent_instances = self.instance_count_var.get()
            self.config.emulator_type_selection = self.emulator_type_var.get()
            
            # 保存定时运行配置
            self.config.scheduled_run_enabled = self.scheduled_run_enabled.get()
            self.config.scheduled_run_time = self.scheduled_run_time.get()
            
            self.config_loader.save(self.config)
        except Exception as e:
            # 静默失败，避免干扰用户
            print(f"自动保存配置失败: {e}")
    
    def _log(self, message: str):
        """添加日志(支持实例过滤)"""
        import datetime
        timestamp = datetime.datetime.now().strftime("%H:%M:%S")
        log_entry = f"[{timestamp}] {message}"
        
        # 存储到所有日志列表
        self.all_logs.append(log_entry)
        
        # 根据当前过滤器决定是否显示
        current_filter = self.log_filter_var.get()
        should_display = False
        
        if current_filter == "全部":
            should_display = True
        elif current_filter.startswith("实例"):
            # 提取实例编号(例如"实例0" -> "0")
            instance_num = current_filter.replace("实例", "")
            # 检查消息是否包含该实例的标记
            if f"[实例{instance_num}]" in message:
                should_display = True
        
        # 只显示符合过滤条件的日志
        if should_display:
            self.log_text.config(state=tk.NORMAL)
            self.log_text.insert(tk.END, f"{log_entry}\n")
            
            # 只有在自动滚动开启时才滚动到底部
            if self.log_auto_scroll:
                self.log_text.see(tk.END)
            
            self.log_text.config(state=tk.DISABLED)
    
    def _log_error(self, phone: str, user_id: str, nickname: str, error_message: str):
        """添加错误日志（每个账号一行）
        
        Args:
            phone: 手机号
            user_id: 用户ID
            nickname: 昵称
            error_message: 错误内容
        """
        import datetime
        timestamp = datetime.datetime.now().strftime("%H:%M:%S")
        
        # 格式：[时间] 手机号 | ID | 昵称 | 错误内容
        error_entry = f"[{timestamp}] {phone} | {user_id} | {nickname} | {error_message}"
        
        # 存储到错误日志列表（新格式）
        self.all_error_logs.append({
            'phone': phone,
            'type': 'error',
            'message': error_entry,
            'timestamp': timestamp
        })
        
        # 显示错误日志
        self.error_log_text.config(state=tk.NORMAL)
        self.error_log_text.insert(tk.END, f"{error_entry}\n", "error")
        self.error_log_text.see(tk.END)
        self.error_log_text.config(state=tk.DISABLED)
    
    def _on_log_filter_changed(self, event=None):
        """日志过滤器改变时重新显示日志"""
        current_filter = self.log_filter_var.get()
        
        # 清空当前显示
        self.log_text.config(state=tk.NORMAL)
        self.log_text.delete(1.0, tk.END)
        
        # 根据过滤器重新显示日志
        for log_entry in self.all_logs:
            should_display = False
            
            if current_filter == "全部":
                should_display = True
            elif current_filter.startswith("实例"):
                # 提取实例编号
                instance_num = current_filter.replace("实例", "")
                # 检查消息是否包含该实例的标记
                if f"[实例{instance_num}]" in log_entry:
                    should_display = True
            
            if should_display:
                self.log_text.insert(tk.END, f"{log_entry}\n")
        
        # 滚动到底部
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)
        
        # 重新启用自动滚动
        self.log_auto_scroll = True
    
    def _on_log_scroll(self, event):
        """用户滚动日志时停止自动滚动"""
        self.log_auto_scroll = False
        return None  # 允许事件继续传播
    
    def _on_log_scrollbar_drag(self, event):
        """用户拖动滚动条时停止自动滚动"""
        self.log_auto_scroll = False
        return None
    
    def _on_tree_click(self, event):
        """处理表格点击事件(切换勾选状态)"""
        region = self.results_tree.identify("region", event.x, event.y)
        if region == "tree":  # 点击了勾选框列
            item = self.results_tree.identify_row(event.y)
            if item:
                self._toggle_check(item)
    
    def _on_tree_double_click(self, event):
        """处理表格双击事件(快速勾选/取消勾选)
        
        双击任意列都可以切换勾选状态，不限于勾选框列
        """
        item = self.results_tree.identify_row(event.y)
        if item:
            self._toggle_check(item)
    
    def _toggle_check(self, item):
        """切换单个项目的勾选状态"""
        # 保存当前状态到历史
        self._save_selection_state()
        
        # 切换状态
        current_state = self.checked_items.get(item, False)
        new_state = not current_state
        self.checked_items[item] = new_state
        
        # 更新显示
        if new_state:
            self.results_tree.item(item, text=self.checkbox_checked_text)
        else:
            self.results_tree.item(item, text=self.checkbox_unchecked_text)
        
        # 勾选状态变化后，延迟保存（防抖）
        if hasattr(self, '_save_timer'):
            self.root.after_cancel(self._save_timer)
        
        # 300毫秒后保存
        self._save_timer = self.root.after(300, self._save_selections_to_file)
    
    def _select_all_results(self):
        """全选或取消全选（只操作当前显示的账户）"""
        # 保存当前状态到历史
        self._save_selection_state()
        
        # 获取当前显示的项目（不包括被筛选隐藏的）
        all_items = self.results_tree.get_children()
        if not all_items:
            return
        
        all_checked = all(self.checked_items.get(item, False) for item in all_items)
        
        # 如果全部已选中，则取消全选；否则全选
        new_state = not all_checked
        
        for item in all_items:
            self.checked_items[item] = new_state
            if new_state:
                self.results_tree.item(item, text=self.checkbox_checked_text)
            else:
                self.results_tree.item(item, text=self.checkbox_unchecked_text)
        
        # 记录日志
        action = "全选" if new_state else "取消全选"
        self._log(f"✓ 已{action} {len(all_items)} 个显示的账户")
        
        # 勾选状态变化后，延迟保存（防抖）
        if hasattr(self, '_save_timer'):
            self.root.after_cancel(self._save_timer)
        
        # 300毫秒后保存
        self._save_timer = self.root.after(300, self._save_selections_to_file)
    
    def _invert_selection(self):
        """反选：选中未选中的，取消选中已选中的（只操作当前显示的账户）"""
        # 保存当前状态到历史
        self._save_selection_state()
        
        # 获取当前显示的项目（不包括被筛选隐藏的）
        all_items = self.results_tree.get_children()
        
        checked_count = 0
        unchecked_count = 0
        
        for item in all_items:
            current_state = self.checked_items.get(item, False)
            new_state = not current_state
            self.checked_items[item] = new_state
            
            if new_state:
                self.results_tree.item(item, text=self.checkbox_checked_text)
                checked_count += 1
            else:
                self.results_tree.item(item, text=self.checkbox_unchecked_text)
                unchecked_count += 1
        
        # 记录日志
        self._log(f"✓ 已反选 {len(all_items)} 个显示的账户（勾选 {checked_count} 个，取消 {unchecked_count} 个）")
        
        # 勾选状态变化后，延迟保存（防抖）
        if hasattr(self, '_save_timer'):
            self.root.after_cancel(self._save_timer)
        
        # 300毫秒后保存
        self._save_timer = self.root.after(300, self._save_selections_to_file)
    
    def _save_selection_state(self):
        """保存当前选择状态到历史(最多5个)"""
        # 复制当前状态
        current_state = self.checked_items.copy()
        
        # 添加到历史
        self.selection_history.append(current_state)
        
        # 只保留最近5个
        if len(self.selection_history) > 5:
            self.selection_history.pop(0)
    
    def _save_selections_to_file(self):
        """保存勾选状态到文件"""
        # 构建手机号到勾选状态的映射
        selections = {}
        for item_id in self.results_tree.get_children():
            values = self.results_tree.item(item_id, 'values')
            if values and len(values) > 0:
                phone = values[0]  # 第一列是手机号
                checked = self.checked_items.get(item_id, False)
                selections[phone] = checked
        
        # 保存到文件
        self.selection_manager.save_selections(selections)
    
    def _undo_selection(self):
        """撤回到上一个选择状态"""
        if not self.selection_history:
            self._log("没有可撤回的历史记录")
            return
        
        # 恢复上一个状态
        previous_state = self.selection_history.pop()
        self.checked_items = previous_state
        
        # 更新所有项目的显示
        all_items = self.results_tree.get_children()
        for item in all_items:
            is_checked = self.checked_items.get(item, False)
            if is_checked:
                self.results_tree.item(item, text=self.checkbox_checked_text)
            else:
                self.results_tree.item(item, text=self.checkbox_unchecked_text)
        
        self._log(f"已撤回到上一个状态(剩余{len(self.selection_history)}个历史记录)")
    
    
    def _filter_failed(self):
        """筛选执行失败的账户（只显示失败的账户）"""
        # 保存所有项目ID（如果还没保存）
        if not self.all_tree_items:
            self.all_tree_items = list(self.results_tree.get_children())
        
        if not self.all_tree_items:
            self._log("⚠️ 表格中没有数据")
            messagebox.showinfo("提示", "表格中没有数据")
            return
        
        # 先detach所有项目
        for item in self.all_tree_items:
            self.results_tree.detach(item)
        
        # 只reattach失败的账户
        failed_count = 0
        for item in self.all_tree_items:
            values = self.results_tree.item(item, 'values')
            if values and len(values) > 13:  # 确保有足够的列
                status = values[13]  # 状态列是第14列（索引13）
                if '失败' in str(status):
                    self.results_tree.reattach(item, '', 'end')
                    failed_count += 1
        
        if failed_count > 0:
            self._log(f"✓ 已筛选出 {failed_count} 个执行失败的账户（可勾选操作）")
        else:
            self._log("✓ 没有找到失败的账户")
            messagebox.showinfo("提示", "没有找到失败的账户")
    
    def _filter_has_balance(self):
        """筛选有余额的账户（余额不为0）"""
        # 保存所有项目ID（如果还没保存）
        if not self.all_tree_items:
            self.all_tree_items = list(self.results_tree.get_children())
        
        if not self.all_tree_items:
            self._log("⚠️ 表格中没有数据")
            messagebox.showinfo("提示", "表格中没有数据")
            return
        
        # 先detach所有项目
        for item in self.all_tree_items:
            self.results_tree.detach(item)
        
        # 只reattach有余额的账户（余额 > 0）
        has_balance_count = 0
        for item in self.all_tree_items:
            values = self.results_tree.item(item, 'values')
            if values and len(values) > 9:  # 确保有足够的列
                balance_after = values[9]  # 余额列是第10列（索引9）
                try:
                    balance = float(balance_after) if balance_after and balance_after != 'N/A' else 0.0
                    if balance > 0:
                        self.results_tree.reattach(item, '', 'end')
                        has_balance_count += 1
                except:
                    pass
        
        if has_balance_count > 0:
            self._log(f"✓ 已筛选出 {has_balance_count} 个有余额的账户（可勾选操作）")
        else:
            self._log("✓ 没有找到有余额的账户")
            messagebox.showinfo("提示", "没有找到有余额的账户")
    
    def _filter_no_balance(self):
        """筛选无余额的账户（余额为0）"""
        # 保存所有项目ID（如果还没保存）
        if not self.all_tree_items:
            self.all_tree_items = list(self.results_tree.get_children())
        
        if not self.all_tree_items:
            self._log("⚠️ 表格中没有数据")
            messagebox.showinfo("提示", "表格中没有数据")
            return
        
        # 先detach所有项目
        for item in self.all_tree_items:
            self.results_tree.detach(item)
        
        # 只reattach无余额的账户（余额 == 0）
        no_balance_count = 0
        for item in self.all_tree_items:
            values = self.results_tree.item(item, 'values')
            if values and len(values) > 9:  # 确保有足够的列
                balance_after = values[9]  # 余额列是第10列（索引9）
                try:
                    balance = float(balance_after) if balance_after and balance_after != 'N/A' else 0.0
                    if balance == 0:
                        self.results_tree.reattach(item, '', 'end')
                        no_balance_count += 1
                except:
                    pass
        
        if no_balance_count > 0:
            self._log(f"✓ 已筛选出 {no_balance_count} 个无余额的账户（可勾选操作）")
        else:
            self._log("✓ 没有找到无余额的账户")
            messagebox.showinfo("提示", "没有找到无余额的账户")
    
    def _show_all(self):
        """显示全部账户（清除筛选）"""
        # 如果有保存的项目，恢复所有项目
        if self.all_tree_items:
            # 先detach所有
            for item in self.all_tree_items:
                try:
                    self.results_tree.detach(item)
                except:
                    pass
            
            # 重新attach所有项目
            for item in self.all_tree_items:
                try:
                    self.results_tree.reattach(item, '', 'end')
                except:
                    pass
            
            self._log(f"✓ 已显示全部账户（共 {len(self.all_tree_items)} 个）")
        else:
            self._log("✓ 已显示全部账户")
    
    def _search_main_table(self):
        """搜索主界面表格（根据手机号或ID）"""
        search_text = self.main_search_var.get().strip()
        
        if not search_text:
            # 如果搜索框为空，显示全部
            self._show_all()
            return
        
        # 先显示全部（清除之前的筛选）
        if self.all_tree_items:
            for item in self.all_tree_items:
                try:
                    self.results_tree.detach(item)
                except:
                    pass
        
        # 搜索匹配的项目
        matched_items = []
        for item in self.all_tree_items:
            try:
                values = self.results_tree.item(item, 'values')
                if values and len(values) > 2:
                    phone = str(values[0])  # 手机号在第一列
                    user_id = str(values[2])  # 用户ID在第三列
                    
                    # 模糊匹配：手机号或ID包含搜索文本
                    if search_text in phone or search_text in user_id:
                        matched_items.append(item)
            except:
                pass
        
        # 显示匹配的项目
        for item in matched_items:
            try:
                self.results_tree.reattach(item, '', 'end')
            except:
                pass
        
        # 高亮显示匹配的行（使用蓝色）
        for item in matched_items:
            try:
                # 获取当前标签
                current_tags = list(self.results_tree.item(item, 'tags'))
                # 添加checked标签（蓝色）
                if 'checked' not in current_tags:
                    current_tags.append('checked')
                    self.results_tree.item(item, tags=current_tags)
            except:
                pass
        
        if matched_items:
            self._log(f"🔍 找到 {len(matched_items)} 个匹配的账户")
        else:
            self._log(f"🔍 未找到匹配 '{search_text}' 的账户")
            messagebox.showinfo("提示", f"未找到匹配 '{search_text}' 的账户")
    
    
    def _refresh_account_list(self):
        """刷新账号列表（重新加载以显示更新后的管理员）"""
        self._auto_load_accounts()
    
    def _clear_log(self):
        """清空日志"""
        self.log_text.config(state=tk.NORMAL)
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state=tk.DISABLED)
        # 清空所有日志存储
        self.all_logs = []
        # 清空日志后恢复自动滚动
        self.log_auto_scroll = True
    
    def _clear_error_log(self):
        """清空错误日志"""
        self.error_log_text.config(state=tk.NORMAL)
        self.error_log_text.delete(1.0, tk.END)
        self.error_log_text.config(state=tk.DISABLED)
        # 清空所有错误日志存储
        self.all_error_logs = []
    
    def _log_warning(self, phone: str, warning_message: str):
        """添加警告日志（黄色显示）
        
        Args:
            phone: 手机号
            warning_message: 警告内容
        """
        import datetime
        timestamp = datetime.datetime.now().strftime("%H:%M:%S")
        
        # 格式：[时间] 手机号 | 警告内容
        warning_entry = f"[{timestamp}] {phone} | ⚠️ {warning_message}"
        
        # 存储到错误日志列表（带标记，方便后续清理）
        self.all_error_logs.append({
            'phone': phone,
            'type': 'warning',
            'message': warning_entry,
            'timestamp': timestamp
        })
        
        # 显示警告日志（黄色）
        self.error_log_text.config(state=tk.NORMAL)
        self.error_log_text.tag_configure("warning", foreground="orange")
        self.error_log_text.insert(tk.END, f"{warning_entry}\n", "warning")
        self.error_log_text.see(tk.END)
        self.error_log_text.config(state=tk.DISABLED)
    
    def _clear_account_warnings(self, phone: str):
        """清除指定账号的所有错误日志（重试成功后调用）
        
        Args:
            phone: 手机号
        """
        # 过滤掉该账号的所有错误日志（包括error和warning）
        self.all_error_logs = [
            log for log in self.all_error_logs 
            if not (isinstance(log, dict) and log.get('phone') == phone)
        ]
        
        # 重新显示错误日志
        self.error_log_text.config(state=tk.NORMAL)
        self.error_log_text.delete(1.0, tk.END)
        
        for log in self.all_error_logs:
            if isinstance(log, dict):
                # 新格式（带类型）
                tag = log.get('type', 'error')
                self.error_log_text.insert(tk.END, f"{log['message']}\n", tag)
            else:
                # 旧格式（字符串）
                self.error_log_text.insert(tk.END, f"{log}\n", "error")
        
        self.error_log_text.see(tk.END)
        self.error_log_text.config(state=tk.DISABLED)
    
    def _update_account_status_in_table(self, phone: str, status: str):
        """更新表格中指定账号的状态
        
        Args:
            phone: 手机号
            status: 新状态(如"执行中"、"成功"、"失败"等)
        """
        for item_id in self.results_tree.get_children():
            values = self.results_tree.item(item_id, 'values')
            if values and len(values) > 13 and values[0] == phone:  # 确保至少有14列（索引0-13）
                # 找到对应的行，更新状态列(第14列，索引13)
                new_values = list(values)
                new_values[13] = status  # 状态列（索引13）
                self.results_tree.item(item_id, values=tuple(new_values))
                break
    
    def _update_pending_count(self):
        """更新待处理账号数量（统计未勾选的账号）"""
        pending_count = 0
        for item_id in self.results_tree.get_children():
            # 检查该行是否未被勾选（未勾选 = 待处理）
            if not self.checked_items.get(item_id, False):
                pending_count += 1
        
        self.account_pending_var.set(f"待处理: {pending_count}")
    
    def _update_stats(self, total: int, success: int, failed: int, 
                     total_checkin_reward: float = 0.0, 
                     total_balance: float = 0.0,
                     total_points: int = 0,
                     total_vouchers: float = 0.0,
                     total_coupons: int = 0):
        """更新统计信息（运行时使用传入的实时统计数据）
        
        注意：此方法在运行时使用传入的实时统计数据，不重新遍历表格
        只有在加载账号或手动刷新时才调用 _update_stats_from_table()
        
        Args:
            total: 总账号数
            success: 成功数
            failed: 失败数
            total_checkin_reward: 总签到奖励(所有账号累加)
            total_balance: 总余额(所有账号累加)
            total_points: 总积分(所有账号累加)
            total_vouchers: 总抵扣券(所有账号累加)
            total_coupons: 总优惠券(所有账号累加)
        """
        # 运行时直接使用传入的统计数据，不重新遍历表格（性能优化）
        self.total_var.set(f"总计: {total}")
        self.success_var.set(f"成功: {success}")
        self.failed_var.set(f"失败: {failed}")
        self.total_balance_var.set(f"总余额: {total_balance:.2f} 元")
        self.total_checkin_reward_var.set(f"总签到奖励: {total_checkin_reward:.2f} 元")
    
    def _update_stats_from_table(self):
        """从表格中统计所有账号的数据
        
        统计规则：
        1. 总计：表格中所有账号
        2. 成功/失败：根据状态列统计
        3. 余额/签到奖励：只统计成功账号的数据（避免混合历史数据）
        """
        total = 0
        success = 0
        failed = 0
        total_balance = 0.0
        total_checkin_reward = 0.0
        
        # 遍历表格中的所有行
        for item_id in self.results_tree.get_children():
            values = self.results_tree.item(item_id, 'values')
            if not values or len(values) < 16:  # 现在有16列
                continue
            
            total += 1
            
            # 表格列顺序：
            # 0:phone, 1:nickname, 2:user_id, 3:balance_before, 4:points, 5:vouchers, 6:coupons,
            # 7:checkin_reward, 8:checkin_total_times, 9:balance_after, 10:transfer_amount, 
            # 11:transfer_recipient, 12:duration, 13:status, 14:login_method, 15:owner
            
            # 状态列（索引13）
            status = values[13]
            is_success = "成功" in status  # 使用in判断，因为状态可能是"✅ 成功"
            is_failed = "失败" in status or "❌" in status  # 包含失败和错误状态
            
            if is_success:
                success += 1
            elif is_failed:
                failed += 1
            # 注意：待处理状态不计入成功或失败
            
            # 只统计成功账号的余额和签到奖励（避免混合历史数据）
            if is_success:
                # 余额（索引9，balance_after）
                try:
                    balance_str = values[9]
                    if balance_str and balance_str != "N/A" and balance_str != "-" and balance_str != "待处理" and balance_str != "None":
                        total_balance += float(balance_str)
                except (ValueError, IndexError, TypeError) as e:
                    # 记录解析错误，便于调试
                    phone = values[0] if values else 'unknown'
                    print(f"[统计] 解析余额失败 - 账号: {phone}, 值: {balance_str}, 错误: {e}")
                
                # 签到奖励（索引7）
                try:
                    checkin_reward_str = values[7]
                    if checkin_reward_str and checkin_reward_str != "N/A" and checkin_reward_str != "-" and checkin_reward_str != "待处理" and checkin_reward_str != "None":
                        total_checkin_reward += float(checkin_reward_str)
                except (ValueError, IndexError, TypeError) as e:
                    # 记录解析错误，便于调试
                    phone = values[0] if values else 'unknown'
                    print(f"[统计] 解析签到奖励失败 - 账号: {phone}, 值: {checkin_reward_str}, 错误: {e}")
        
        # 更新统计显示
        self.total_var.set(f"总计: {total}")
        self.success_var.set(f"成功: {success}")
        self.failed_var.set(f"失败: {failed}")
        self.total_balance_var.set(f"总余额: {total_balance:.2f} 元")
        self.total_checkin_reward_var.set(f"总签到奖励: {total_checkin_reward:.2f} 元")
    
    def _get_success_failed_from_table(self) -> tuple:
        """从表格中快速获取成功和失败数量（用于进度条显示）
        
        Returns:
            tuple: (成功数, 失败数)
        """
        success = 0
        failed = 0
        
        # 遍历表格中的所有行
        for item_id in self.results_tree.get_children():
            values = self.results_tree.item(item_id, 'values')
            if not values or len(values) < 16:
                continue
            
            # 状态列（索引13）
            status = values[13]
            is_success = "成功" in status
            is_failed = "失败" in status or "❌" in status
            
            if is_success:
                success += 1
            elif is_failed:
                failed += 1
        
        return (success, failed)
    
    def _format_status(self, account_result: AccountResult) -> str:
        """格式化状态文本
        
        Args:
            account_result: 账号处理结果
            
        Returns:
            格式化后的状态文本，最长30个字符
        """
        if account_result.success:
            return "✅ 成功"
        
        # 如果有error_type，使用映射的错误文本
        if account_result.error_type:
            status_text = ErrorType.to_display_text(account_result.error_type)
        else:
            # 兼容旧代码：如果没有error_type，显示"失败"
            status_text = "失败"
        
        # 限制长度在30个字符以内
        if len(status_text) > 30:
            status_text = status_text[:27] + "..."
        
        return status_text
    
    def _add_result_to_table(self, account_result):
        """添加账号结果到表格(更新现有行)
        
        Args:
            account_result: AccountResult对象
        """
        # 准备显示数据 - 确保所有字符串都是正确的UTF-8编码
        # 昵称 - 确保编码正确
        if account_result.nickname:
            # 如果是bytes，解码为str
            if isinstance(account_result.nickname, bytes):
                nickname = account_result.nickname.decode('utf-8', errors='replace')
            else:
                nickname = str(account_result.nickname)
        else:
            nickname = "N/A"
        
        # 用户ID
        user_id = account_result.user_id if account_result.user_id else "N/A"
        phone = account_result.phone
        
        # 管理员 - 从数据库获取
        owner_name = "-"
        try:
            db = LocalDatabase()
            records = db.get_history_records(phone, limit=1)
            if records and records[0].get('owner'):
                owner_name = records[0].get('owner')
        except:
            pass
        
        # 余额前
        balance_before = f"{account_result.balance_before:.2f}" if account_result.balance_before is not None else "N/A"
        
        # 积分
        points = str(account_result.points) if account_result.points is not None else "N/A"
        
        # 抵扣券
        vouchers = f"{account_result.vouchers:.2f}" if account_result.vouchers is not None else "N/A"
        
        # 优惠券
        coupons = str(account_result.coupons) if account_result.coupons is not None else "N/A"
        
        # 签到奖励
        checkin_reward = f"{account_result.checkin_reward:.2f}" if account_result.checkin_reward else "0"
        
        # 签到总次数
        checkin_total_times = str(account_result.checkin_total_times) if account_result.checkin_total_times is not None else "N/A"
        
        # 余额（最终余额）
        balance_after = f"{account_result.balance_after:.2f}" if account_result.balance_after is not None else "N/A"
        
        # 转账金额
        transfer_amount = f"{account_result.transfer_amount:.2f}" if hasattr(account_result, 'transfer_amount') and account_result.transfer_amount is not None and account_result.transfer_amount > 0 else "-"
        
        # 转账收款人
        transfer_recipient = account_result.transfer_recipient if hasattr(account_result, 'transfer_recipient') and account_result.transfer_recipient else "-"
        
        # 耗时
        duration = f"{round(account_result.duration, 3)}" if account_result.duration is not None else "N/A"
        
        # 状态 - 使用新的格式化方法
        status = self._format_status(account_result)
        
        # 登录方式
        login_method = account_result.login_method if account_result.login_method else "N/A"
        
        # 准备新的值（包含owner和转账字段）- 注意顺序要与columns定义一致
        values = (
            phone, nickname, user_id, balance_before, points, vouchers, coupons,
            checkin_reward, checkin_total_times,
            balance_after, transfer_amount, transfer_recipient, duration, status, login_method, owner_name
        )
        
        # 根据余额变化确定标签(用于颜色高亮)
        tags = []
        if account_result.balance_change is not None:
            if account_result.balance_change > 0:
                tags.append("positive")
            elif account_result.balance_change < 0:
                tags.append("negative")
            else:
                tags.append("neutral")
        else:
            tags.append("neutral")
        
        # 查找表格中是否已有该账号的行
        item_id = None
        for existing_item in self.results_tree.get_children():
            existing_values = self.results_tree.item(existing_item, 'values')
            if existing_values and len(existing_values) > 0 and existing_values[0] == phone:
                item_id = existing_item
                break
        
        # 如果成功完成，添加checked标签并自动勾选
        if account_result.success:
            tags.append("checked")
            if item_id:
                # 更新现有行
                self.results_tree.item(item_id, text=self.checkbox_checked_text, values=values, tags=tuple(tags))
                self.checked_items[item_id] = True
            else:
                # 插入新行
                item_id = self.results_tree.insert("", tk.END, text=self.checkbox_checked_text, values=values, tags=tuple(tags))
                self.checked_items[item_id] = True
        else:
            # 失败的显示未勾选
            if item_id:
                # 更新现有行
                self.results_tree.item(item_id, text=self.checkbox_unchecked_text, values=values, tags=tuple(tags))
                self.checked_items[item_id] = False
            else:
                # 插入新行
                item_id = self.results_tree.insert("", tk.END, text=self.checkbox_unchecked_text, values=values, tags=tuple(tags))
                self.checked_items[item_id] = False
        
        # 更新待处理数量
        self._update_pending_count()
        
        # 更新统计（从表格重新计算）
        self._update_stats_from_table()
        
        # 自动滚动到最新结果
        self.results_tree.see(item_id)
        
        # 只保存当前完成的账号到历史记录（不是批量保存所有账号）
        self._save_single_account_to_history(account_result)
    
    def _update_progress(self, current: int, total: int, message: str):
        """更新进度条和进度信息
        
        Args:
            current: 当前已完成数量
            total: 总数量（待处理账号数，不是所有账号数）
            message: 进度消息
        """
        if total > 0:
            progress = (current / total) * 100
            self.progress_var.set(progress)
        else:
            # 如果总数为0，设置进度为100%
            self.progress_var.set(100)
        self.progress_label_var.set(message)
    
    def _save_single_account_to_history(self, account_result):
        """保存单个账号到历史记录数据库
        
        Args:
            account_result: AccountResult对象
        
        规则：
        - 只保存成功的记录
        - 失败的记录不更新数据库
        - 防止同一账号在同一次运行中被重复保存（避免签到奖励累加）
        """
        try:
            # 只保存成功的账号
            if account_result.success:
                # 检查是否已经保存过（防止重复保存导致签到奖励累加）
                with self.saved_accounts_lock:
                    if account_result.phone in self.saved_accounts:
                        print(f"[历史记录] - 跳过重复保存: {account_result.phone} (已在 {self.saved_accounts[account_result.phone]} 保存)")
                        return
                    
                    # 标记为已保存
                    from datetime import datetime
                    self.saved_accounts[account_result.phone] = datetime.now().strftime('%H:%M:%S')
                
                db = LocalDatabase()
                
                # 获取当前日期
                run_date = datetime.now().strftime('%Y-%m-%d')
                
                # 调试：检查account_result对象的转账属性
                print(f"[历史记录] 账号 {account_result.phone} 状态: {account_result.success}")
                print(f"[历史记录]   - hasattr transfer_amount: {hasattr(account_result, 'transfer_amount')}")
                print(f"[历史记录]   - transfer_amount value: {getattr(account_result, 'transfer_amount', 'N/A')}")
                print(f"[历史记录]   - hasattr transfer_recipient: {hasattr(account_result, 'transfer_recipient')}")
                print(f"[历史记录]   - transfer_recipient value: {getattr(account_result, 'transfer_recipient', 'N/A')}")
                
                # 准备记录数据（金额字段格式化为2位小数）
                record = {
                    'phone': account_result.phone,
                    'nickname': account_result.nickname if account_result.nickname else '',
                    'user_id': account_result.user_id if account_result.user_id else '',
                    'balance_before': round(account_result.balance_before, 2) if account_result.balance_before is not None else None,
                    'points': account_result.points if account_result.points is not None else None,
                    'vouchers': round(account_result.vouchers, 2) if account_result.vouchers is not None else None,
                    'coupons': account_result.coupons if account_result.coupons is not None else None,
                    'checkin_reward': round(account_result.checkin_reward, 2) if account_result.checkin_reward else 0.0,
                    'checkin_total_times': account_result.checkin_total_times if account_result.checkin_total_times is not None else None,
                    'checkin_balance_after': round(account_result.checkin_balance_after, 2) if account_result.checkin_balance_after is not None else None,
                    'balance_after': round(account_result.balance_after, 2) if account_result.balance_after is not None else None,
                    'duration': round(account_result.duration, 2) if account_result.duration is not None else 0.0,
                    'status': '成功',
                    'login_method': account_result.login_method if account_result.login_method else '',
                    'transfer_amount': round(account_result.transfer_amount, 2) if hasattr(account_result, 'transfer_amount') and account_result.transfer_amount is not None else 0.0,
                    'transfer_recipient': account_result.transfer_recipient if hasattr(account_result, 'transfer_recipient') and account_result.transfer_recipient else '',
                    'run_date': run_date
                }
                
                # 调试：打印转账信息
                print(f"[历史记录]   - record transfer_amount: {record['transfer_amount']}")
                print(f"[历史记录]   - record transfer_recipient: {record['transfer_recipient']}")
                if record['transfer_amount'] and record['transfer_amount'] > 0:
                    print(f"[历史记录] ✓ 账号 {account_result.phone} 转账信息: {record['transfer_amount']:.2f} 元 → {record['transfer_recipient']}")
                
                # 保存到数据库
                if db.upsert_history_record(record):
                    print(f"[历史记录] ✓ 已保存账号 {account_result.phone}")
                else:
                    print(f"[历史记录] ✗ 保存失败: {account_result.phone}")
            else:
                print(f"[历史记录] - 跳过失败账号: {account_result.phone}")
                    
        except Exception as e:
            # 静默失败，不影响主流程
            print(f"保存单个账号历史记录失败: {e}")
            import traceback
            traceback.print_exc()
    
    def _save_to_history(self):
        """历史记录保存（已废弃批量保存功能）
        
        注意：此函数不再执行批量保存。
        每个账号在执行完成后会立即保存到数据库（在 account_manager.update_account_result 中）。
        这样可以避免字段错位和数据丢失问题。
        """
        try:
            # 不再需要批量保存，只输出提示信息
            print(f"[历史记录] ✓ 所有账号的结果已在执行过程中实时保存到数据库")
            
        except Exception as e:
            # 静默失败，不影响主流程
            print(f"[历史记录] ⚠️ 保存历史记录失败: {e}")
            import traceback
            traceback.print_exc()
    
    def _load_from_history(self, date_filter: str = None):
        """从历史记录数据库加载最新数据
        
        Args:
            date_filter: 日期过滤器（格式：YYYY-MM-DD），如果为None则加载所有历史记录
        
        Returns:
            list: 历史记录列表
        """
        try:
            db = LocalDatabase()
            
            # 获取所有历史记录（不限制日期，按手机号去重保留最新）
            if date_filter:
                # 如果指定了日期，只加载该日期的记录
                all_records = db.get_history_records(start_date=date_filter, end_date=date_filter, limit=10000)
            else:
                # 不指定日期，加载所有历史记录
                all_records = db.get_history_records(limit=10000)
            
            # 按手机号去重，保留最新的记录（同一天可能有多次执行）
            phone_dict = {}
            for record in all_records:
                phone = record['phone']
                # 如果该手机号还没有记录，或者当前记录更新，则更新
                if phone not in phone_dict:
                    phone_dict[phone] = record
                else:
                    # 比较创建时间，保留最新的
                    if record.get('created_at', '') > phone_dict[phone].get('created_at', ''):
                        phone_dict[phone] = record
            
            # 转换为旧格式（兼容现有代码）- 确保所有字符串都是正确的UTF-8编码
            history_data = []
            for phone, record in phone_dict.items():
                # 处理昵称编码
                nickname = record['nickname']
                if nickname:
                    # 如果是bytes，解码为str
                    if isinstance(nickname, bytes):
                        nickname = nickname.decode('utf-8', errors='replace')
                    else:
                        # 确保是字符串
                        nickname = str(nickname)
                
                history_data.append({
                    '手机号': record['phone'],
                    '昵称': nickname,
                    '用户ID': record['user_id'],
                    '余额前(元)': str(record['balance_before']),
                    '积分': str(record['points']),
                    '抵扣券(张)': str(record['vouchers']),
                    '优惠券(张)': str(record['coupons']),
                    '签到奖励(元)': str(record['checkin_reward']),
                    '签到总次数': str(record['checkin_total_times']),
                    '余额(元)': str(record['balance_after']),
                    '耗时(秒)': str(record['duration']),
                    '状态': record['status'],
                    '登录方式': record['login_method']
                })
            
            return history_data
            
        except Exception as e:
            print(f"加载历史记录失败: {e}")
            import traceback
            traceback.print_exc()
            return []
            traceback.print_exc()
            return []
    
    def _test_launch_emulator(self):
        """测试启动模拟器"""
        path = self.emulator_path_var.get()
        if not path:
            messagebox.showerror("错误", "请先配置模拟器路径")
            return
        
        self._log("正在测试启动模拟器...")
        self.status_var.set("启动中...")
        
        thread = threading.Thread(target=self._test_launch_thread, daemon=True)
        thread.start()
    
    def _test_launch_thread(self):
        """测试启动模拟器线程"""
        try:
            asyncio.run(self._test_launch_async())
        except Exception as e:
            self.root.after(0, lambda: self._log(f"启动失败: {e}"))
        finally:
            self.root.after(0, lambda: self.status_var.set("就绪"))
    
    async def _test_launch_async(self):
        """异步测试启动模拟器"""
        path = self.emulator_path_var.get()
        controller = EmulatorController(path)
        
        if not controller.is_available():
            self.root.after(0, lambda: self._log("模拟器控制台程序未找到"))
            return
        
        self.root.after(0, lambda: self._log(f"使用: {controller.get_emulator_info()}"))
        
        # 检查模拟器是否已经在运行
        is_running = await controller._is_running(0)
        
        if is_running:
            self.root.after(0, lambda: self._log("模拟器已在运行"))
        else:
            self.root.after(0, lambda: self._log("正在启动模拟器实例 0..."))
            success = await controller.launch_instance(0, timeout=120)
            
            if not success:
                self.root.after(0, lambda: self._log("模拟器启动失败"))
                return
            
            self.root.after(0, lambda: self._log("模拟器启动成功！"))
        
        # 获取 ADB 端口
        adb_port = await controller.get_adb_port(0)
        self.root.after(0, lambda port=adb_port: self._log(f"ADB 端口: {port}"))
        
        # 测试 ADB 连接
        adb_path = controller.get_adb_path()
        if adb_path:
            self.root.after(0, lambda p=adb_path: self._log(f"ADB 路径: {p}"))
            
            adb = ADBBridge(adb_path)
            device_id = f"127.0.0.1:{adb_port}"
            
            self.root.after(0, lambda: self._log(f"正在连接 ADB: {device_id}"))
            connected = await adb.connect(device_id)
            
            if connected:
                self.root.after(0, lambda: self._log(f"ADB 连接成功: {device_id}"))
                
                # 检查分辨率
                resolution = await controller.get_resolution(0)
                if resolution:
                    width, height = resolution
                    self.root.after(0, lambda w=width, h=height: self._log(f"当前分辨率: {w}x{h}"))
                    
                    if width == 540 and height == 960:
                        self.root.after(0, lambda: self._log("OK 分辨率正确 (540x960)"))
                    else:
                        self.root.after(0, lambda: self._log(f"! 分辨率不正确！需要 540x960"))
                        self.root.after(0, lambda: self._log("请在模拟器设置中修改分辨率为 540x960"))
                else:
                    self.root.after(0, lambda: self._log("无法获取分辨率"))
                
                # 搜索溪盟商城应用
                self.root.after(0, lambda: self._log("正在搜索溪盟商城应用..."))
                target_app = await adb.find_package_by_name(device_id, "溪盟")
                if target_app:
                    self.root.after(0, lambda app=target_app: self._log(f"OK 找到应用: {app}"))
                else:
                    self.root.after(0, lambda: self._log("! 未找到溪盟商城应用"))
                    # 列出所有第三方应用
                    packages = await adb.get_installed_packages(device_id)
                    third_party = [p for p in packages if not p.startswith("com.android") and not p.startswith("com.google")]
                    if third_party:
                        self.root.after(0, lambda pkgs=third_party[:10]: self._log(f"已安装的第三方应用: {pkgs}"))
            else:
                self.root.after(0, lambda: self._log("ADB 连接失败"))
        else:
            self.root.after(0, lambda: self._log("未找到 ADB 路径"))

    def _start_automation(self):
        """开始自动化"""
        # 检查模型是否已加载
        if not self.models_loaded:
            from .model_manager import ModelManager
            model_manager = ModelManager.get_instance()
            if not model_manager.is_initialized():
                messagebox.showwarning("提示", "模型正在后台加载中，请稍候...\n\n加载完成后会在日志中显示提示。")
                return
        
        path = self.emulator_path_var.get()
        if not path:
            messagebox.showerror("错误", "请先配置模拟器路径")
            return
        
        accounts_file = self.accounts_file_var.get()
        if not accounts_file:
            messagebox.showerror("错误", "请选择有效的账号文件")
            return
        
        # 检查明文文件或加密文件是否存在
        plain_file_exists = Path(accounts_file).exists()
        enc_file_exists = Path(f"{accounts_file}.enc").exists()
        
        if not plain_file_exists and not enc_file_exists:
            messagebox.showerror("错误", f"账号文件不存在\n\n请检查以下路径：\n- {accounts_file}\n- {accounts_file}.enc")
            return
        
        self._save_config()
        
        # 重置事件标志
        self.stop_event.clear()
        self.pause_event.clear()
        
        # 清空已保存账号集合（新的一次运行）
        with self.saved_accounts_lock:
            self.saved_accounts.clear()
            print("[历史记录] 已清空已保存账号集合（新的一次运行）")
        
        self.is_running = True
        self.is_paused = False
        self.start_btn.config(state=tk.DISABLED)
        self.pause_btn.config(state=tk.NORMAL, text="⏸ 暂停")
        self.stop_btn.config(state=tk.NORMAL)
        self.status_var.set("运行中...")
        
        thread = threading.Thread(target=self._run_automation_thread, daemon=True)
        thread.start()
    
    def _pause_automation(self):
        """暂停/继续自动化"""
        if self.is_paused:
            # 继续运行
            self.is_paused = False
            self.pause_event.clear()  # 清除暂停标志
            self.pause_btn.config(text="⏸ 暂停")
            self.status_var.set("运行中...")
            self._log("▶ 继续运行")
        else:
            # 暂停
            self.is_paused = True
            self.pause_event.set()  # 设置暂停标志
            self.pause_btn.config(text="▶ 继续")
            self.status_var.set("已暂停")
            self._log("⏸ 已暂停，点击'继续'按钮恢复运行")
    
    def _stop_automation(self):
        """停止自动化（强制终止）"""
        self.is_running = False
        self.is_paused = False
        self.stop_event.set()  # 设置停止标志
        self.pause_event.clear()  # 清除暂停标志
        
        # 更新按钮状态
        self.start_btn.config(state=tk.NORMAL)
        self.pause_btn.config(state=tk.DISABLED, text="⏸ 暂停")
        self.stop_btn.config(state=tk.DISABLED)
        
        # 将所有"执行中"的账号标记为"手动停止"
        for item_id in self.results_tree.get_children():
            values = self.results_tree.item(item_id, 'values')
            if values and len(values) > 13:  # 至少要有14列（索引0-13）
                status = values[13]  # 状态列（索引13）
                if status == "执行中":
                    # 更新状态为"手动停止"
                    new_values = list(values)
                    new_values[13] = "手动停止"  # 状态列（索引13）
                    self.results_tree.item(item_id, values=tuple(new_values))
        
        # 取消所有未开始的任务
        if self.pending_futures:
            cancelled_count = 0
            for future in self.pending_futures:
                if future.cancel():  # 尝试取消任务
                    cancelled_count += 1
            if cancelled_count > 0:
                self._log(f"⏹ 取消了 {cancelled_count} 个未开始的任务")
        
        # 强制终止线程池中的所有任务
        if self.executor:
            self._log("⏹ 强制终止所有运行中的任务...")
            try:
                # 立即关闭线程池，不等待任务完成
                self.executor.shutdown(wait=False, cancel_futures=True)
                self._log("✓ 线程池已强制关闭")
            except Exception as e:
                self._log(f"⚠️ 关闭线程池时出错: {e}")
            finally:
                self.executor = None
        
        # 更新统计（从表格重新统计，确保数据准确）
        self._update_stats_from_table()
        
        # 更新状态
        self.status_var.set("已停止")
        self._log("⏹ 已强制停止所有任务")
        self._log("提示：如需重新运行，请点击'开始运行'按钮")
    
    def _check_stop_or_pause(self):
        """检查是否需要停止或暂停
        
        Returns:
            bool: True表示需要停止，False表示可以继续
        """
        # 检查停止标志
        if self.stop_event.is_set():
            return True
        
        # 检查暂停标志，如果暂停则等待
        while self.pause_event.is_set():
            time.sleep(0.1)
            # 在暂停期间也要检查停止标志
            if self.stop_event.is_set():
                return True
        
        return False
    
    def _run_automation_thread(self):
        """后台运行自动化任务"""
        try:
            asyncio.run(self._run_automation_async())
        except Exception as e:
            error_msg = f"运行出错: {e}"
            self.root.after(0, lambda msg=error_msg: self._log(msg))
        finally:
            self.root.after(0, self._on_automation_complete)
    
    async def _run_automation_async(self):
        """异步运行自动化"""
        self.root.after(0, lambda: self._log("开始运行..."))
        
        # 检查停止标志(使用Event)
        if self.stop_event.is_set():
            return
        
        # 初始化模拟器控制器
        path = self.emulator_path_var.get()
        controller = EmulatorController(path)
        
        if not controller.is_available():
            self.root.after(0, lambda: self._log("模拟器控制台程序未找到"))
            return
        
        # 检查停止标志(使用Event)
        if self.stop_event.is_set():
            return
        
        # 加载账号
        account_manager = AccountManager(self.accounts_file_var.get())
        accounts = account_manager.load_accounts()
        
        if not accounts:
            self.root.after(0, lambda: self._log("未能加载账号，请检查账号文件"))
            return
        
        total = len(accounts)
        
        # 检查停止标志（使用Event）
        if self.stop_event.is_set():
            return
        
        # 统计未勾选账号数（需要处理的账号数）
        # 只统计当前显示的（未被detach的）未勾选账号
        unchecked_phones = set()
        with self.stats_lock:
            # 获取当前显示的项目（未被detach的）
            visible_items = self.results_tree.get_children()
            for item_id in visible_items:
                # 只统计未勾选的账户
                if not self.checked_items.get(item_id, False):
                    values = self.results_tree.item(item_id, 'values')
                    if values and len(values) > 0:
                        unchecked_phones.add(values[0])
        
        unchecked_count = len(unchecked_phones)
        
        # 简化日志：只显示需要处理的账号数
        self.root.after(0, lambda c=unchecked_count: 
                       self._log(f"需要处理 {c} 个账号"))
        
        # 如果没有未勾选的账号，直接返回
        if unchecked_count == 0:
            self.root.after(0, lambda: self._log("没有需要处理的账号（所有账号都已完成）"))
            return
        
        # 批量启动模拟器实例（支持多开）
        auto_launch = self.auto_launch_var.get()
        max_workers_config = self.instance_count_var.get()
        max_workers = min(unchecked_count, max_workers_config)
        
        if auto_launch:
            self.root.after(0, lambda w=max_workers: self._log(f"准备启动 {w} 个实例"))
        self.root.after(0, lambda: self._update_progress(0, unchecked_count, f"正在检测模拟器实例... (待处理: {unchecked_count})"))
        
        # 检测正在运行的模拟器实例
        running_instances = await controller.get_running_instances()
        
        if running_instances:
            # 使用检测到的实例ID初始化实例池
            self.instance_pool = running_instances[:max_workers]
            self.root.after(0, lambda: self._log(f"使用 {len(self.instance_pool)} 个已运行的实例"))
        else:
            if not auto_launch:
                # 未启用自动启动，且没有运行中的实例
                self.root.after(0, lambda: self._log("未检测到运行中的实例"))
                self.root.after(0, lambda: self._log("请手动启动模拟器，或在设置中启用'自动启动模拟器'"))
                return
            
            # 初始化实例池
            self.instance_pool = list(range(max_workers))
            
            # 批量启动所有需要的模拟器实例
            for instance_id in range(max_workers):
                if self.stop_event.is_set():
                    return
                
                self.root.after(0, lambda i=instance_id: self._log(f"启动实例 {i}..."))
                
                # 检查实例是否已运行
                is_running = await controller._is_running(instance_id)
                
                if is_running:
                    self.root.after(0, lambda i=instance_id: self._log(f"✓ 实例 {i} 已运行"))
                else:
                    # 启动实例
                    success = await controller.launch_instance(instance_id, timeout=120)
                    
                    if not success:
                        self.root.after(0, lambda i=instance_id: self._log(f"✗ 实例 {i} 启动失败"))
                        # 从实例池中移除失败的实例
                        with self.instance_lock:
                            if instance_id in self.instance_pool:
                                self.instance_pool.remove(instance_id)
                    else:
                        self.root.after(0, lambda i=instance_id: self._log(f"✓ 实例 {i} 启动成功"))
                
                # 短暂延迟，避免同时启动太多实例
                await asyncio.sleep(2)
        
        # 检查是否有可用实例
        if not self.instance_pool:
            self.root.after(0, lambda: self._log("没有可用的模拟器实例"))
            return
        
        self.root.after(0, lambda: self._log(f"✓ 成功启动 {len(self.instance_pool)} 个实例"))
        
        # 检查停止标志（使用Event）
        if self.stop_event.is_set():
            return
        
        # 处理账号
        processed = 0
        success_count = 0
        failed_count = 0
        total_draw = 0.0
        total_checkin_reward = 0.0
        total_balance = 0.0
        total_points = 0
        total_vouchers = 0.0
        total_coupons = 0
        
        # 重置统计显示（从0开始）
        self.root.after(0, lambda: self._update_stats(unchecked_count, 0, 0, 0.0, 0.0, 0, 0.0, 0))
        
        # 直接使用已知的溪盟商城包名
        target_app = "com.ry.xmsc"
        
        # 获取Activity名称
        target_activity = self.config.target_app_activity
        if target_activity:
            self.root.after(0, lambda: self._log(f"使用Activity: {target_activity}"))
        
        # 使用动态任务队列（所有实例共享一个账号队列）
        # 删除"使用 X 个工作线程"消息
        
        # 创建共享的账号队列（线程安全）
        import queue
        account_queue = queue.Queue()
        
        # 统计变量（线程安全）
        processed = 0
        success_count = 0
        failed_count = 0
        total_checkin_reward = 0.0
        total_balance = 0.0
        total_points = 0
        total_vouchers = 0.0
        total_coupons = 0
        
        # 失败账号列表（用于重试）
        failed_accounts = []
        
        # 将未勾选的账号加入队列（只运行未勾选的账号）
        queued_count = 0
        for i, account in enumerate(accounts):
            # 检查是否需要停止
            if self.stop_event.is_set():
                self.root.after(0, lambda: self._log("用户中断操作"))
                break
            
            # 只处理未勾选的账号（已勾选的账号跳过）
            if account.phone not in unchecked_phones:
                continue
            
            # 加入共享队列
            account_queue.put((i, account))
            queued_count += 1
        
        self.root.after(0, lambda c=queued_count: self._log(f"✓ 已将 {c} 个账号加入处理队列"))
        
        # 为每个实例创建一个工作线程
        instance_threads = []
        
        # 删除详细日志提示，减少噪音
        
        # 为每个实例创建专属处理线程（从共享队列获取账号）
        def process_instance_accounts(instance_id):
            """处理账号（从共享队列动态获取）"""
            # 为该实例创建独立的日志文件
            import os
            log_dir = Path("runtime_data") / "instance_logs"
            log_dir.mkdir(parents=True, exist_ok=True)
            log_file = log_dir / f"instance_{instance_id}.log"
            
            # 清空旧的日志文件（重新开始执行时清理旧日志）
            try:
                with open(log_file, "w", encoding="utf-8") as f:
                    from datetime import datetime
                    f.write(f"{'='*80}\n")
                    f.write(f"实例 {instance_id} 日志 - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                    f.write(f"{'='*80}\n\n")
            except Exception as e:
                print(f"清空日志文件失败: {e}")
            
            # 为该实例创建带前缀的日志回调（同时输出到GUI和文件）
            def instance_log_callback(msg):
                # 如果消息已经包含实例编号，不要重复添加
                if not msg.startswith(f"[实例{instance_id}]") and not msg.startswith("[实例"):
                    prefixed_msg = f"[实例{instance_id}] {msg}"
                else:
                    prefixed_msg = msg
                
                # 输出到GUI
                self.root.after(0, lambda m=prefixed_msg: self._log(m))
                
                # 检测警告信息并记录到错误日志
                if "⚠️" in msg or "警告" in msg or "失败" in msg or "超时" in msg:
                    # 提取账号信息（如果有）
                    phone = account.phone if 'account' in locals() else "未知"
                    # 记录警告到错误日志
                    self.root.after(0, lambda p=phone, m=msg: self._log_warning(p, m))
                
                # 同时输出到独立的日志文件
                try:
                    from datetime import datetime
                    timestamp = datetime.now().strftime("%H:%M:%S")
                    with open(log_file, "a", encoding="utf-8") as f:
                        f.write(f"[{timestamp}] {prefixed_msg}\n")
                except Exception:
                    pass  # 静默失败，不影响主流程
            
            # 持续从队列获取账号处理
            while True:
                # 检查是否需要停止
                if self.stop_event.is_set():
                    instance_log_callback("收到停止信号，退出")
                    break
                
                # 从队列获取账号（非阻塞，超时1秒）
                try:
                    account_index, account = account_queue.get(timeout=1)
                except queue.Empty:
                    # 队列为空，检查是否所有账号都处理完了
                    if account_queue.empty():
                        instance_log_callback("队列已空，完成所有任务")
                        break
                    else:
                        # 队列暂时为空，继续等待
                        continue
                
                # 检查是否暂停
                while self.pause_event.is_set():
                    time.sleep(0.1)
                    if self.stop_event.is_set():
                        break
                
                if self.stop_event.is_set():
                    break
                
                # 添加分隔线，清晰分隔不同账号
                instance_log_callback("=" * 60)
                instance_log_callback(f"开始处理账号 {account_index+1}/{total}: {account.phone}")
                
                try:
                    # 在新的事件循环中运行异步处理
                    result = asyncio.run(
                        self._process_account_with_instance(
                            controller, instance_id, account, target_app, target_activity,
                            account_manager,
                            instance_log_callback  # 使用带前缀的日志回调
                        )
                    )
                    
                    nonlocal processed, success_count, failed_count
                    nonlocal total_checkin_reward, total_balance, total_points, total_vouchers, total_coupons
                    
                    # P0修复: 使用锁保护所有nonlocal变量的修改
                    with self.stats_lock:
                        processed += 1
                        
                        # 直接使用 result.success，不需要额外判断
                        if result.success:
                            success_count += 1
                            
                            # 累积统计数据
                            if result.checkin_reward:
                                total_checkin_reward += result.checkin_reward
                            if result.balance_after is not None:
                                total_balance += result.balance_after
                            if result.points is not None:
                                total_points += result.points
                            if result.vouchers is not None:
                                total_vouchers += result.vouchers
                            if result.coupons is not None:
                                total_coupons += result.coupons
                            
                            # 在锁内获取当前值的副本，用于UI更新
                            current_success = success_count
                            current_failed = failed_count
                            current_tcr = total_checkin_reward
                            current_tb = total_balance
                            current_tp = total_points
                            current_tv = total_vouchers
                            current_tc = total_coupons
                            current_processed = processed
                            
                            # 更新统计和表格（在锁外执行UI操作）
                            self.root.after(0, lambda s=current_success, f=current_failed,
                                           tcr=current_tcr, tb=current_tb,
                                           tp=current_tp, tv=current_tv, tc=current_tc:
                                           self._update_stats(total, s, f, tcr, tb, tp, tv, tc))
                            self.root.after(0, lambda ar=result: self._add_result_to_table(ar))
                            # 删除重复的成功消息，因为 _process_account 已经输出了 "✓ 账号处理完成"
                        else:
                            failed_count += 1
                            
                            # 使用 result.error_message 作为失败原因
                            error_msg = result.error_message if result.error_message else "未知错误"
                            error_type = "处理失败"
                            
                            # 记录到失败日志文件
                            try:
                                from .failure_logger import get_failure_logger
                                failure_logger = get_failure_logger()
                                failure_logger.log_failure(
                                    phone=account.phone,
                                    user_id=result.user_id if result else None,
                                    nickname=result.nickname if result else None,
                                    error_message=error_msg,
                                    error_type=error_type
                                )
                            except Exception as log_err:
                                instance_log_callback(f"记录失败日志时出错: {log_err}")
                            
                            failed_accounts.append((account, error_msg))
                            instance_log_callback(f"✗ 账号 {account.phone} 处理失败: {error_msg}")
                            
                            # 在锁内获取当前值的副本
                            current_success = success_count
                            current_failed = failed_count
                            current_tcr = total_checkin_reward
                            current_tb = total_balance
                            current_tp = total_points
                            current_tv = total_vouchers
                            current_tc = total_coupons
                            current_processed = processed
                            
                            # 更新统计（在锁外执行UI操作）
                            self.root.after(0, lambda s=current_success, f=current_failed,
                                           tcr=current_tcr, tb=current_tb,
                                           tp=current_tp, tv=current_tv, tc=current_tc:
                                           self._update_stats(total, s, f, tcr, tb, tp, tv, tc))
                        
                        # 更新进度（显示实际处理的账号进度，成功/失败数从表格统计）
                        # 从表格统计成功/失败数，确保与统计区域一致
                        table_success, table_failed = self._get_success_failed_from_table()
                        self.root.after(0, lambda p=current_processed, t=queued_count, s=table_success, f=table_failed: 
                                       self._update_progress(p, t, f"进度: {p}/{t} | 成功: {s} | 失败: {f}"))
                
                except Exception as e:
                    # 捕获所有异常,确保实例不会因为单个账号的异常而停止
                    instance_log_callback(f"❌ 账号 {account.phone} 处理时发生异常: {e}")
                    import traceback
                    instance_log_callback(f"异常详情: {traceback.format_exc()}")
                    
                    # 记录到失败日志文件
                    try:
                        from .failure_logger import get_failure_logger
                        failure_logger = get_failure_logger()
                        failure_logger.log_failure(
                            phone=account.phone,
                            user_id=None,
                            nickname=None,
                            error_message=str(e),
                            error_type="异常错误"
                        )
                    except Exception as log_err:
                        instance_log_callback(f"记录失败日志时出错: {log_err}")
                    
                    # 记录错误日志
                    error_msg = str(e)
                    self.root.after(0, lambda p=account.phone, e=error_msg: 
                                  self._log_error(p, "未知", "未知", e))
                    
                    # P0修复: 使用锁保护nonlocal变量的修改
                    with self.stats_lock:
                        processed += 1
                        failed_count += 1
                        
                        # 在锁内获取当前值的副本
                        current_success = success_count
                        current_failed = failed_count
                        current_tcr = total_checkin_reward
                        current_tb = total_balance
                        current_tp = total_points
                        current_tv = total_vouchers
                        current_tc = total_coupons
                        current_processed = processed
                    
                    if "用户中断" not in error_msg:
                        failed_accounts.append((account, error_msg))
                    
                    # 更新统计（在锁外执行UI操作）
                    self.root.after(0, lambda s=current_success, f=current_failed,
                                   tcr=current_tcr, tb=current_tb,
                                   tp=current_tp, tv=current_tv, tc=current_tc:
                                   self._update_stats(total, s, f, tcr, tb, tp, tv, tc))
                    
                    # 更新进度（显示实际处理的账号进度，成功/失败数从表格统计）
                    # 从表格统计成功/失败数，确保与统计区域一致
                    table_success, table_failed = self._get_success_failed_from_table()
                    self.root.after(0, lambda p=current_processed, t=queued_count, s=table_success, f=table_failed: 
                                   self._update_progress(p, t, f"进度: {p}/{t} | 成功: {s} | 失败: {f}"))
                    
                    # 继续处理下一个账号,不要停止实例
                    instance_log_callback(f"继续处理下一个账号...")
                
                # 标记任务完成
                account_queue.task_done()
            
            # 实例处理完成
            instance_log_callback(f"实例 {instance_id} 已完成所有任务")
        
        # 启动所有实例的处理线程
        for inst_id in self.instance_pool:
            thread = threading.Thread(
                target=process_instance_accounts,
                args=(inst_id,),
                name=f"Instance-{inst_id}-Worker",
                daemon=True
            )
            thread.start()
            instance_threads.append(thread)
        
        # 等待所有线程完成
        self.root.after(0, lambda: self._log(f"等待 {len(instance_threads)} 个实例线程完成..."))
        for thread in instance_threads:
            thread.join()
        
        # 处理失败账号的重试
        if failed_accounts and not self.stop_event.is_set():
            self.root.after(0, lambda: self._log(f"\n{'='*50}"))
            self.root.after(0, lambda: self._log(f"开始重试失败的账号（共 {len(failed_accounts)} 个）"))
            self.root.after(0, lambda: self._log(f"{'='*50}\n"))
            
            # 重新创建线程池
            instance_count = self.instance_count_var.get()
            self.executor = ThreadPoolExecutor(max_workers=instance_count)
            self.pending_futures = []
            
            # 提交失败账号重试
            for account, error_msg in failed_accounts:
                if self.stop_event.is_set():
                    break
                
                self.root.after(0, lambda a=account, e=error_msg: self._log(f"重试账号 {a.phone}（上次失败原因: {e}）"))
                future = self.executor.submit(
                    self._process_account_sync,
                    controller, account, target_app, target_activity,
                    account_manager,
                    0, len(failed_accounts)
                )
                self.pending_futures.append(future)
            
            # 等待重试任务完成
            retry_processed = 0
            for future in as_completed(self.pending_futures):
                if self.stop_event.is_set():
                    break
                
                try:
                    status, account, result = future.result()
                    retry_processed += 1
                    
                    if status == 'success' and result:
                        success_count += 1
                        
                        # 线程安全地累积统计数据
                        with self.stats_lock:
                            if result.checkin_reward:
                                total_checkin_reward += result.checkin_reward
                            if result.balance_after is not None:
                                total_balance += result.balance_after
                            if result.points is not None:
                                total_points += result.points
                            if result.vouchers is not None:
                                total_vouchers += result.vouchers
                            if result.coupons is not None:
                                total_coupons += result.coupons
                        
                        # 更新统计和表格
                        self.root.after(0, lambda s=success_count, f=failed_count,
                                       tcr=total_checkin_reward, tb=total_balance,
                                       tp=total_points, tv=total_vouchers, tc=total_coupons:
                                       self._update_stats(total, s, f, tcr, tb, tp, tv, tc))
                        self.root.after(0, lambda ar=result: self._add_result_to_table(ar))
                        self.root.after(0, lambda a=account: self._log(f"OK 重试成功: {a.phone}"))
                        
                    else:
                        failed_count += 1
                        
                        # 创建失败结果
                        from .models import AccountResult
                        failed_result = AccountResult(
                            phone=account.phone,
                            success=False,
                            error_message=result if status == 'error' else "重试后仍然失败",
                            timestamp=datetime.now(),
                            duration=0,
                            login_method="未知"
                        )
                        
                        # 更新统计和表格
                        self.root.after(0, lambda s=success_count, f=failed_count,
                                       tcr=total_checkin_reward, tb=total_balance,
                                       tp=total_points, tv=total_vouchers, tc=total_coupons:
                                       self._update_stats(total, s, f, tcr, tb, tp, tv, tc))
                        self.root.after(0, lambda fr=failed_result: self._add_result_to_table(fr))
                        self.root.after(0, lambda a=account: self._log(f"X 重试失败: {a.phone}"))
                    
                    # 更新进度
                    self.root.after(0, lambda p=retry_processed, t=len(failed_accounts): 
                                   self._update_progress(p, t, f"重试进度: {p}/{t}"))
                    
                except Exception as e:
                    failed_count += 1
                    self.root.after(0, lambda err=e, a=account: self._log(f"X 重试异常: {a.phone} - {err}"))
            
            # 关闭重试线程池
            if self.executor:
                self.root.after(0, lambda: self._log("正在关闭重试线程池..."))
                self.executor.shutdown(wait=True)
                self.executor = None
                self.pending_futures = []
        
        # 生成报告
        Path(self.config.report_dir).mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # 生成TXT格式报告（带时间戳，保留历史）
        txt_report_path = f"{self.config.report_dir}/report_{timestamp}.txt"
        account_manager.generate_report(txt_report_path)
        self.root.after(0, lambda: self._log(f"TXT报告已生成: {txt_report_path}"))
        
        # 数据已自动保存到数据库
        self.root.after(0, lambda: self._log("处理结果已保存到数据库"))
        
        # 关闭所有模拟器实例
        self.root.after(0, lambda: self._log("正在关闭所有模拟器实例..."))
        for instance_id in range(max_workers):
            try:
                await controller.quit_instance(instance_id)
                self.root.after(0, lambda i=instance_id: self._log(f"已关闭实例 {i}"))
            except:
                pass
    
    def _process_account_sync(self, controller, account, target_app, target_activity,
                              account_manager, account_index, total_accounts):
        """同步包装方法：在线程池中处理单个账号（支持多模拟器实例）
        
        这个方法会在独立的线程中运行：
        1. 从实例池获取一个可用的模拟器实例
        2. 为该实例创建独立的 ADB 连接和自动化组件
        3. 处理账号
        4. 释放实例回池中
        """
        # 创建日志回调（线程安全）
        def log_callback(msg):
            self.root.after(0, lambda m=msg: self._log(m))
        
        # 获取一个可用的模拟器实例
        instance_id = self._acquire_instance()
        if instance_id is None:
            log_callback(f"❌ 账号 {account.phone}: 无可用模拟器实例")
            return ('error', account, "无可用模拟器实例")
        
        log_callback(f"账号 {account.phone} 使用模拟器实例 {instance_id}")
        
        try:
            # 在新的事件循环中运行异步处理
            result = asyncio.run(
                self._process_account_with_instance(
                    controller, instance_id, account, target_app, target_activity,
                    account_manager, log_callback
                )
            )
            return ('success', account, result)
        except Exception as e:
            log_callback(f"账号 {account.phone} 处理异常: {e}")
            return ('error', account, str(e))
        finally:
            # 释放实例回池中
            self._release_instance(instance_id)
            log_callback(f"释放模拟器实例 {instance_id}")
    
    async def _process_account_with_instance(self, controller, instance_id, account, target_app, target_activity,
                                            account_manager, log_callback):
        """使用指定的模拟器实例处理账号
        
        Args:
            controller: 模拟器控制器
            instance_id: 模拟器实例编号
            account: 账号对象
            target_app: 目标应用包名
            target_activity: 目标Activity
            account_manager: 账号管理器
            log_callback: 日志回调函数（已包含实例前缀）
        """
        import time
        start_time = time.time()
        
        # 获取该实例的 ADB 端口
        adb_port = await controller.get_adb_port(instance_id)
        adb_path = controller.get_adb_path()
        device_id = f"127.0.0.1:{adb_port}"
        
        # 删除"使用设备"消息，减少日志噪音
        
        # P1修复: 创建该实例的独立 ADB 连接，使用try...finally确保资源释放
        adb = ADBBridge(adb_path)
        await adb.connect(device_id)
        
        try:
            # 创建该实例的独立自动化组件
            screen_capture = ScreenCapture(adb, self.config.screenshot_dir)
            ui_automation = UIAutomation(adb, screen_capture)
            
            # 根据模拟器类型设置坐标
            emulator_type_str = "mumu"
            
            # 从ModelManager获取共享的检测器实例（不再为每个实例创建新的模型）
            from .model_manager import ModelManager
            from .page_detector import PageDetector
            
            model_manager = ModelManager.get_instance()
            integrated_detector = model_manager.get_page_detector_integrated()
            
            # 创建AutoLogin，传递integrated_detector
            # AutoLogin会从ModelManager获取整合检测器，不需要手动设置
            auto_login = AutoLogin(ui_automation, screen_capture, adb, 
                                  emulator_type=emulator_type_str,
                                  integrated_detector=integrated_detector)
            
            # 不再需要手动设置这些属性，AutoLogin会从ModelManager获取
            # auto_login.detector = ...  # 已在AutoLogin.__init__中从ModelManager获取
            auto_login.page_detector = PageDetector(adb)
            
            # XimengAutomation会从ModelManager获取所有模型
            ximeng = XimengAutomation(ui_automation, screen_capture, auto_login, adb, log_callback=log_callback)
            
            # 记录开始时间
            import time
            import logging
            start_time = time.time()
            
            # 更新表格状态为"执行中"
            self.root.after(0, lambda: self._update_account_status_in_table(account.phone, "执行中"))
            self.root.after(0, lambda: self._update_pending_count())
            
            # 准备阶段：停止应用、检查缓存、启动应用
            await adb.stop_app(device_id, target_app)
            await asyncio.sleep(0.5)  # 优化：减少等待时间从1秒到0.5秒
            
            # 处理缓存验证
            has_valid_cache = False
            
            if auto_login.enable_cache and auto_login.cache_manager:
                # 获取预期的user_id
                expected_user_id = auto_login.cache_manager._get_expected_user_id(account.phone)
                
                # 检查是否有该账号的缓存
                if auto_login.cache_manager.has_cache(account.phone, expected_user_id):
                    log_callback(f"检测到登录缓存")
                    
                    # 恢复缓存
                    if await auto_login.cache_manager.restore_login_cache(device_id, account.phone, user_id=expected_user_id):
                        log_callback("缓存恢复成功")
                        has_valid_cache = True
                    else:
                        await auto_login.cache_manager.clear_app_login_data(device_id, target_app)
                else:
                    await auto_login.cache_manager.clear_app_login_data(device_id, target_app)
            else:
                if auto_login.cache_manager:
                    await auto_login.cache_manager.clear_app_login_data(device_id, target_app)
            
            # 启动应用
            success = await adb.start_app(device_id, target_app, target_activity)
            if not success:
                raise Exception("应用启动失败")
            await asyncio.sleep(1.5)  # 优化：减少等待时间从3秒到1.5秒，让启动流程智能检测
            
            # 处理启动流程（跳过广告、弹窗等）
            # 获取文件日志记录器
            file_logger = logging.getLogger(__name__)
            
            startup_ok = await ximeng.handle_startup_flow_integrated(
                device_id, 
                log_callback=log_callback,
                stop_check=self._check_stop_or_pause,
                package_name=target_app,
                activity_name=target_activity,
                max_retries=3,
                file_logger=file_logger
            )
            
            if not startup_ok:
                if self.stop_event.is_set():
                    raise Exception("用户中断操作")
                raise Exception("启动流程失败")
            
            # 如果有缓存，验证用户ID（快速签到模式除外）
            # 快速签到模式下，跳过ID验证，直接进入登录流程
            enable_profile = getattr(self.config, 'workflow_enable_profile', True)
            
            if has_valid_cache and enable_profile:
                # 完整流程模式：需要验证用户ID
                from .navigator import Navigator
                from .model_manager import ModelManager
                from .page_detector import PageState
                
                # 从ModelManager获取共享的检测器实例
                model_manager = ModelManager.get_instance()
                detector = model_manager.get_page_detector_integrated()
                
                navigator = Navigator(adb, detector)
                
                # 导航到个人页面
                nav_success = await navigator.navigate_to_profile(device_id)
                if nav_success:
                    # 检测页面状态
                    profile_templates = ['已登陆个人页.png', '未登陆个人页.png']
                    page_result = await detector.detect_page_with_priority(
                        device_id, profile_templates, use_cache=False
                    )
                    
                    if page_result and page_result.state == PageState.PROFILE_LOGGED:
                        # 获取当前用户ID验证
                        profile_info = await ximeng.profile_reader.get_full_profile_parallel(device_id)
                        
                        if profile_info and profile_info.get('user_id'):
                            current_user_id = profile_info['user_id']
                            expected_user_id = auto_login.cache_manager._get_expected_user_id(account.phone)
                            
                            if expected_user_id and current_user_id != expected_user_id:
                                log_callback(f"用户ID不匹配，重新登录")
                                
                                # 停止应用并清理
                                await adb.stop_app(device_id, target_app)
                                await asyncio.sleep(0.5)  # 优化：减少等待时间
                                await auto_login.cache_manager.clear_app_login_data(device_id, target_app)
                                
                                # 重新启动
                                await adb.start_app(device_id, target_app, target_activity)
                                await asyncio.sleep(1.5)  # 优化：减少等待时间，让启动流程智能检测
                                
                                # 清理页面检测缓存
                                ximeng.detector.clear_cache()
                                
                                # 重新处理启动流程
                                file_logger = logging.getLogger(__name__)
                                startup_ok = await ximeng.handle_startup_flow_integrated(
                                    device_id, log_callback=log_callback,
                                    stop_check=self._check_stop_or_pause,
                                    package_name=target_app, activity_name=target_activity,
                                    max_retries=3,
                                    file_logger=file_logger
                                )
                                if not startup_ok:
                                    raise Exception("重新启动失败")
                                
                                has_valid_cache = False
                            else:
                                log_callback("OK 用户ID验证通过，缓存有效！")
                                # 清理页面检测缓存，因为当前已在个人页
                                ximeng.detector.clear_cache()
                        else:
                            # 无法获取用户ID，清理后重新登录
                            await adb.stop_app(device_id, target_app)
                            await asyncio.sleep(0.5)  # 优化：减少等待时间
                            await auto_login.cache_manager.clear_app_login_data(device_id, target_app)
                            await adb.start_app(device_id, target_app, target_activity)
                            await asyncio.sleep(1.5)  # 优化：减少等待时间
                            ximeng.detector.clear_cache()
                            file_logger = logging.getLogger(__name__)
                            startup_ok = await ximeng.handle_startup_flow_integrated(
                                device_id, log_callback=log_callback,
                                stop_check=self._check_stop_or_pause,
                                package_name=target_app, activity_name=target_activity,
                                max_retries=3,
                                file_logger=file_logger
                            )
                            if not startup_ok:
                                raise Exception("重新启动失败")
                            has_valid_cache = False
                    else:
                        has_valid_cache = False
                        ximeng.detector.clear_cache()
                else:
                    has_valid_cache = False
                    ximeng.detector.clear_cache()
            elif has_valid_cache and not enable_profile:
                # 快速签到模式：跳过ID验证，直接使用缓存
                log_callback("快速签到模式：跳过用户ID验证，直接使用缓存")
            
            # 调用完整工作流
            ximeng._stop_check = self._check_stop_or_pause
            
            # 获取流程控制配置
            workflow_config = {
                'enable_login': getattr(self.config, 'workflow_enable_login', True),
                'enable_profile': getattr(self.config, 'workflow_enable_profile', True),
                'enable_checkin': getattr(self.config, 'workflow_enable_checkin', True),
                'enable_transfer': getattr(self.config, 'workflow_enable_transfer', True),
            }
            
            if has_valid_cache:
                # 使用缓存登录
                result = await ximeng.run_full_workflow(device_id, account, skip_login=True, workflow_config=workflow_config)
            else:
                # 正常登录
                result = await ximeng.run_full_workflow(device_id, account, skip_login=False, workflow_config=workflow_config)
            
            # 设置登录方式
            if has_valid_cache:
                result.login_method = "缓存"
            else:
                result.login_method = "正常登录"
            
            # 计算耗时
            duration = time.time() - start_time
            result.duration = duration
            
            # 保存到数据库
            log_callback("保存处理结果到数据库...")
            try:
                account_manager.update_account_result(account.phone, result)
                log_callback("✓ 数据已保存到数据库")
            except Exception as db_error:
                log_callback(f"❌ 数据库保存失败: {db_error}")
                import traceback
                log_callback(f"详细错误: {traceback.format_exc()}")
            
            # 更新表格
            if result.success:
                self.root.after(0, lambda: self._add_result_to_table(result))
                self.root.after(0, lambda: self._update_pending_count())
                log_callback(f"✓ 账号处理完成 (耗时: {round(duration, 3)}秒)")
                log_callback("")  # 空行分隔
                
                # 清除该账号的警告日志（重试成功后）
                self.root.after(0, lambda p=account.phone: self._clear_account_warnings(p))
            else:
                self.root.after(0, lambda: self._update_account_status_in_table(account.phone, "失败"))
                self.root.after(0, lambda: self._update_pending_count())
                log_callback(f"✗ 账号处理失败: {result.error_message}")
                log_callback("")  # 空行分隔
                
                # 记录错误日志
                self.root.after(0, lambda p=account.phone, u=result.user_id or "未知", n=result.nickname or "未知", e=result.error_message: 
                              self._log_error(p, u, n, e))
            
            return result
            
        except Exception as e:
            # P1修复: 完善异常处理，返回包含错误信息的AccountResult对象
            duration = time.time() - start_time
            log_callback(f"✗ 账号处理异常: {e}")
            
            # 创建失败结果
            from .models.models import AccountResult
            result = AccountResult(
                phone=account.phone,
                success=False,
                error_message=str(e),
                duration=duration
            )
            
            # 保存到数据库（重要！即使失败也要保存）
            log_callback("保存失败结果到数据库...")
            try:
                account_manager.update_account_result(account.phone, result)
                log_callback("✓ 失败结果已保存到数据库")
            except Exception as db_error:
                log_callback(f"❌ 数据库保存失败: {db_error}")
            
            self.root.after(0, lambda: self._update_account_status_in_table(account.phone, "失败"))
            self.root.after(0, lambda: self._update_pending_count())
            
            # 记录错误日志
            self.root.after(0, lambda p=account.phone, e=str(e): 
                          self._log_error(p, "未知", "未知", e))
            
            log_callback("")  # 空行分隔
            return result
        
        finally:
            # P1修复: 确保ADB连接被关闭，避免资源泄漏
            try:
                await adb.disconnect(device_id)
                # 只记录到文件日志，不显示在GUI
                file_logger = logging.getLogger(__name__)
                file_logger.info("ADB连接已关闭")
            except Exception as e:
                # 错误仍然显示在GUI
                log_callback(f"⚠️ 关闭ADB连接时出错: {e}")
    
    # ==================== 已废弃的方法 ====================
    # 以下方法已被 run_full_workflow 替代，保留用于参考
    # 新的GUI流程直接调用 ximeng.run_full_workflow(device_id, account)
    
    # ⚠️ 已删除废弃函数 _process_single_account_monitored（2026-02-02）
    # 原因：该函数调用了已删除的 handle_startup_flow_optimized
    # 替代方案：使用 ximeng.run_full_workflow(device_id, account)
    
    async def _save_exception_screenshot(self, adb: ADBBridge, device_id: str, phone: str, exception_type: str):
        """保存异常截图
        
        Args:
            adb: ADB桥接对象
            device_id: 设备ID
            phone: 手机号
            exception_type: 异常类型（用于文件命名）
        """
        try:
            from datetime import datetime
            
            # 创建异常截图目录
            exception_dir = Path("exception_screenshots")
            exception_dir.mkdir(exist_ok=True)
            
            # 截图
            screenshot_data = await adb.screencap(device_id)
            if not screenshot_data:
                return
            
            # 保存截图
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"exception_{phone}_{exception_type}_{timestamp}.png"
            screenshot_path = exception_dir / filename
            
            with open(screenshot_path, 'wb') as f:
                f.write(screenshot_data)
            
            self.root.after(0, lambda: self._log(f"异常截图已保存: {screenshot_path}"))
        except Exception as e:
            self.root.after(0, lambda: self._log(f"保存异常截图失败: {e}"))
    
    def _open_transfer_config(self):
        """打开转账配置窗口"""
        # 检查窗口是否已打开
        if hasattr(self, '_transfer_config_window') and self._transfer_config_window and hasattr(self._transfer_config_window, 'window') and self._transfer_config_window.window.winfo_exists():
            # 窗口已存在，激活它
            self._transfer_config_window.window.lift()
            self._transfer_config_window.window.focus_force()
            return
        
        # 创建新窗口
        self._transfer_config_window = TransferConfigWindow(self.root, self._log, self.accounts_file_var.get(), self)
    
    def _open_workflow_control(self):
        """打开流程控制窗口"""
        # 检查窗口是否已打开
        if hasattr(self, '_workflow_control_window') and self._workflow_control_window and hasattr(self._workflow_control_window, 'window') and self._workflow_control_window.window.winfo_exists():
            # 窗口已存在，激活它
            self._workflow_control_window.window.lift()
            self._workflow_control_window.window.focus_force()
            return
        
        # 显示当前模式
        mode_names = {
            "complete": "完整流程",
            "quick_checkin": "快速签到",
            "login_only": "只登录",
            "transfer_only": "只转账",
            "custom": "自定义"
        }
        current_mode = getattr(self.config, 'workflow_mode', 'complete')
        mode_name = mode_names.get(current_mode, "完整流程")
        self._log(f"当前流程模式: {mode_name}")
        
        # 创建新窗口
        self._workflow_control_window = WorkflowControlWindow(self.root, self)
    
    def _open_transfer_history(self):
        """打开转账历史窗口"""
        # 检查窗口是否已打开
        if hasattr(self, '_transfer_history_window') and self._transfer_history_window and self._transfer_history_window.winfo_exists():
            # 窗口已存在，激活它
            self._transfer_history_window.lift()
            self._transfer_history_window.focus_force()
            return
        
        try:
            from .transfer_history_gui import show_transfer_history
            self._transfer_history_window = show_transfer_history(self.root)
        except Exception as e:
            self._log(f"打开转账历史窗口失败: {e}")
            messagebox.showerror("错误", f"打开转账历史窗口失败:\n{e}")
    
    def _on_auto_transfer_changed(self, state: bool):
        """自动转账开关状态改变回调
        
        Args:
            state: True=开启, False=关闭
        """
        from .transfer_config import get_transfer_config
        transfer_config = get_transfer_config()
        
        # 保存到配置
        transfer_config.set_enabled(state)
        
        # 记录日志
        status_text = "开启" if state else "关闭"
        self._log(f"自动转账已{status_text}")
        
        # 同步更新转账配置窗口（如果打开的话）
        self._sync_transfer_config_window(state)
    
    def _on_scheduled_run_changed(self, state: bool):
        """定时运行开关状态改变回调
        
        Args:
            state: True=开启, False=关闭
        """
        self.scheduled_run_enabled.set(state)
        
        # 更新scheduled_run_time
        self._update_scheduled_time()
        
        # 保存配置
        self._auto_save_config()
        
        # 记录日志
        status_text = "开启" if state else "关闭"
        scheduled_time = self.scheduled_run_time.get()
        self._log(f"定时运行已{status_text}（运行时间: {scheduled_time}）")
        
        if state:
            # 启动定时检查线程
            self._start_schedule_check_thread()
        else:
            # 停止定时检查线程
            self._stop_schedule_check_thread()
    
    def _on_scheduled_time_changed(self):
        """定时运行时间改变回调"""
        self._update_scheduled_time()
        self._auto_save_config()
    
    def _update_scheduled_time(self):
        """更新scheduled_run_time变量（从小时和分钟变量合成）"""
        try:
            hour = int(self.scheduled_hour_var.get())
            minute = int(self.scheduled_minute_var.get())
            
            # 验证范围
            if 0 <= hour <= 23 and 0 <= minute <= 59:
                self.scheduled_run_time.set(f"{hour:02d}:{minute:02d}")
            else:
                # 超出范围，重置为默认值
                self.scheduled_hour_var.set("08")
                self.scheduled_minute_var.set("00")
                self.scheduled_run_time.set("08:00")
        except:
            # 解析失败，重置为默认值
            self.scheduled_hour_var.set("08")
            self.scheduled_minute_var.set("00")
            self.scheduled_run_time.set("08:00")
    
    def _start_schedule_check_thread(self):
        """启动定时检查线程"""
        if self.schedule_check_thread and self.schedule_check_thread.is_alive():
            return  # 线程已在运行
        
        self.schedule_check_thread = threading.Thread(target=self._schedule_check_loop, daemon=True)
        self.schedule_check_thread.start()
        self._log("定时检查线程已启动")
    
    def _stop_schedule_check_thread(self):
        """停止定时检查线程"""
        # 线程会在下次检查时自动退出（检查scheduled_run_enabled状态）
        self._log("定时检查线程将在下次检查时停止")
    
    def _schedule_check_loop(self):
        """定时检查循环（在后台线程中运行）"""
        import time
        from datetime import datetime, date
        
        while True:
            try:
                # 检查是否已禁用定时运行
                if not self.scheduled_run_enabled.get():
                    break
                
                # 获取当前时间
                now = datetime.now()
                current_time = now.strftime("%H:%M")
                current_date = now.date()
                
                # 获取设定的运行时间
                scheduled_time = self.scheduled_run_time.get()
                
                # 更新状态显示
                self.root.after(0, lambda: self.scheduled_status_var.set(f"下次运行: 今天 {scheduled_time}"))
                
                # 检查是否到达运行时间
                if current_time == scheduled_time:
                    # 检查今天是否已经运行过
                    if self.last_scheduled_run_date != current_date:
                        # 触发自动运行
                        self.root.after(0, self._trigger_scheduled_run)
                        # 记录运行日期
                        self.last_scheduled_run_date = current_date
                        # 等待60秒，避免重复触发
                        time.sleep(60)
                
                # 每30秒检查一次
                time.sleep(30)
                
            except Exception as e:
                print(f"定时检查线程异常: {e}")
                time.sleep(60)  # 出错后等待1分钟再继续
    
    def _trigger_scheduled_run(self):
        """触发定时运行"""
        self._log("=" * 50)
        self._log(f"⏰ 定时运行触发 - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        self._log("=" * 50)
        
        # 检查是否正在运行
        if hasattr(self, 'automation_thread') and self.automation_thread and self.automation_thread.is_alive():
            self._log("⚠️ 任务正在运行中，跳过本次定时运行")
            return
        
        # 触发开始运行
        self._start_automation()
    
    def _toggle_auto_transfer(self):
        """切换自动转账开关（保留用于兼容）"""
        current_state = self.auto_transfer_switch.get_state()
        self.auto_transfer_switch.toggle()
    
    def _update_auto_transfer_button(self):
        """更新自动转账按钮显示（已废弃，保留用于兼容）"""
        pass
    
    def _sync_transfer_config_window(self, state: bool):
        """同步转账配置窗口的复选框状态
        
        Args:
            state: True=开启, False=关闭
        """
        # 遍历所有子窗口，找到转账配置窗口
        for child in self.root.winfo_children():
            if isinstance(child, tk.Toplevel) and child.title() == "转账配置":
                # 找到窗口，更新其中的复选框
                try:
                    # 通过窗口对象找到对应的TransferConfigWindow实例
                    # 这需要在创建窗口时保存引用
                    if hasattr(self, '_transfer_config_window'):
                        self._transfer_config_window.enabled_var.set(state)
                except:
                    pass
                break
    
    def _open_history_results(self):
        """打开历史结果查看窗口"""
        # 检查窗口是否已打开
        if hasattr(self, '_history_results_window') and self._history_results_window and hasattr(self._history_results_window, 'window') and self._history_results_window.window.winfo_exists():
            # 窗口已存在，激活它
            self._history_results_window.window.lift()
            self._history_results_window.window.focus_force()
            return
        
        # 创建新窗口
        self._history_results_window = HistoryResultsWindow(self.root, self._log, self)
    
    def _open_user_management(self):
        """打开用户管理窗口"""
        # 检查窗口是否已打开
        if hasattr(self, '_user_management_window') and self._user_management_window and hasattr(self._user_management_window, 'dialog') and self._user_management_window.dialog.winfo_exists():
            # 窗口已存在，激活它
            self._user_management_window.dialog.lift()
            self._user_management_window.dialog.focus_force()
            return
        
        # 创建新窗口
        from .user_management_gui import UserManagementDialog
        self._user_management_window = UserManagementDialog(self.root, self._log)
    
    def _auto_check_new_models(self):
        """自动检查并注册新模型和新页面类型（启动时调用）"""
        try:
            from .auto_model_registry import check_and_register_new_models
            from .auto_page_type_registry import check_and_register_page_types
            
            # 在后台线程中执行，避免阻塞GUI
            def check_thread():
                # 检查YOLO模型
                yolo_result = check_and_register_new_models(log_callback=None)
                
                # 检查页面类型
                page_result = check_and_register_page_types(log_callback=None)
                
                # 如果发现新内容，在主线程中显示提示
                if yolo_result['new_models_count'] > 0 or page_result['new_types_count'] > 0:
                    self.root.after(0, lambda: self._show_auto_register_notification(yolo_result, page_result))
            
            thread = threading.Thread(target=check_thread, daemon=True)
            thread.start()
            
        except Exception as e:
            # 静默失败，不影响程序启动
            print(f"自动检测失败: {e}")
    
    def _show_auto_register_notification(self, yolo_result: dict, page_result: dict):
        """显示自动注册通知"""
        has_yolo = yolo_result['registered_count'] > 0
        has_page = page_result['registered_count'] > 0
        
        if has_yolo or has_page:
            self._log("")
            self._log("=" * 60)
            self._log("🎉 自动注册完成")
            self._log("=" * 60)
            
            if has_yolo:
                self._log(f"✅ YOLO模型: {yolo_result['registered_count']} 个")
                self._log(f"   版本: {yolo_result['version']}")
            
            if has_page:
                self._log(f"✅ 页面类型: {page_result['registered_count']} 个")
                self._log(f"   新类型: {', '.join(page_result.get('new_types', []))}")
                self._log("")
                self._log("💡 提示: 页面类型已自动注册，重启程序即可生效")
                self._log("   无需手动修改代码！")
            
            self._log("=" * 60)
            self._log("")
    
    def _register_new_models(self):
        """手动注册新模型（按钮点击）"""
        try:
            from .auto_model_registry import check_and_register_new_models
            from .auto_page_type_registry import check_and_register_page_types
            
            self._log("")
            self._log("=" * 60)
            self._log("🔍 正在扫描新模型和新页面类型...")
            self._log("=" * 60)
            
            # 在后台线程中执行
            def register_thread():
                # 1. 注册YOLO模型
                self.root.after(0, lambda: self._log(""))
                self.root.after(0, lambda: self._log("📦 步骤1: 扫描YOLO模型..."))
                
                yolo_result = check_and_register_new_models(
                    log_callback=lambda msg: self.root.after(0, lambda m=msg: self._log(m))
                )
                
                # 2. 注册页面类型
                self.root.after(0, lambda: self._log(""))
                self.root.after(0, lambda: self._log("📦 步骤2: 扫描页面类型..."))
                
                page_result = check_and_register_page_types(
                    log_callback=lambda msg: self.root.after(0, lambda m=msg: self._log(m))
                )
                
                # 显示结果
                self.root.after(0, lambda: self._log(""))
                self.root.after(0, lambda: self._log("=" * 60))
                self.root.after(0, lambda: self._log("📊 注册结果汇总"))
                self.root.after(0, lambda: self._log("=" * 60))
                
                # YOLO模型结果
                if yolo_result['new_models_count'] == 0:
                    self.root.after(0, lambda: self._log("✅ YOLO模型: 未发现新模型"))
                elif yolo_result['registered_count'] > 0:
                    self.root.after(0, lambda: self._log(f"✅ YOLO模型: 成功注册 {yolo_result['registered_count']} 个"))
                    self.root.after(0, lambda: self._log(f"   版本已更新: {yolo_result['version']}"))
                
                if yolo_result['errors']:
                    self.root.after(0, lambda: self._log(f"⚠️ YOLO模型: {len(yolo_result['errors'])} 个注册失败"))
                
                # 页面类型结果
                if page_result['new_types_count'] == 0:
                    self.root.after(0, lambda: self._log("✅ 页面类型: 未发现新类型"))
                elif page_result['registered_count'] > 0:
                    self.root.after(0, lambda: self._log(f"✅ 页面类型: 成功注册 {page_result['registered_count']} 个"))
                    self.root.after(0, lambda: self._log(f"   新类型: {', '.join(page_result.get('new_types', []))}"))
                
                if page_result['errors']:
                    self.root.after(0, lambda: self._log(f"⚠️ 页面类型: {len(page_result['errors'])} 个注册失败"))
                
                # 总结
                self.root.after(0, lambda: self._log(""))
                self.root.after(0, lambda: self._log("=" * 60))
                
                total_registered = yolo_result['registered_count'] + page_result['registered_count']
                if total_registered > 0:
                    self.root.after(0, lambda: self._log(f"🎉 总计注册: {total_registered} 项"))
                    self.root.after(0, lambda: self._log(""))
                    self.root.after(0, lambda: self._log("💡 提示："))
                    if yolo_result['registered_count'] > 0:
                        self.root.after(0, lambda: self._log("  • YOLO模型将在下次启动程序时生效"))
                    if page_result['registered_count'] > 0:
                        self.root.after(0, lambda: self._log("  • 页面类型已自动注册，重启程序即可生效"))
                        self.root.after(0, lambda: self._log("  • 无需手动修改代码！"))
                else:
                    self.root.after(0, lambda: self._log("✅ 所有模型和类型都已是最新"))
                
                self.root.after(0, lambda: self._log("=" * 60))
                self.root.after(0, lambda: self._log(""))
            
            thread = threading.Thread(target=register_thread, daemon=True)
            thread.start()
            
        except Exception as e:
            self._log(f"❌ 注册失败: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def _open_window_arranger(self):
        """打开窗口排列对话框"""
        # 检查窗口是否已打开
        if hasattr(self, '_window_arranger_dialog') and self._window_arranger_dialog and hasattr(self._window_arranger_dialog, 'dialog') and self._window_arranger_dialog.dialog.winfo_exists():
            # 窗口已存在，激活它
            self._window_arranger_dialog.dialog.lift()
            self._window_arranger_dialog.dialog.focus_force()
            return
        
        # 创建新窗口
        self._window_arranger_dialog = WindowArrangerDialog(self.root, self._log)
    
    def _on_automation_complete(self):
        """自动化完成回调"""
        self.is_running = False
        self.is_paused = False
        self.start_btn.config(state=tk.NORMAL)
        self.pause_btn.config(state=tk.DISABLED, text="⏸ 暂停")
        self.stop_btn.config(state=tk.DISABLED)
        
        # 先更新统计（从表格重新统计，确保数据准确）
        self._update_stats_from_table()
        
        # 然后更新状态
        self.status_var.set("已完成")
        self._log("自动化任务完成")
        
        # 历史记录已在执行过程中实时保存
        try:
            self._save_to_history()
            self._log("✓ 所有账号的结果已在执行过程中实时保存到数据库")
        except Exception as e:
            self._log(f"⚠️ 保存历史记录失败: {e}")
    
    def run(self):
        """运行 GUI"""
        try:
            self.root.mainloop()
        except KeyboardInterrupt:
            self._on_closing()
        except Exception as e:
            print(f"GUI运行错误: {e}")
            self._on_closing()
    
    def _on_closing(self):
        """安全关闭窗口"""
        try:
            # 如果正在运行，先停止
            if self.is_running:
                self._log("正在停止运行中的任务...")
                self.stop_event.set()
                
                # 取消所有待处理的Future对象
                if self.pending_futures:
                    for future in self.pending_futures:
                        try:
                            future.cancel()
                        except:
                            pass
                    self.pending_futures.clear()
                
                # 等待线程池关闭（最多等待5秒）
                if self.executor:
                    try:
                        # 先尝试优雅关闭
                        self.executor.shutdown(wait=True, cancel_futures=True)
                    except Exception as e:
                        print(f"线程池关闭失败: {e}")
                        # 强制关闭
                        try:
                            self.executor.shutdown(wait=False)
                        except:
                            pass
                
                # 短暂等待让线程有机会清理
                time.sleep(0.5)
            
            # 保存配置
            try:
                self._save_config()
            except Exception as e:
                print(f"保存配置失败: {e}")
            
            # 保存历史记录
            try:
                self._save_to_history()
            except Exception as e:
                print(f"保存历史记录失败: {e}")
            
            # 清理资源
            try:
                # 清空表格
                for item in self.results_tree.get_children():
                    self.results_tree.delete(item)
                
                # 清空日志
                self.log_text.config(state=tk.NORMAL)
                self.log_text.delete(1.0, tk.END)
                self.log_text.config(state=tk.DISABLED)
                
                # 清空变量
                self.checked_items.clear()
                self.selection_history.clear()
                self.all_logs.clear()
                self.instance_pool.clear()
                
            except Exception as e:
                print(f"清理资源失败: {e}")
            
            # 销毁窗口（使用destroy而不是quit+destroy）
            try:
                # 只调用destroy，不调用quit
                # quit会导致mainloop退出，可能引起销毁顺序问题
                self.root.destroy()
            except Exception as e:
                print(f"销毁窗口失败: {e}")
                
        except Exception as e:
            print(f"关闭窗口时出错: {e}")
            import traceback
            traceback.print_exc()
            # 强制退出
            try:
                self.root.destroy()
            except:
                pass
            import sys
            sys.exit(0)


class TransferConfigWindow:
    """转账配置窗口"""
    
    def __init__(self, parent, log_callback, accounts_file, gui_instance=None):
        """初始化转账配置窗口
        
        Args:
            parent: 父窗口
            log_callback: 日志回调函数
            accounts_file: 账号文件路径
            gui_instance: AutomationGUI实例（用于同步主界面开关）
        """
        self.parent = parent
        self.log = log_callback
        self.accounts_file = accounts_file
        self.gui_instance = gui_instance
        
        # 加载转账配置
        from .transfer_config import get_transfer_config
        self.transfer_config = get_transfer_config()
        
        # 创建窗口
        self.window = tk.Toplevel(parent)
        self.window.title("转账配置")
        self.window.geometry("800x700")
        self.window.resizable(True, True)
        
        # 先隐藏窗口，避免白屏
        self.window.withdraw()
        
        # 设置窗口关闭协议
        self.window.protocol("WM_DELETE_WINDOW", self._on_closing)
        
        # 账号列表（从账号缓存中读取）
        self.accounts = []
        self._load_accounts()
        
        # 创建界面
        self._create_widgets()
        
        # 居中显示（不使用grab_set，允许主窗口操作）
        self.window.transient(parent)
        
        # 所有内容准备完成后再显示窗口
        self.window.deiconify()
        self.window.lift()
        self.window.focus_force()
    
    def _load_accounts(self):
        """加载账号列表（从数据库）"""
        try:
            db = LocalDatabase()
            
            # 从数据库加载历史记录
            records = db.get_all_history_records()
            
            if not records:
                self.log("未找到历史记录，尝试从账号缓存加载")
                self._load_accounts_from_cache()
                return
            
            self.log(f"从数据库加载历史记录")
            
            # 转换为账号列表（注意：数据库返回的是中文字段名）
            self.accounts = []
            phone_dict = {}  # 用于去重，保留最新记录
            
            for record in records:
                # 使用中文字段名获取数据
                user_id = record.get('用户ID', '')
                nickname = record.get('昵称', '未知')
                phone = record.get('手机号', '')
                balance = record.get('余额(元)', 0.0)
                
                # 处理余额：可能是字符串或数字
                if isinstance(balance, str):
                    try:
                        balance = float(balance) if balance and balance != 'N/A' and balance != '-' else 0.0
                    except:
                        balance = 0.0
                elif balance is None or balance == '-':
                    balance = 0.0
                
                if user_id and phone:
                    # 去重：保留最新的记录
                    if phone not in phone_dict:
                        phone_dict[phone] = {
                            'phone': phone,
                            'user_id': user_id,
                            'nickname': nickname,
                            'balance': balance
                        }
            
            # 转换为列表
            self.accounts = list(phone_dict.values())
            
            self.log(f"从数据库加载了 {len(self.accounts)} 个账号")
            
        except Exception as e:
            self.log(f"从数据库加载账号失败: {e}，尝试从账号缓存加载")
            import traceback
            traceback.print_exc()
            self._load_accounts_from_cache()
    
    def _load_accounts_from_cache(self):
        """从账号缓存加载账号列表（备用方法）"""
        try:
            from .account_cache import get_account_cache
            cache = get_account_cache()
            
            # 获取所有账号的用户ID
            self.accounts = []
            for phone, data in cache._cache.items():
                user_id = data.get('user_id')
                nickname = data.get('nickname', '未知')
                if user_id:
                    self.accounts.append({
                        'phone': phone,
                        'user_id': user_id,
                        'nickname': nickname,
                        'balance': 0.0  # 缓存中没有余额信息
                    })
            self.log(f"从账号缓存加载了 {len(self.accounts)} 个账号（无余额信息）")
        except Exception as e:
            self.log(f"从账号缓存加载失败: {e}")
    
    def _create_widgets(self):
        """创建界面组件"""
        main_frame = ttk.Frame(self.window, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # === 顶部配置区域 ===
        config_frame = ttk.LabelFrame(main_frame, text="转账设置", padding="10")
        config_frame.pack(fill=tk.X, pady=(0, 10))
        
        # 启用转账
        self.enabled_var = tk.BooleanVar(value=self.transfer_config.enabled)
        ttk.Checkbutton(config_frame, text="启用自动转账", variable=self.enabled_var,
                       command=self._on_enabled_changed).pack(anchor=tk.W, pady=(0, 5))
        
        # 收款人选择策略
        strategy_frame = ttk.Frame(config_frame)
        strategy_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(strategy_frame, text="收款人选择策略:").pack(side=tk.LEFT)
        
        # 获取当前策略
        current_strategy = getattr(self.transfer_config, 'recipient_selection_strategy', 'rotation')
        
        # 策略选项
        strategy_options = [
            ("轮询（平均分配）", "rotation"),
            ("随机选择", "random")
        ]
        
        self.recipient_strategy_var = tk.StringVar(value=current_strategy)
        
        for display_name, strategy_value in strategy_options:
            ttk.Radiobutton(
                strategy_frame,
                text=display_name,
                variable=self.recipient_strategy_var,
                value=strategy_value,
                command=self._on_recipient_strategy_changed
            ).pack(side=tk.LEFT, padx=(10, 0))
        
        # 策略说明
        strategy_info_label = ttk.Label(
            config_frame,
            text="说明：轮询=按顺序循环选择收款人，确保负载均衡 | 随机=随机选择收款人，增加不可预测性",
            foreground="green"
        )
        strategy_info_label.pack(anchor=tk.W, pady=(0, 5))
        
        # 多级转账设置
        multi_level_frame = ttk.Frame(config_frame)
        multi_level_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.multi_level_enabled_var = tk.BooleanVar(value=getattr(self.transfer_config, 'multi_level_enabled', False))
        ttk.Checkbutton(multi_level_frame, text="启用多级转账", variable=self.multi_level_enabled_var,
                       command=self._on_multi_level_changed).pack(side=tk.LEFT)
        
        ttk.Label(multi_level_frame, text="  最大级数:").pack(side=tk.LEFT, padx=(10, 0))
        self.max_transfer_level_var = tk.IntVar(value=getattr(self.transfer_config, 'max_transfer_level', 1))
        ttk.Spinbox(multi_level_frame, from_=1, to=3, textvariable=self.max_transfer_level_var, 
                   width=5, command=self._on_level_changed).pack(side=tk.LEFT, padx=(5, 0))
        ttk.Label(multi_level_frame, text="级 (收款账号也会继续转账)", foreground="orange").pack(side=tk.LEFT, padx=(5, 0))
        
        # 起步金额（最小转账金额）
        threshold_frame = ttk.Frame(config_frame)
        threshold_frame.pack(fill=tk.X, pady=(0, 5))
        ttk.Label(threshold_frame, text="起步金额:").pack(side=tk.LEFT)
        self.min_transfer_amount_var = tk.DoubleVar(value=getattr(self.transfer_config, 'min_transfer_amount', 30.0))
        ttk.Entry(threshold_frame, textvariable=self.min_transfer_amount_var, width=10).pack(side=tk.LEFT, padx=(5, 5))
        ttk.Label(threshold_frame, text="元 (余额达到此金额才开始转账)").pack(side=tk.LEFT)
        ttk.Button(threshold_frame, text="保存", command=self._save_threshold).pack(side=tk.LEFT, padx=(10, 0))
        
        # 最小保留余额
        balance_frame = ttk.Frame(config_frame)
        balance_frame.pack(fill=tk.X, pady=(0, 5))
        ttk.Label(balance_frame, text="保留余额:").pack(side=tk.LEFT)
        self.min_balance_var = tk.DoubleVar(value=self.transfer_config.min_balance)
        ttk.Entry(balance_frame, textvariable=self.min_balance_var, width=10).pack(side=tk.LEFT, padx=(5, 5))
        ttk.Label(balance_frame, text="元 (转账后保留的余额)").pack(side=tk.LEFT)
        ttk.Button(balance_frame, text="保存", command=self._save_min_balance).pack(side=tk.LEFT, padx=(10, 0))
        
        # 说明文字
        info_label = ttk.Label(config_frame, 
                              text="说明：上方为待转账号，下方为收款账户。当余额 >= (起步金额 + 保留余额) 时才会转账",
                              foreground="blue")
        info_label.pack(anchor=tk.W, pady=(5, 0))
        
        # 示例说明
        example_label = ttk.Label(config_frame,
                                 text="示例：起步金额30元 + 保留余额5元 = 余额需>=35元才转账，转账后保留5元；保留余额0元 = 达到起步金额则全部转出",
                                 foreground="gray")
        example_label.pack(anchor=tk.W, pady=(2, 0))
        
        # 多级转账说明
        multi_level_info = ttk.Label(config_frame,
                                    text="多级转账：收款账号收到钱后也会继续转账。例如2级：A→B→C，3级：A→B→C→D",
                                    foreground="purple")
        multi_level_info.pack(anchor=tk.W, pady=(2, 0))
        
        # === 账号列表区域（上下分栏）===
        lists_frame = ttk.Frame(main_frame)
        lists_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # 上方：待转账号列表
        top_frame = ttk.LabelFrame(lists_frame, text="待转账号", padding="10")
        top_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True, pady=(0, 5))
        
        # 创建待转账号Treeview
        columns = ("phone", "user_id", "nickname", "balance", "owner")
        self.transfer_tree = ttk.Treeview(top_frame, columns=columns, show="headings", height=8)
        
        self.transfer_tree.heading("phone", text="手机号")
        self.transfer_tree.heading("user_id", text="用户ID")
        self.transfer_tree.heading("nickname", text="昵称")
        self.transfer_tree.heading("balance", text="余额(元)")
        self.transfer_tree.heading("owner", text="管理员")
        
        self.transfer_tree.column("phone", width=120, anchor=tk.CENTER)
        self.transfer_tree.column("user_id", width=100, anchor=tk.CENTER)
        self.transfer_tree.column("nickname", width=150, anchor=tk.CENTER)
        self.transfer_tree.column("balance", width=100, anchor=tk.CENTER)
        self.transfer_tree.column("owner", width=80, anchor=tk.CENTER)
        
        # 添加滚动条
        transfer_scrollbar = ttk.Scrollbar(top_frame, orient=tk.VERTICAL, command=self.transfer_tree.yview)
        self.transfer_tree.configure(yscrollcommand=transfer_scrollbar.set)
        
        self.transfer_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        transfer_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 绑定双击事件
        self.transfer_tree.bind("<Double-Button-1>", self._on_transfer_tree_double_click)
        
        # 待转账号按钮（改为上下排列）
        transfer_btn_frame = ttk.Frame(top_frame)
        transfer_btn_frame.pack(fill=tk.X, pady=(5, 0))
        ttk.Button(transfer_btn_frame, text="↓ 移到收款", command=self._move_to_recipient, width=15).pack(pady=(0, 3))
        ttk.Button(transfer_btn_frame, text="↓ 全部移到收款", command=self._move_all_to_recipient, width=15).pack()
        
        # 下方：收款账户列表
        self.bottom_frame = ttk.LabelFrame(lists_frame, text="收款账户", padding="10")
        self.bottom_frame.pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True, pady=(5, 0))
        
        # 存储收款账户的容器（会根据多级转账设置动态切换）
        self.recipient_container = ttk.Frame(self.bottom_frame)
        self.recipient_container.pack(fill=tk.BOTH, expand=True)
        
        # 创建单级和多级两种界面
        self.recipient_trees = {}  # 存储每个级别的树形控件
        self.recipient_notebook = None  # Notebook控件（多级模式）
        self.recipient_tree = None  # 单个树形控件（单级模式）
        
        # 根据当前配置创建界面
        self._create_recipient_interface()
        
        # 填充数据
        self._refresh_trees()
        
        # === 底部按钮区域 ===
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X)
        
        ttk.Button(button_frame, text="刷新账号列表", command=self._refresh_accounts).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(button_frame, text="关闭", command=self.window.destroy).pack(side=tk.RIGHT)
    
    def _create_recipient_interface(self):
        """创建收款账户界面（根据多级转账设置动态创建）"""
        # 清空容器
        for widget in self.recipient_container.winfo_children():
            widget.destroy()
        
        columns = ("phone", "user_id", "nickname", "balance", "owner")
        
        if self.multi_level_enabled_var.get():
            # 多级模式：创建Notebook标签页
            self.bottom_frame.config(text="收款账户（多级配置）")
            
            self.recipient_notebook = ttk.Notebook(self.recipient_container)
            self.recipient_notebook.pack(fill=tk.BOTH, expand=True)
            
            self.recipient_trees = {}
            
            for level in [1, 2, 3]:
                # 创建标签页框架
                tab_frame = ttk.Frame(self.recipient_notebook)
                self.recipient_notebook.add(tab_frame, text=f"{level}级收款账号")
                
                # 创建树形控件
                tree = ttk.Treeview(tab_frame, columns=columns, show="headings", height=6)
                
                tree.heading("phone", text="手机号")
                tree.heading("user_id", text="用户ID")
                tree.heading("nickname", text="昵称")
                tree.heading("balance", text="余额(元)")
                tree.heading("owner", text="管理员")
                
                tree.column("phone", width=120, anchor=tk.CENTER)
                tree.column("user_id", width=100, anchor=tk.CENTER)
                tree.column("nickname", width=150, anchor=tk.CENTER)
                tree.column("balance", width=100, anchor=tk.CENTER)
                tree.column("owner", width=80, anchor=tk.CENTER)
                
                # 添加滚动条
                scrollbar = ttk.Scrollbar(tab_frame, orient=tk.VERTICAL, command=tree.yview)
                tree.configure(yscrollcommand=scrollbar.set)
                
                tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
                scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
                
                # 绑定双击事件
                tree.bind("<Double-Button-1>", lambda e, lv=level: self._on_recipient_tree_double_click(e, lv))
                
                # 存储树形控件
                self.recipient_trees[level] = tree
                
                # 按钮框架
                btn_frame = ttk.Frame(tab_frame)
                btn_frame.pack(fill=tk.X, pady=(5, 0))
                ttk.Button(btn_frame, text="↑ 移到待转", 
                          command=lambda lv=level: self._move_to_transfer(lv), width=15).pack(pady=(0, 3))
                ttk.Button(btn_frame, text="↑ 全部移到待转", 
                          command=lambda lv=level: self._move_all_to_transfer(lv), width=15).pack(pady=(0, 3))
                ttk.Button(btn_frame, text="➕ 手动添加", 
                          command=lambda lv=level: self._manual_add_recipient(lv), width=15).pack()
            
            # 保持对第一级的引用
            self.recipient_tree = self.recipient_trees[1]
        else:
            # 单级模式：只显示1级收款账号
            self.bottom_frame.config(text="收款账户")
            
            # 创建树形控件
            self.recipient_tree = ttk.Treeview(self.recipient_container, columns=columns, show="headings", height=8)
            
            self.recipient_tree.heading("phone", text="手机号")
            self.recipient_tree.heading("user_id", text="用户ID")
            self.recipient_tree.heading("nickname", text="昵称")
            self.recipient_tree.heading("balance", text="余额(元)")
            self.recipient_tree.heading("owner", text="管理员")
            
            self.recipient_tree.column("phone", width=120, anchor=tk.CENTER)
            self.recipient_tree.column("user_id", width=100, anchor=tk.CENTER)
            self.recipient_tree.column("nickname", width=150, anchor=tk.CENTER)
            self.recipient_tree.column("balance", width=100, anchor=tk.CENTER)
            self.recipient_tree.column("owner", width=80, anchor=tk.CENTER)
            
            # 添加滚动条
            scrollbar = ttk.Scrollbar(self.recipient_container, orient=tk.VERTICAL, command=self.recipient_tree.yview)
            self.recipient_tree.configure(yscrollcommand=scrollbar.set)
            
            self.recipient_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            # 绑定双击事件
            self.recipient_tree.bind("<Double-Button-1>", lambda e: self._on_recipient_tree_double_click(e, 1))
            
            # 存储到字典中（保持一致性）
            self.recipient_trees = {1: self.recipient_tree}
            
            # 按钮框架
            btn_frame = ttk.Frame(self.recipient_container)
            btn_frame.pack(fill=tk.X, pady=(5, 0))
            ttk.Button(btn_frame, text="↑ 移到待转", 
                      command=lambda: self._move_to_transfer(1), width=15).pack(pady=(0, 3))
            ttk.Button(btn_frame, text="↑ 全部移到待转", 
                      command=lambda: self._move_all_to_transfer(1), width=15).pack(pady=(0, 3))
            ttk.Button(btn_frame, text="➕ 手动添加", 
                      command=lambda: self._manual_add_recipient(1), width=15).pack()
    
    def _refresh_trees(self):
        """刷新所有树形列表"""
        self.log(f"开始刷新树形列表，共有 {len(self.accounts)} 个账号")
        
        # 从数据库加载账号管理员信息
        from .local_db import LocalDatabase
        db = LocalDatabase()
        
        # 清空待转账号列表
        for item in self.transfer_tree.get_children():
            self.transfer_tree.delete(item)
        
        # 清空收款账号列表
        if self.multi_level_enabled_var.get():
            # 多级模式：清空所有级别
            for level in [1, 2, 3]:
                if level in self.recipient_trees:
                    tree = self.recipient_trees[level]
                    for item in tree.get_children():
                        tree.delete(item)
        else:
            # 单级模式：只清空1级
            if self.recipient_tree:
                for item in self.recipient_tree.get_children():
                    self.recipient_tree.delete(item)
        
        # 获取所有级别的收款账号ID
        all_recipient_ids = set()
        if self.multi_level_enabled_var.get():
            # 多级模式：获取所有级别
            for level in [1, 2, 3]:
                recipients = self.transfer_config.get_recipients(level)
                all_recipient_ids.update(recipients)
                self.log(f"{level}级收款账号: {len(recipients)} 个")
        else:
            # 单级模式：只获取1级
            recipients = self.transfer_config.get_recipients(1)
            all_recipient_ids.update(recipients)
            self.log(f"1级收款账号: {len(recipients)} 个")
        
        # 创建账号ID到账号信息的映射
        account_map = {acc['user_id']: acc for acc in self.accounts}
        
        # 分类账号
        transfer_count = 0
        recipient_count = 0
        processed_ids = set()  # 记录已处理的账号ID
        
        # 导入用户管理器，用于检查账号是否有收款ID
        from .user_manager import UserManager
        user_manager = UserManager()
        
        # 先处理历史记录中的账号
        for account in self.accounts:
            user_id = account['user_id']
            phone = account['phone']
            balance = account.get('balance', 0.0)
            balance_str = f"{balance:.2f}"
            
            # 从数据库读取管理员（而不是实时查询user_manager）
            owner_name = "-"
            try:
                records = db.get_history_records(phone, limit=1)
                if records and records[0].get('owner'):
                    owner_name = records[0].get('owner')
            except:
                pass
            
            values = (phone, user_id, account['nickname'], balance_str, owner_name)
            
            processed_ids.add(user_id)
            
            # 检查账号属于哪个级别
            is_recipient = False
            if self.multi_level_enabled_var.get():
                # 多级模式：检查所有级别
                for level in [1, 2, 3]:
                    if user_id in self.transfer_config.get_recipients(level):
                        # 添加到对应级别的收款账号列表
                        if level in self.recipient_trees:
                            self.recipient_trees[level].insert("", tk.END, values=values)
                            recipient_count += 1
                        is_recipient = True
                        break
            else:
                # 单级模式：只检查1级
                if user_id in self.transfer_config.get_recipients(1):
                    if self.recipient_tree:
                        self.recipient_tree.insert("", tk.END, values=values)
                        recipient_count += 1
                    is_recipient = True
            
            if not is_recipient:
                # 检查账号是否有收款ID（通过用户管理）
                has_recipient_id = False
                try:
                    user = user_manager.get_account_user(phone)
                    if user and user.enabled:
                        # 检查管理员是否配置了收款人
                        if user.transfer_recipients and len(user.transfer_recipients) > 0:
                            has_recipient_id = True
                        # 或者管理员自己的ID也算有收款ID
                        elif user.user_id:
                            has_recipient_id = True
                except:
                    pass
                
                # 只有没有收款ID的账号才显示在待转账号列表中
                if not has_recipient_id:
                    self.transfer_tree.insert("", tk.END, values=values)
                    transfer_count += 1
        
        # 处理手动添加的收款账号（不在历史记录中的）
        if self.multi_level_enabled_var.get():
            # 多级模式：检查所有级别
            for level in [1, 2, 3]:
                for user_id in self.transfer_config.get_recipients(level):
                    if user_id not in processed_ids:
                        # 这是手动添加的账号，不在历史记录中
                        values = ("未知", user_id, "手动添加", "N/A", "-")
                        if level in self.recipient_trees:
                            self.recipient_trees[level].insert("", tk.END, values=values)
                            recipient_count += 1
                        processed_ids.add(user_id)
        else:
            # 单级模式：只检查1级
            for user_id in self.transfer_config.get_recipients(1):
                if user_id not in processed_ids:
                    # 这是手动添加的账号，不在历史记录中
                    values = ("未知", user_id, "手动添加", "N/A", "-")
                    if self.recipient_tree:
                        self.recipient_tree.insert("", tk.END, values=values)
                        recipient_count += 1
                    processed_ids.add(user_id)
        
        self.log(f"刷新完成：待转账号 {transfer_count} 个，收款账号 {recipient_count} 个")
    
    def _move_to_recipient(self):
        """将选中的账号移到收款账户"""
        selection = self.transfer_tree.selection()
        
        # 调试：打印选中的项目数量
        print(f"[DEBUG] _move_to_recipient: selection count = {len(selection)}")
        
        if not selection:
            messagebox.showwarning("提示", "请先选择要移动的账号")
            return
        
        # 确定目标级别
        if self.multi_level_enabled_var.get() and self.recipient_notebook:
            # 多级模式：获取当前选中的标签页级别
            current_tab = self.recipient_notebook.index(self.recipient_notebook.select())
            level = current_tab + 1
        else:
            # 单级模式：固定为1级
            level = 1
        
        # 调试：打印要移动的账号
        moved_phones = []
        for item in selection:
            values = self.transfer_tree.item(item, "values")
            user_id = values[1]
            phone = values[0]
            moved_phones.append(phone)
            self.transfer_config.add_recipient(user_id, level)
        
        print(f"[DEBUG] _move_to_recipient: moved phones = {moved_phones}")
        
        self._refresh_trees()
        self.log(f"已将 {len(selection)} 个账号移到 {level}级收款账户")
    
    def _move_all_to_recipient(self):
        """将所有待转账号移到收款账户"""
        # 确定目标级别
        if self.multi_level_enabled_var.get() and self.recipient_notebook:
            # 多级模式：获取当前选中的标签页级别
            current_tab = self.recipient_notebook.index(self.recipient_notebook.select())
            level = current_tab + 1
        else:
            # 单级模式：固定为1级
            level = 1
        
        count = 0
        for item in self.transfer_tree.get_children():
            values = self.transfer_tree.item(item, "values")
            user_id = values[1]
            self.transfer_config.add_recipient(user_id, level)
            count += 1
        
        self._refresh_trees()
        self.log(f"已将所有 {count} 个待转账号移到 {level}级收款账户")
    
    def _move_to_transfer(self, level: int = 1):
        """将选中的账号移到待转账号
        
        Args:
            level: 收款级别（1-3）
        """
        tree = self.recipient_trees[level]
        selection = tree.selection()
        if not selection:
            messagebox.showwarning("提示", "请先选择要移动的账号")
            return
        
        for item in selection:
            values = tree.item(item, "values")
            user_id = values[1]
            self.transfer_config.remove_recipient(user_id, level)
        
        self._refresh_trees()
        self.log(f"已将 {len(selection)} 个账号从 {level}级收款账户移到待转账号")
    
    def _move_all_to_transfer(self, level: int = 1):
        """将指定级别的所有收款账户移到待转账号
        
        Args:
            level: 收款级别（1-3）
        """
        tree = self.recipient_trees[level]
        count = 0
        for item in tree.get_children():
            values = tree.item(item, "values")
            user_id = values[1]
            self.transfer_config.remove_recipient(user_id, level)
            count += 1
        
        self._refresh_trees()
        self.log(f"已将 {level}级的所有 {count} 个收款账户移到待转账号")
    
    def _on_transfer_tree_double_click(self, event):
        """双击待转账号列表，移到收款账户"""
        item = self.transfer_tree.identify_row(event.y)
        if item:
            values = self.transfer_tree.item(item, "values")
            user_id = values[1]
            
            # 确定目标级别
            if self.multi_level_enabled_var.get() and self.recipient_notebook:
                # 多级模式：获取当前选中的标签页级别
                current_tab = self.recipient_notebook.index(self.recipient_notebook.select())
                level = current_tab + 1
            else:
                # 单级模式：固定为1级
                level = 1
            
            self.transfer_config.add_recipient(user_id, level)
            self._refresh_trees()
            self.log(f"已将账号 {values[0]} 移到 {level}级收款账户")
    
    def _on_recipient_tree_double_click(self, event, level: int):
        """双击收款账户列表，移到待转账号
        
        Args:
            event: 事件对象
            level: 收款级别（1-3）
        """
        tree = self.recipient_trees[level]
        item = tree.identify_row(event.y)
        if item:
            values = tree.item(item, "values")
            user_id = values[1]
            self.transfer_config.remove_recipient(user_id, level)
            self._refresh_trees()
            self.log(f"已将账号 {values[0]} 从 {level}级收款账户移到待转账号")
    
    def _on_enabled_changed(self):
        """启用状态改变"""
        self.transfer_config.set_enabled(self.enabled_var.get())
        self.log(f"转账功能已{'启用' if self.enabled_var.get() else '禁用'}")
        
        # 同步更新主界面的滑动开关
        try:
            if self.gui_instance and hasattr(self.gui_instance, 'auto_transfer_switch'):
                self.gui_instance.auto_transfer_switch.set_state(self.enabled_var.get())
        except:
            pass
    
    def _on_recipient_strategy_changed(self):
        """收款人选择策略改变"""
        strategy = self.recipient_strategy_var.get()
        try:
            self.transfer_config.recipient_selection_strategy = strategy
            self.transfer_config.save()
            strategy_name = "轮询（平均分配）" if strategy == "rotation" else "随机选择"
            self.log(f"收款人选择策略已设置为: {strategy_name}")
        except Exception as e:
            messagebox.showerror("错误", f"设置失败: {e}")
            # 恢复到之前的值
            self.recipient_strategy_var.set(self.transfer_config.recipient_selection_strategy)
    
    def _on_multi_level_changed(self):
        """多级转账启用状态改变"""
        self.transfer_config.multi_level_enabled = self.multi_level_enabled_var.get()
        self.transfer_config.save()
        self.log(f"多级转账已{'启用' if self.multi_level_enabled_var.get() else '禁用'}")
        
        # 重新创建收款账户界面
        self._create_recipient_interface()
        # 刷新数据
        self._refresh_trees()
    
    def _on_level_changed(self):
        """转账级数改变"""
        level = self.max_transfer_level_var.get()
        self.transfer_config.max_transfer_level = level
        self.transfer_config.save()
        self.log(f"最大转账级数已设置为: {level} 级")
    
    def _save_threshold(self):
        """保存起步金额"""
        try:
            min_transfer_amount = self.min_transfer_amount_var.get()
            # 保存到配置中
            self.transfer_config.min_transfer_amount = min_transfer_amount
            self.transfer_config.save()
            self.log(f"起步金额已设置为: {min_transfer_amount:.2f} 元")
            messagebox.showinfo("成功", f"起步金额已设置为: {min_transfer_amount:.2f} 元")
        except Exception as e:
            messagebox.showerror("错误", f"保存失败: {e}")
    
    def _save_min_balance(self):
        """保存最小余额"""
        try:
            min_balance = self.min_balance_var.get()
            self.transfer_config.set_min_balance(min_balance)
            self.log(f"保留余额已设置为: {min_balance:.2f} 元")
            messagebox.showinfo("成功", f"保留余额已设置为: {min_balance:.2f} 元")
        except Exception as e:
            messagebox.showerror("错误", f"保存失败: {e}")
    
    def _refresh_accounts(self):
        """刷新账号列表"""
        self._load_accounts()
        self._refresh_trees()
        self.log(f"已刷新账号列表，共 {len(self.accounts)} 个账号")
    
    def _manual_add_recipient(self, level: int = 1):
        """手动添加收款账号
        
        Args:
            level: 收款级别（1-3）
        """
        from tkinter import simpledialog
        
        # 弹出输入对话框
        user_id = simpledialog.askstring(
            f"手动添加 {level}级收款账号",
            f"请输入 {level}级收款账号的用户ID：",
            parent=self.window
        )
        
        if not user_id:
            return
        
        # 去除空格
        user_id = user_id.strip()
        
        if not user_id:
            messagebox.showwarning("提示", "用户ID不能为空")
            return
        
        # 检查是否已经是该级别的收款账号
        if user_id in self.transfer_config.get_recipients(level):
            messagebox.showinfo("提示", f"用户ID {user_id} 已经是 {level}级收款账号")
            return
        
        # 添加到指定级别的收款账号列表
        self.transfer_config.add_recipient(user_id, level)
        
        # 检查是否在账号列表中
        found = False
        for account in self.accounts:
            if account['user_id'] == user_id:
                found = True
                break
        
        # 如果不在账号列表中，添加一个临时账号记录
        if not found:
            self.accounts.append({
                'phone': '手动添加',
                'user_id': user_id,
                'nickname': '手动添加',
                'balance': 0.0
            })
        
        # 刷新显示
        self._refresh_trees()
        self.log(f"已手动添加 {level}级收款账号: {user_id}")
        messagebox.showinfo("成功", f"已添加 {level}级收款账号: {user_id}")
    
    def _on_closing(self):
        """安全关闭窗口"""
        try:
            # 清理树形视图
            # 清理待转账号树
            if hasattr(self, 'transfer_tree') and self.transfer_tree:
                for item in self.transfer_tree.get_children():
                    self.transfer_tree.delete(item)
            
            # 清理收款账号树（可能是单级或多级）
            if hasattr(self, 'recipient_trees') and self.recipient_trees:
                for tree in self.recipient_trees.values():
                    if tree:
                        for item in tree.get_children():
                            tree.delete(item)
            elif hasattr(self, 'recipient_tree') and self.recipient_tree:
                for item in self.recipient_tree.get_children():
                    self.recipient_tree.delete(item)
            
            # 销毁窗口
            self.window.destroy()
        except Exception as e:
            print(f"关闭转账配置窗口时出错: {e}")
            try:
                self.window.destroy()
            except:
                pass





class HistoryResultsWindow:
    """历史结果查看窗口"""
    
    def __init__(self, parent, log_callback, main_gui=None):
        """初始化历史结果窗口
        
        Args:
            parent: 父窗口
            log_callback: 日志回调函数
            main_gui: 主界面GUI实例(可选)
        """
        self.parent = parent
        self.log = log_callback
        self.main_gui = main_gui  # 保存主界面GUI实例
        
        # 创建窗口
        self.window = tk.Toplevel(parent)
        self.window.title("历史结果查看")
        self.window.geometry("1200x700")
        self.window.resizable(True, True)
        
        # 先隐藏窗口，避免白屏
        self.window.withdraw()
        
        # 设置窗口关闭协议
        self.window.protocol("WM_DELETE_WINDOW", self._on_closing)
        
        # 初始化日期过滤器（默认为当天）
        from datetime import datetime
        self.selected_date = datetime.now().strftime('%Y-%m-%d')
        
        # 初始化筛选相关变量
        self.all_tree_items = []  # 存储所有表格项目ID（用于筛选恢复）
        
        # 加载历史结果
        self.results = []
        self._load_results()
        
        # 创建界面
        self._create_widgets()
        
        # 居中显示（不使用grab_set，允许主窗口操作）
        self.window.transient(parent)
        
        # 所有内容准备完成后再显示窗口
        self.window.deiconify()
        self.window.lift()
        self.window.focus_force()
    
    def _load_results(self, date_filter=None):
        """加载历史结果（直接从数据库读取）
        
        统计逻辑：
        1. 每个账号每天只显示一条记录（去重）
        2. 如果当天有多次有效执行，累加数据（签到奖励、余额变化等）
        3. 选择"全部"时，显示每个账号每天的累计数据
        
        Args:
            date_filter: 日期过滤器（格式：YYYY-MM-DD），如果为None则加载所有记录
        """
        try:
            from datetime import datetime
            
            # 如果没有指定日期，使用当前选择的日期
            if date_filter is None:
                date_filter = self.selected_date
            
            self.log(f"开始加载历史记录（日期：{date_filter if date_filter else '全部'}）...")
            
            # 创建数据库实例
            db = LocalDatabase()
            self.log("✓ 成功创建数据库实例")
            
            # 获取所有历史记录
            records = db.get_all_history_records()
            self.log(f"✓ 从数据库查询到 {len(records)} 条记录")
            
            # 按日期过滤
            if date_filter:
                # 只保留指定日期的记录
                records = [r for r in records if r.get('运行日期', '') == date_filter]
                self.log(f"✓ 过滤后剩余 {len(records)} 条记录（日期：{date_filter}）")
            
            # 按"日期-手机号"分组并累加数据
            date_phone_groups = {}  # key: "日期-手机号", value: [记录列表]
            
            for record in records:
                run_date = record.get('运行日期', '')
                phone = record['手机号']
                key = f"{run_date}-{phone}"
                
                if key not in date_phone_groups:
                    date_phone_groups[key] = []
                date_phone_groups[key].append(record)
            
            # 对每组数据进行累加处理
            self.results = []
            for key, group_records in date_phone_groups.items():
                # 按创建时间排序，最新的在前
                group_records.sort(key=lambda x: x.get('创建时间', ''), reverse=True)
                
                # 基础信息使用最新记录
                latest_record = group_records[0]
                
                # 累加数值字段
                total_checkin_reward = 0.0  # 签到奖励累加
                total_transfer_amount = 0.0  # 转账金额累加
                
                # 余额：使用最早记录的"余额前"和最新记录的"余额后"
                earliest_record = group_records[-1]  # 最早的记录
                balance_before = earliest_record.get('余额前(元)', 'N/A')
                balance_after = latest_record.get('余额(元)', 'N/A')
                
                # 累加所有有效记录的数值
                for record in group_records:
                    # 只累加成功的记录
                    if '成功' in record.get('状态', ''):
                        # 签到奖励累加
                        reward = record.get('签到奖励(元)', 0)
                        if reward and reward != '-' and reward != 'N/A':
                            try:
                                total_checkin_reward += float(reward)
                            except:
                                pass
                        
                        # 转账金额累加
                        transfer = record.get('转账金额(元)', 0)
                        if transfer and transfer != '-' and transfer != 'N/A':
                            try:
                                total_transfer_amount += float(transfer)
                            except:
                                pass
                
                # 格式化转账信息（使用累加后的金额）
                transfer_recipient = latest_record.get('转账收款人', '')
                
                # 构建结果字典
                result_dict = {
                    '手机号': latest_record['手机号'],
                    '昵称': latest_record['昵称'],
                    '用户ID': latest_record['用户ID'],
                    '余额前(元)': balance_before,  # 使用最早记录的余额前
                    '积分': latest_record['积分'],
                    '抵扣券(张)': latest_record['抵扣券(张)'],
                    '优惠券(张)': latest_record['优惠券(张)'],
                    '签到奖励(元)': total_checkin_reward,  # 累加后的签到奖励
                    '签到总次数': latest_record['签到总次数'],
                    '余额(元)': balance_after,  # 使用最新记录的余额后
                    '转账金额(元)': total_transfer_amount,  # 累加后的转账金额
                    '转账收款人': transfer_recipient,
                    '耗时(秒)': latest_record['耗时(秒)'],
                    '状态': latest_record['状态'],
                    '登录方式': latest_record['登录方式'],
                    '管理员': latest_record.get('管理员', '-'),
                    '时间戳': latest_record.get('创建时间', ''),
                    '执行次数': len(group_records)  # 记录当天执行次数
                }
                self.results.append(result_dict)
            
            self.log(f"✓ 成功加载 {len(self.results)} 条历史结果（日期：{date_filter if date_filter else '全部'}）")
            if date_filter:
                # 统计当天多次执行的账号
                multi_exec_count = sum(1 for r in self.results if r.get('执行次数', 1) > 1)
                if multi_exec_count > 0:
                    self.log(f"  其中 {multi_exec_count} 个账号当天有多次执行（已累加数据）")
            
        except Exception as e:
            self.log(f"❌ 加载历史结果失败: {e}")
            import traceback
            error_details = traceback.format_exc()
            self.log(f"错误详情:\n{error_details}")
            # 确保 results 是空列表，避免后续代码出错
            self.results = []
    
    def _create_widgets(self):
        """创建界面组件"""
        main_frame = ttk.Frame(self.window, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # === 日期选择区域 ===
        date_frame = ttk.LabelFrame(main_frame, text="日期筛选", padding="10")
        date_frame.pack(fill=tk.X, pady=(0, 10))
        
        # 日期选择控件
        date_control_frame = ttk.Frame(date_frame)
        date_control_frame.pack(fill=tk.X)
        
        ttk.Label(date_control_frame, text="选择日期：").pack(side=tk.LEFT, padx=(0, 5))
        
        # 日期输入框
        self.date_entry = ttk.Entry(date_control_frame, width=12)
        self.date_entry.insert(0, self.selected_date)
        self.date_entry.pack(side=tk.LEFT, padx=(0, 5))
        
        # 绑定回车键自动刷新
        self.date_entry.bind('<Return>', lambda e: self._refresh_by_date())
        
        # 绑定鼠标滚轮事件（滚动日期）
        self.date_entry.bind('<MouseWheel>', self._on_date_scroll)
        
        # 快捷按钮
        ttk.Button(date_control_frame, text="◀ 前一天", command=self._select_previous_day).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(date_control_frame, text="今天", command=self._select_today).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(date_control_frame, text="后一天 ▶", command=self._select_next_day).pack(side=tk.LEFT, padx=(0, 5))
        
        # 时间范围筛选按钮
        ttk.Button(date_control_frame, text="最近一周", command=self._select_last_week).pack(side=tk.LEFT, padx=(10, 5))
        ttk.Button(date_control_frame, text="最近半月", command=self._select_last_half_month).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(date_control_frame, text="最近一月", command=self._select_last_month).pack(side=tk.LEFT, padx=(0, 5))
        
        ttk.Button(date_control_frame, text="全部", command=self._select_all).pack(side=tk.LEFT, padx=(10, 5))
        
        # 提示信息
        ttk.Label(date_control_frame, text="（提示：在日期框上滚动鼠标滚轮可切换日期）", 
                 foreground="gray", font=("TkDefaultFont", 8)).pack(side=tk.LEFT, padx=(10, 0))
        
        # === 顶部信息区域 ===
        info_frame = ttk.LabelFrame(main_frame, text="统计信息", padding="10")
        info_frame.pack(fill=tk.X, pady=(0, 10))
        
        # 统计信息
        stats_text = f"共 {len(self.results)} 条记录"
        if self.results:
            success_count = sum(1 for r in self.results if r.get('状态') == '成功')
            failed_count = len(self.results) - success_count
            stats_text += f"  |  成功: {success_count}  |  失败: {failed_count}"
            
            # 计算总余额变化
            total_balance_change = 0.0
            for r in self.results:
                balance_before = r.get('余额前(元)', 'N/A')
                balance_after = r.get('余额(元)', 'N/A')
                if balance_before != 'N/A' and balance_after != 'N/A':
                    try:
                        total_balance_change += float(balance_after) - float(balance_before)
                    except:
                        pass
            
            stats_text += f"  |  总余额变化: {total_balance_change:.2f} 元"
        
        # 保存统计标签以便更新
        self.stats_label = ttk.Label(info_frame, text=stats_text, font=("TkDefaultFont", 10))
        self.stats_label.pack(anchor=tk.W)
        
        # === 结果表格区域 ===
        table_frame = ttk.LabelFrame(main_frame, text="历史结果", padding="10")
        table_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # 创建Treeview
        columns = (
            "nickname", "user_id", "phone", "balance_before", "points", "vouchers",
            "checkin_reward", "checkin_total_times",
            "balance_after", "transfer_info", "status", "timestamp", "owner"
        )
        
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=20)
        
        # 定义列标题和宽度
        column_config = {
            "nickname": ("昵称", 80),
            "user_id": ("ID", 80),
            "phone": ("手机号", 100),
            "balance_before": ("余额前", 70),
            "points": ("积分", 60),
            "vouchers": ("抵扣券", 60),
            "checkin_reward": ("签到奖励", 80),
            "checkin_total_times": ("签到次数", 70),
            "balance_after": ("余额", 70),
            "transfer_info": ("转账", 120),
            "status": ("状态", 60),
            "timestamp": ("时间", 140),
            "owner": ("管理员", 80)
        }
        
        for col, (heading, width) in column_config.items():
            self.tree.heading(col, text=heading)
            self.tree.column(col, width=width, anchor=tk.CENTER)
        
        # 添加滚动条
        scrollbar_y = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.tree.yview)
        scrollbar_x = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        
        # 布局
        self.tree.grid(row=0, column=0, sticky="nsew")
        scrollbar_y.grid(row=0, column=1, sticky="ns")
        scrollbar_x.grid(row=1, column=0, sticky="ew")
        
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)
        
        # 配置标签颜色（整行文字颜色）
        self.tree.tag_configure("success", foreground="green")
        self.tree.tag_configure("failed", foreground="red")
        self.tree.tag_configure("transfer_success", foreground="blue")  # 转账成功：蓝色
        
        # 绑定右键菜单
        self.tree.bind("<Button-3>", self._show_main_context_menu)
        
        # 填充数据
        self._refresh_tree()
        
        # === 底部按钮区域 ===
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X)
        
        ttk.Button(button_frame, text="刷新", command=self._refresh_data).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(button_frame, text="🔍 定位失败账户", command=self._locate_failed_account).pack(side=tk.LEFT, padx=(0, 5))
        
        # 快速筛选按钮
        ttk.Button(button_frame, text="🔍 执行失败", command=self._filter_failed).pack(side=tk.LEFT, padx=(10, 5))
        ttk.Button(button_frame, text="💰 有余额", command=self._filter_has_balance).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(button_frame, text="📭 无余额", command=self._filter_no_balance).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(button_frame, text="🔄 显示全部", command=self._show_all).pack(side=tk.LEFT, padx=(0, 5))
        
        # 搜索框
        ttk.Label(button_frame, text="搜索:", width=6).pack(side=tk.LEFT, padx=(10, 5))
        self.history_search_var = tk.StringVar()
        self.history_search_entry = ttk.Entry(button_frame, textvariable=self.history_search_var, width=15)
        self.history_search_entry.pack(side=tk.LEFT, padx=(0, 5))
        self.history_search_entry.bind('<Return>', lambda e: self._search_history_table())
        ttk.Button(button_frame, text="🔍 搜索", command=self._search_history_table, width=8).pack(side=tk.LEFT, padx=(0, 5))
        
        ttk.Button(button_frame, text="导出Excel", command=self._export_excel).pack(side=tk.LEFT, padx=(10, 5))
        ttk.Button(button_frame, text="关闭", command=self.window.destroy).pack(side=tk.RIGHT)
    
    def _refresh_tree(self):
        """刷新树形列表"""
        # 清空现有项
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # 添加结果
        for result in self.results:
            # 格式化转账信息
            transfer_amount = result.get('转账金额(元)', 0.0)
            transfer_recipient = result.get('转账收款人', '')
            
            if transfer_amount and transfer_amount > 0:
                transfer_info = f"{transfer_amount:.2f}→{transfer_recipient}"
            else:
                transfer_info = ""
            
            # 从数据库字段读取管理员（而不是实时查询user_manager）
            owner_name = result.get('管理员', '-')
            if not owner_name or owner_name == '-':
                owner_name = "-"
            
            # 处理None值，确保显示正确
            def format_value(value, default='N/A'):
                """格式化显示值，处理None和数值"""
                if value is None:
                    return default
                if isinstance(value, (int, float)):
                    # 数值类型：如果是整数显示为整数，否则保留2位小数
                    if isinstance(value, float) and value == int(value):
                        return str(int(value))
                    elif isinstance(value, float):
                        return f"{value:.2f}"
                    else:
                        return str(value)
                return str(value) if value else default
            
            values = (
                format_value(result.get('昵称'), 'N/A'),
                format_value(result.get('用户ID'), 'N/A'),
                format_value(result.get('手机号'), 'N/A'),
                format_value(result.get('余额前(元)'), '0.0'),
                format_value(result.get('积分'), '0'),
                format_value(result.get('抵扣券(张)'), '0'),
                format_value(result.get('签到奖励(元)'), '0.0'),
                format_value(result.get('签到总次数'), '0'),
                format_value(result.get('余额(元)'), '0.0'),
                transfer_info,
                format_value(result.get('状态'), 'N/A'),
                format_value(result.get('时间戳'), 'N/A'),
                owner_name
            )
            
            # 根据状态和转账情况选择标签颜色
            status = result.get('状态', 'N/A')
            transfer_amount = result.get('转账金额(元)', 0.0)
            
            # 优先级：转账成功 > 失败 > 成功
            if '成功' in status and transfer_amount and transfer_amount > 0:
                # 转账成功：蓝色
                tag = "transfer_success"
            elif '失败' in status:
                # 失败：红色
                tag = "failed"
            elif '成功' in status:
                # 成功（无转账）：绿色
                tag = "success"
            else:
                tag = ""
            
            self.tree.insert("", tk.END, values=values, tags=(tag,))
        
        # 保存所有项目ID（用于筛选）
        self.all_tree_items = list(self.tree.get_children())
    
    def _on_date_scroll(self, event):
        """鼠标滚轮滚动日期
        
        Args:
            event: 鼠标滚轮事件
        """
        # 如果当前是"全部"模式，不响应滚轮
        date_str = self.date_entry.get().strip()
        if date_str == "全部":
            return
        
        try:
            from datetime import datetime, timedelta
            
            # 获取当前日期
            current_date = datetime.strptime(date_str, '%Y-%m-%d')
            
            # 根据滚轮方向调整日期
            # event.delta > 0 表示向上滚动（前一天）
            # event.delta < 0 表示向下滚动（后一天）
            if event.delta > 0:
                # 向上滚动 - 前一天
                new_date = current_date - timedelta(days=1)
            else:
                # 向下滚动 - 后一天
                new_date = current_date + timedelta(days=1)
            
            # 更新日期并刷新
            self.selected_date = new_date.strftime('%Y-%m-%d')
            self.date_entry.delete(0, tk.END)
            self.date_entry.insert(0, self.selected_date)
            self._refresh_by_date()
            
        except ValueError:
            # 日期格式错误，忽略
            pass
    
    def _select_previous_day(self):
        """选择前一天"""
        from datetime import datetime, timedelta
        
        # 如果当前是"全部"，则从今天开始
        date_str = self.date_entry.get().strip()
        if date_str == "全部":
            current_date = datetime.now()
        else:
            try:
                current_date = datetime.strptime(date_str, '%Y-%m-%d')
            except ValueError:
                current_date = datetime.now()
        
        previous_day = current_date - timedelta(days=1)
        self.selected_date = previous_day.strftime('%Y-%m-%d')
        self.date_entry.delete(0, tk.END)
        self.date_entry.insert(0, self.selected_date)
        self._refresh_by_date()
    
    def _select_next_day(self):
        """选择后一天"""
        from datetime import datetime, timedelta
        
        # 如果当前是"全部"，则从今天开始
        date_str = self.date_entry.get().strip()
        if date_str == "全部":
            current_date = datetime.now()
        else:
            try:
                current_date = datetime.strptime(date_str, '%Y-%m-%d')
            except ValueError:
                current_date = datetime.now()
        
        next_day = current_date + timedelta(days=1)
        self.selected_date = next_day.strftime('%Y-%m-%d')
        self.date_entry.delete(0, tk.END)
        self.date_entry.insert(0, self.selected_date)
        self._refresh_by_date()
    
    def _select_today(self):
        """选择今天"""
        from datetime import datetime
        self.selected_date = datetime.now().strftime('%Y-%m-%d')
        self.date_entry.delete(0, tk.END)
        self.date_entry.insert(0, self.selected_date)
        self._refresh_by_date()
    
    def _select_yesterday(self):
        """选择昨天"""
        from datetime import datetime, timedelta
        yesterday = datetime.now() - timedelta(days=1)
        self.selected_date = yesterday.strftime('%Y-%m-%d')
        self.date_entry.delete(0, tk.END)
        self.date_entry.insert(0, self.selected_date)
        self._refresh_by_date()
    
    def _select_all(self):
        """选择全部（不过滤日期）"""
        self.selected_date = None
        self.date_entry.delete(0, tk.END)
        self.date_entry.insert(0, "全部")
        self._refresh_by_date()
    
    def _select_last_week(self):
        """选择最近一周（7天）"""
        from datetime import datetime, timedelta
        end_date = datetime.now()
        start_date = end_date - timedelta(days=6)  # 包括今天，共7天
        
        date_range = f"{start_date.strftime('%Y-%m-%d')} ~ {end_date.strftime('%Y-%m-%d')}"
        self.selected_date = date_range
        self.date_entry.delete(0, tk.END)
        self.date_entry.insert(0, date_range)
        self._refresh_by_date_range(start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d'))
    
    def _select_last_half_month(self):
        """选择最近半月（15天）"""
        from datetime import datetime, timedelta
        end_date = datetime.now()
        start_date = end_date - timedelta(days=14)  # 包括今天，共15天
        
        date_range = f"{start_date.strftime('%Y-%m-%d')} ~ {end_date.strftime('%Y-%m-%d')}"
        self.selected_date = date_range
        self.date_entry.delete(0, tk.END)
        self.date_entry.insert(0, date_range)
        self._refresh_by_date_range(start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d'))
    
    def _select_last_month(self):
        """选择最近一月（30天）"""
        from datetime import datetime, timedelta
        end_date = datetime.now()
        start_date = end_date - timedelta(days=29)  # 包括今天，共30天
        
        date_range = f"{start_date.strftime('%Y-%m-%d')} ~ {end_date.strftime('%Y-%m-%d')}"
        self.selected_date = date_range
        self.date_entry.delete(0, tk.END)
        self.date_entry.insert(0, date_range)
        self._refresh_by_date_range(start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d'))
    
    def _refresh_by_date_range(self, start_date, end_date):
        """根据日期范围刷新数据
        
        Args:
            start_date: 开始日期（格式：YYYY-MM-DD）
            end_date: 结束日期（格式：YYYY-MM-DD）
        """
        try:
            from datetime import datetime
            
            self.log(f"开始加载历史记录（日期范围：{start_date} ~ {end_date}）...")
            
            # 创建数据库实例
            db = LocalDatabase()
            
            # 获取所有历史记录
            records = db.get_all_history_records()
            self.log(f"✓ 从数据库查询到 {len(records)} 条记录")
            
            # 按日期范围过滤
            filtered_records = []
            for r in records:
                run_date = r.get('运行日期', '')
                if run_date and start_date <= run_date <= end_date:
                    filtered_records.append(r)
            
            self.log(f"✓ 过滤后剩余 {len(filtered_records)} 条记录（日期范围：{start_date} ~ {end_date}）")
            
            # 按"日期-手机号"分组并累加数据
            date_phone_groups = {}
            
            for record in filtered_records:
                run_date = record.get('运行日期', '')
                phone = record['手机号']
                key = f"{run_date}-{phone}"
                
                if key not in date_phone_groups:
                    date_phone_groups[key] = []
                date_phone_groups[key].append(record)
            
            # 对每组数据进行累加处理
            self.results = []
            for key, group_records in date_phone_groups.items():
                # 按创建时间排序，最新的在前
                group_records.sort(key=lambda x: x.get('创建时间', ''), reverse=True)
                
                # 基础信息使用最新记录
                latest_record = group_records[0]
                
                # 累加数值字段
                total_checkin_reward = 0.0
                total_transfer_amount = 0.0
                
                # 余额：使用最早记录的"余额前"和最新记录的"余额后"
                earliest_record = group_records[-1]
                balance_before = earliest_record.get('余额前(元)', 'N/A')
                balance_after = latest_record.get('余额(元)', 'N/A')
                
                # 累加所有有效记录的数值
                for record in group_records:
                    if '成功' in record.get('状态', ''):
                        reward = record.get('签到奖励(元)', 0)
                        if reward and reward != '-' and reward != 'N/A':
                            try:
                                total_checkin_reward += float(reward)
                            except:
                                pass
                        
                        transfer = record.get('转账金额(元)', 0)
                        if transfer and transfer != '-' and transfer != 'N/A':
                            try:
                                total_transfer_amount += float(transfer)
                            except:
                                pass
                
                # 格式化转账信息
                transfer_recipient = latest_record.get('转账收款人', '')
                
                # 构建结果字典
                result_dict = {
                    '手机号': latest_record['手机号'],
                    '昵称': latest_record['昵称'],
                    '用户ID': latest_record['用户ID'],
                    '余额前(元)': balance_before,
                    '积分': latest_record['积分'],
                    '抵扣券(张)': latest_record['抵扣券(张)'],
                    '优惠券(张)': latest_record['优惠券(张)'],
                    '签到奖励(元)': total_checkin_reward,
                    '签到总次数': latest_record['签到总次数'],
                    '余额(元)': balance_after,
                    '转账金额(元)': total_transfer_amount,
                    '转账收款人': transfer_recipient,
                    '耗时(秒)': latest_record['耗时(秒)'],
                    '状态': latest_record['状态'],
                    '登录方式': latest_record['登录方式'],
                    '管理员': latest_record.get('管理员', '-'),
                    '时间戳': latest_record.get('创建时间', ''),
                    '执行次数': len(group_records)
                }
                self.results.append(result_dict)
            
            self.log(f"✓ 成功加载 {len(self.results)} 条历史结果（日期范围：{start_date} ~ {end_date}）")
            
            # 刷新界面
            self._refresh_tree()
            self._update_stats()
            
        except Exception as e:
            self.log(f"❌ 加载历史结果失败: {e}")
            import traceback
            error_details = traceback.format_exc()
            self.log(f"错误详情:\n{error_details}")
            self.results = []
    
    def _refresh_by_date(self):
        """根据选择的日期刷新数据（自动刷新，不需要点击按钮）"""
        # 获取输入框中的日期
        date_str = self.date_entry.get().strip()
        
        # 如果是"全部"，则不过滤日期
        if date_str == "全部":
            self.selected_date = None
        else:
            # 验证日期格式
            try:
                from datetime import datetime
                datetime.strptime(date_str, '%Y-%m-%d')
                self.selected_date = date_str
            except ValueError:
                self.log(f"❌ 日期格式错误，请使用 YYYY-MM-DD 格式")
                return
        
        # 刷新数据
        self._load_results(self.selected_date)
        self._refresh_tree()
        self._update_stats()
        
        if self.selected_date:
            self.log(f"✓ 已加载 {self.selected_date} 的历史记录")
        else:
            self.log(f"✓ 已加载全部历史记录")
    
    def _update_stats(self):
        """更新统计信息"""
        stats_text = f"共 {len(self.results)} 条记录"
        if self.results:
            success_count = sum(1 for r in self.results if r.get('状态') == '成功')
            failed_count = len(self.results) - success_count
            stats_text += f"  |  成功: {success_count}  |  失败: {failed_count}"
            
            # 计算总余额变化
            total_balance_change = 0.0
            for r in self.results:
                balance_before = r.get('余额前(元)', 'N/A')
                balance_after = r.get('余额(元)', 'N/A')
                if balance_before != 'N/A' and balance_after != 'N/A':
                    try:
                        total_balance_change += float(balance_after) - float(balance_before)
                    except:
                        pass
            
            stats_text += f"  |  总余额变化: {total_balance_change:.2f} 元"
        
        self.stats_label.config(text=stats_text)
    
    def _refresh_data(self):
        """刷新数据"""
        self._load_results(self.selected_date)
        self._refresh_tree()
        self._update_stats()
        self.log("历史结果已刷新")
    
    def _show_main_context_menu(self, event):
        """显示主界面账号列表的右键菜单"""
        # 选中右键点击的项
        item = self.tree.identify_row(event.y)
        if not item:
            return
        
        self.tree.selection_set(item)
        
        # 获取该项的数据
        values = self.tree.item(item, 'values')
        if not values or len(values) < 3:
            return
        
        nickname = values[0]  # 昵称
        user_id = values[1]  # 用户ID
        phone = values[2]  # 手机号
        
        # 创建右键菜单
        context_menu = tk.Menu(self.tree, tearoff=0)
        context_menu.add_command(label=f"📋 复制手机号: {phone}", command=lambda: self._copy_to_clipboard(phone))
        if user_id and user_id != 'N/A' and user_id != '-':
            context_menu.add_command(label=f"📋 复制用户ID: {user_id}", command=lambda: self._copy_to_clipboard(user_id))
        if nickname and nickname != 'N/A' and nickname != '-':
            context_menu.add_command(label=f"📋 复制昵称: {nickname}", command=lambda: self._copy_to_clipboard(nickname))
        
        # 显示菜单
        context_menu.post(event.x_root, event.y_root)
    
    def _copy_to_clipboard(self, text):
        """复制文本到剪贴板"""
        self.window.clipboard_clear()
        self.window.clipboard_append(text)
        self.log(f"✓ 已复制到剪贴板: {text}")
    
    def _export_excel(self):
        """导出Excel - 支持按时间范围导出，每天记录清晰区分"""
        try:
            from tkinter import filedialog, messagebox
            from datetime import datetime, timedelta
            
            # 创建导出选项对话框
            export_dialog = tk.Toplevel(self.window)
            export_dialog.title("导出Excel - 选择时间范围")
            export_dialog.geometry("400x250")
            export_dialog.transient(self.window)
            export_dialog.grab_set()
            
            # 居中显示
            export_dialog.update_idletasks()
            x = (export_dialog.winfo_screenwidth() // 2) - (400 // 2)
            y = (export_dialog.winfo_screenheight() // 2) - (250 // 2)
            export_dialog.geometry(f'400x250+{x}+{y}')
            
            main_frame = ttk.Frame(export_dialog, padding="20")
            main_frame.pack(fill=tk.BOTH, expand=True)
            
            # 标题
            ttk.Label(main_frame, text="选择导出时间范围", font=("TkDefaultFont", 12, "bold")).pack(pady=(0, 15))
            
            # 时间范围选项
            range_var = tk.StringVar(value="today")
            
            ttk.Radiobutton(main_frame, text="今天", variable=range_var, value="today").pack(anchor=tk.W, pady=5)
            ttk.Radiobutton(main_frame, text="最近7天", variable=range_var, value="week").pack(anchor=tk.W, pady=5)
            ttk.Radiobutton(main_frame, text="最近半月（15天）", variable=range_var, value="half_month").pack(anchor=tk.W, pady=5)
            ttk.Radiobutton(main_frame, text="最近一月（30天）", variable=range_var, value="month").pack(anchor=tk.W, pady=5)
            ttk.Radiobutton(main_frame, text="全部记录", variable=range_var, value="all").pack(anchor=tk.W, pady=5)
            
            # 按钮
            button_frame = ttk.Frame(main_frame)
            button_frame.pack(pady=(15, 0))
            
            result = {'confirmed': False, 'range': None}
            
            def on_confirm():
                result['confirmed'] = True
                result['range'] = range_var.get()
                export_dialog.destroy()
            
            def on_cancel():
                export_dialog.destroy()
            
            ttk.Button(button_frame, text="确定", command=on_confirm, width=10).pack(side=tk.LEFT, padx=5)
            ttk.Button(button_frame, text="取消", command=on_cancel, width=10).pack(side=tk.LEFT, padx=5)
            
            # 等待对话框关闭
            self.window.wait_window(export_dialog)
            
            if not result['confirmed']:
                return
            
            # 计算日期范围
            now = datetime.now()
            today = now.strftime('%Y-%m-%d')
            
            range_type = result['range']
            if range_type == "today":
                start_date = today
                end_date = today
                range_name = "今天"
            elif range_type == "week":
                start_date = (now - timedelta(days=6)).strftime('%Y-%m-%d')
                end_date = today
                range_name = "最近7天"
            elif range_type == "half_month":
                start_date = (now - timedelta(days=14)).strftime('%Y-%m-%d')
                end_date = today
                range_name = "最近半月"
            elif range_type == "month":
                start_date = (now - timedelta(days=29)).strftime('%Y-%m-%d')
                end_date = today
                range_name = "最近一月"
            else:  # all
                start_date = None
                end_date = None
                range_name = "全部"
            
            # 选择保存路径
            default_name = f"历史结果_{range_name}_{now.year}年{now.month}月{now.day}日.xlsx"
            filepath = filedialog.asksaveasfilename(
                title="导出Excel",
                defaultextension=".xlsx",
                filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")],
                initialfile=default_name
            )
            
            if not filepath:
                return
            
            # 从数据库加载指定范围的数据
            self.log(f"正在导出{range_name}的数据...")
            db = LocalDatabase()
            
            if start_date and end_date:
                all_records = db.get_history_records(start_date=start_date, end_date=end_date, limit=100000)
            else:
                all_records = db.get_history_records(limit=100000)
            
            if not all_records:
                messagebox.showinfo("提示", "没有数据可导出")
                return
            
            # 按日期+手机号分组并累加数据
            date_phone_groups = {}
            for record in all_records:
                run_date = record.get('run_date', '')
                phone = record.get('phone', '')
                key = f"{run_date}-{phone}"
                
                if key not in date_phone_groups:
                    date_phone_groups[key] = []
                date_phone_groups[key].append(record)
            
            # 处理每组数据
            processed_records = []
            for key, group_records in date_phone_groups.items():
                group_records.sort(key=lambda x: x.get('created_at', ''), reverse=True)
                
                latest_record = group_records[0]
                earliest_record = group_records[-1]
                
                # 累加签到奖励
                total_checkin_reward = 0.0
                for r in group_records:
                    if '成功' in r.get('status', ''):
                        reward = r.get('checkin_reward', 0)
                        if reward:
                            try:
                                total_checkin_reward += float(reward)
                            except:
                                pass
                
                processed_record = {
                    '日期': latest_record.get('run_date', ''),
                    '手机号': latest_record.get('phone', ''),
                    '昵称': latest_record.get('nickname', ''),
                    '用户ID': latest_record.get('user_id', ''),
                    '余额前(元)': earliest_record.get('balance_before', 0),
                    '积分': latest_record.get('points', 0),
                    '抵扣券(张)': latest_record.get('vouchers', 0),
                    '优惠券(张)': latest_record.get('coupons', 0),
                    '签到奖励(元)': total_checkin_reward,
                    '签到总次数': latest_record.get('checkin_total_times', 0),
                    '余额(元)': latest_record.get('balance_after', 0),
                    '转账金额(元)': latest_record.get('transfer_amount', 0),
                    '转账收款人': latest_record.get('transfer_recipient', ''),
                    '耗时(秒)': latest_record.get('duration', 0),
                    '状态': latest_record.get('status', ''),
                    '登录方式': latest_record.get('login_method', ''),
                    '管理员': latest_record.get('owner', '-'),
                    '执行次数': len(group_records)
                }
                processed_records.append(processed_record)
            
            # 按日期排序
            processed_records.sort(key=lambda x: x['日期'], reverse=True)
            
            # 导出到Excel（按日期分组）
            self._export_to_excel_with_date_groups(filepath, processed_records, range_name)
            
            messagebox.showinfo("成功", f"已成功导出 {len(processed_records)} 条记录到:\n{filepath}")
            self.log(f"✓ 成功导出 {len(processed_records)} 条记录")
            
        except Exception as e:
            messagebox.showerror("错误", f"导出失败: {e}")
            self.log(f"❌ 导出失败: {e}")
            import traceback
            traceback.print_exc()
    
    def _export_to_excel_with_date_groups(self, filepath: str, records: list, range_name: str):
        """导出数据到Excel文件（按日期分组显示）"""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from collections import defaultdict
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"历史结果_{range_name}"[:31]  # Excel工作表名称限制31字符
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        date_header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        
        current_row = 1
        
        # 按日期分组
        records_by_date = defaultdict(list)
        for record in records:
            date = record['日期']
            records_by_date[date].append(record)
        
        sorted_dates = sorted(records_by_date.keys(), reverse=True)
        
        headers = ['手机号', '昵称', '用户ID', '余额前(元)', '积分', '抵扣券(张)', '优惠券(张)', 
                  '签到奖励(元)', '签到总次数', '余额(元)', '余额变化(元)', '转账金额(元)', '转账收款人', 
                  '耗时(秒)', '状态', '登录方式', '管理员', '执行次数']
        
        # 遍历每个日期
        for date in sorted_dates:
            date_records = records_by_date[date]
            
            # 日期标题行
            date_cell = ws.cell(row=current_row, column=1, value=f"📅 {date} ({len(date_records)}个账号)")
            date_cell.font = Font(bold=True, size=12, color="FFFFFF")
            date_cell.fill = date_header_fill
            date_cell.alignment = Alignment(horizontal='left', vertical='center')
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
            current_row += 1
            
            # 表头行
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=header)
                cell.font = Font(bold=True, size=10, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            current_row += 1
            
            # 数据行
            date_total_balance_change = 0.0
            date_total_checkin_reward = 0.0
            
            for record in date_records:
                balance_before = record.get('余额前(元)', 0)
                balance_after = record.get('余额(元)', 0)
                try:
                    balance_change = float(balance_after) - float(balance_before)
                    date_total_balance_change += balance_change
                except:
                    balance_change = 0
                
                try:
                    date_total_checkin_reward += float(record.get('签到奖励(元)', 0))
                except:
                    pass
                
                row_data = [
                    record.get('手机号', ''),
                    record.get('昵称', ''),
                    record.get('用户ID', ''),
                    record.get('余额前(元)', 0),
                    record.get('积分', 0),
                    record.get('抵扣券(张)', 0),
                    record.get('优惠券(张)', 0),
                    record.get('签到奖励(元)', 0),
                    record.get('签到总次数', 0),
                    record.get('余额(元)', 0),
                    balance_change,
                    record.get('转账金额(元)', 0),
                    record.get('转账收款人', ''),
                    record.get('耗时(秒)', 0),
                    record.get('状态', ''),
                    record.get('登录方式', ''),
                    record.get('管理员', '-'),
                    record.get('执行次数', 1)
                ]
                
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                    
                    if col_idx in [4, 5, 6, 7, 8, 9, 10, 11, 12, 14]:
                        if col_idx in [4, 7, 9, 10, 11, 12]:
                            cell.number_format = '0.00'
                        else:
                            cell.number_format = '0'
                    elif col_idx in [1, 3]:
                        cell.number_format = '@'
                    
                    if col_idx == 15:
                        if value == '成功':
                            cell.font = Font(color="008000", bold=True)
                        elif value == '失败':
                            cell.font = Font(color="FF0000", bold=True)
                    
                    if col_idx == 11:
                        if isinstance(value, (int, float)) and value > 0:
                            cell.font = Font(color="008000")
                        elif isinstance(value, (int, float)) and value < 0:
                            cell.font = Font(color="FF0000")
                
                current_row += 1
            
            # 当天小计行
            subtotal_cell = ws.cell(row=current_row, column=1, value=f"📊 {date} 小计")
            subtotal_cell.font = Font(bold=True, color="0000FF")
            subtotal_cell.alignment = Alignment(horizontal='left', vertical='center')
            
            ws.cell(row=current_row, column=8, value=date_total_checkin_reward).number_format = '0.00'
            ws.cell(row=current_row, column=8).font = Font(bold=True, color="008000")
            
            ws.cell(row=current_row, column=11, value=date_total_balance_change).number_format = '0.00'
            ws.cell(row=current_row, column=11).font = Font(bold=True, color="008000")
            
            current_row += 2
        
        # 总计统计
        current_row += 1
        summary_title = ws.cell(row=current_row, column=1, value="=== 总计统计 ===")
        summary_title.font = Font(bold=True, size=12, color="FF0000")
        current_row += 1
        
        total_count = len(records)
        success_count = sum(1 for r in records if r.get('状态') == '成功')
        total_checkin_reward = sum(float(r.get('签到奖励(元)', 0)) for r in records)
        total_balance_change = sum(
            float(r.get('余额(元)', 0)) - float(r.get('余额前(元)', 0))
            for r in records
            if r.get('余额(元)') and r.get('余额前(元)')
        )
        
        summary_data = [
            ("总记录数", total_count),
            ("成功数", success_count),
            ("失败数", total_count - success_count),
            ("成功率", f"{success_count/total_count*100:.1f}%" if total_count > 0 else "0%"),
            ("总签到奖励(元)", f"{total_checkin_reward:.2f}"),
            ("总余额变化(元)", f"{total_balance_change:.2f}")
        ]
        
        for label, value in summary_data:
            ws.cell(row=current_row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=current_row, column=2, value=value)
            current_row += 1
        
        # 自动调整列宽
        for col_idx in range(1, len(headers) + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            
            for cell in ws[column_letter]:
                if cell.value:
                    cell_value = str(cell.value)
                    length = sum(2 if '\u4e00' <= char <= '\u9fff' else 1 for char in cell_value)
                    max_length = max(max_length, length)
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filepath)
    
    def _locate_failed_account(self):
        """定位到第一个失败的账户"""
        # 获取所有表格项
        all_items = self.tree.get_children()
        
        if not all_items:
            self.log("⚠️ 表格中没有数据")
            messagebox.showinfo("提示", "表格中没有数据")
            return
        
        # 查找第一个失败的账户
        failed_item = None
        for item in all_items:
            values = self.tree.item(item, 'values')
            if values and len(values) > 10:  # 确保有足够的列
                status = values[10]  # 状态列是第11列（索引10）
                if '失败' in str(status):
                    failed_item = item
                    break
        
        if failed_item:
            # 清除之前的选择
            self.tree.selection_remove(*self.tree.selection())
            
            # 选中失败的账户
            self.tree.selection_set(failed_item)
            
            # 滚动到该项，使其可见
            self.tree.see(failed_item)
            
            # 获取账户信息
            values = self.tree.item(failed_item, 'values')
            nickname = values[0] if len(values) > 0 else 'N/A'
            phone = values[2] if len(values) > 2 else 'N/A'
            
            self.log(f"✓ 已定位到失败账户: {nickname} ({phone})")
        else:
            self.log("✓ 没有找到失败的账户")
            messagebox.showinfo("提示", "没有找到失败的账户")
    
    def _filter_failed(self):
        """筛选执行失败的账户（只显示失败的账户）"""
        # 保存所有项目ID（如果还没保存）
        if not self.all_tree_items:
            self.all_tree_items = list(self.tree.get_children())
        
        if not self.all_tree_items:
            self.log("⚠️ 表格中没有数据")
            messagebox.showinfo("提示", "表格中没有数据")
            return
        
        # 先detach所有项目
        for item in self.all_tree_items:
            self.tree.detach(item)
        
        # 只reattach失败的账户
        failed_count = 0
        for item in self.all_tree_items:
            values = self.tree.item(item, 'values')
            if values and len(values) > 10:  # 确保有足够的列
                status = values[10]  # 状态列是第11列（索引10）
                if '失败' in str(status):
                    self.tree.reattach(item, '', 'end')
                    failed_count += 1
        
        if failed_count > 0:
            self.log(f"✓ 已筛选出 {failed_count} 个执行失败的账户")
        else:
            self.log("✓ 没有找到失败的账户")
            messagebox.showinfo("提示", "没有找到失败的账户")
    
    def _filter_has_balance(self):
        """筛选有余额的账户（余额不为0）"""
        # 保存所有项目ID（如果还没保存）
        if not self.all_tree_items:
            self.all_tree_items = list(self.tree.get_children())
        
        if not self.all_tree_items:
            self.log("⚠️ 表格中没有数据")
            messagebox.showinfo("提示", "表格中没有数据")
            return
        
        # 先detach所有项目
        for item in self.all_tree_items:
            self.tree.detach(item)
        
        # 只reattach有余额的账户（余额 > 0）
        has_balance_count = 0
        for item in self.all_tree_items:
            values = self.tree.item(item, 'values')
            if values and len(values) > 8:  # 确保有足够的列
                balance_after = values[8]  # 余额列是第9列（索引8）
                try:
                    balance = float(balance_after) if balance_after and balance_after != 'N/A' else 0.0
                    if balance > 0:
                        self.tree.reattach(item, '', 'end')
                        has_balance_count += 1
                except:
                    pass
        
        if has_balance_count > 0:
            self.log(f"✓ 已筛选出 {has_balance_count} 个有余额的账户")
        else:
            self.log("✓ 没有找到有余额的账户")
            messagebox.showinfo("提示", "没有找到有余额的账户")
    
    def _filter_no_balance(self):
        """筛选无余额的账户（余额为0）"""
        # 保存所有项目ID（如果还没保存）
        if not self.all_tree_items:
            self.all_tree_items = list(self.tree.get_children())
        
        if not self.all_tree_items:
            self.log("⚠️ 表格中没有数据")
            messagebox.showinfo("提示", "表格中没有数据")
            return
        
        # 先detach所有项目
        for item in self.all_tree_items:
            self.tree.detach(item)
        
        # 只reattach无余额的账户（余额 == 0）
        no_balance_count = 0
        for item in self.all_tree_items:
            values = self.tree.item(item, 'values')
            if values and len(values) > 8:  # 确保有足够的列
                balance_after = values[8]  # 余额列是第9列（索引8）
                try:
                    balance = float(balance_after) if balance_after and balance_after != 'N/A' else 0.0
                    if balance == 0:
                        self.tree.reattach(item, '', 'end')
                        no_balance_count += 1
                except:
                    pass
        
        if no_balance_count > 0:
            self.log(f"✓ 已筛选出 {no_balance_count} 个无余额的账户")
        else:
            self.log("✓ 没有找到无余额的账户")
            messagebox.showinfo("提示", "没有找到无余额的账户")
    
    def _show_all(self):
        """显示全部账户（清除筛选）"""
        # 如果有保存的项目，恢复所有项目
        if self.all_tree_items:
            # 先detach所有
            for item in self.all_tree_items:
                try:
                    self.tree.detach(item)
                except:
                    pass
            
            # 重新attach所有项目
            for item in self.all_tree_items:
                try:
                    self.tree.reattach(item, '', 'end')
                except:
                    pass
            
            self.log(f"✓ 已显示全部账户（共 {len(self.all_tree_items)} 个）")
        else:
            self.log("✓ 已显示全部账户")
    
    def _search_history_table(self):
        """搜索历史记录表格（根据手机号或ID）"""
        search_text = self.history_search_var.get().strip()
        
        if not search_text:
            # 如果搜索框为空，显示全部
            self._show_all()
            return
        
        # 保存所有项目ID（如果还没保存）
        if not self.all_tree_items:
            self.all_tree_items = list(self.tree.get_children())
        
        # 先detach所有项目
        for item in self.all_tree_items:
            try:
                self.tree.detach(item)
            except:
                pass
        
        # 搜索匹配的项目
        matched_items = []
        for item in self.all_tree_items:
            try:
                values = self.tree.item(item, 'values')
                if values and len(values) > 2:
                    nickname = str(values[0])  # 昵称在第一列
                    user_id = str(values[1])  # 用户ID在第二列
                    phone = str(values[2])  # 手机号在第三列
                    
                    # 模糊匹配：手机号或ID包含搜索文本
                    if search_text in phone or search_text in user_id:
                        matched_items.append(item)
            except:
                pass
        
        # 显示匹配的项目
        for item in matched_items:
            try:
                self.tree.reattach(item, '', 'end')
            except:
                pass
        
        if matched_items:
            self.log(f"🔍 找到 {len(matched_items)} 个匹配的账户")
        else:
            self.log(f"🔍 未找到匹配 '{search_text}' 的账户")
            messagebox.showinfo("提示", f"未找到匹配 '{search_text}' 的账户")
    
    
    def _on_closing(self):
        """安全关闭窗口"""
        try:
            # 清理树形视图
            for item in self.tree.get_children():
                self.tree.delete(item)
            
            # 清空结果列表
            self.results.clear()
            
            # 销毁窗口
            self.window.destroy()
        except Exception as e:
            print(f"关闭历史结果窗口时出错: {e}")
            try:
                self.window.destroy()
            except:
                pass



class WindowArrangerDialog:
    """窗口排列对话框"""
    
    def __init__(self, parent, log_callback=None):
        """初始化对话框
        
        Args:
            parent: 父窗口
            log_callback: 日志回调函数
        """
        self.parent = parent
        self.log_callback = log_callback
        
        # 创建对话框窗口
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("窗口自动排列")
        self.dialog.geometry("500x600")
        self.dialog.resizable(False, False)
        
        # 居中显示
        self._center_window()
        
        # 获取对话框窗口句柄
        self.dialog_hwnd = None
        try:
            # 等待窗口完全创建
            self.dialog.update_idletasks()
            # 获取Tkinter窗口的Windows句柄
            self.dialog_hwnd = int(self.dialog.wm_frame(), 16)
        except Exception as e:
            if log_callback:
                log_callback(f"⚠️ 无法获取对话框窗口句柄: {e}")
        
        # 创建窗口排列器
        try:
            from .window_arranger import WindowArranger
            self.arranger = WindowArranger()
        except ImportError:
            from src.window_arranger import WindowArranger
            self.arranger = WindowArranger()
        
        # 创建界面
        self._create_widgets()
        
        # 刷新窗口列表
        self._refresh_windows()
    
    def _center_window(self):
        """居中显示对话框"""
        self.dialog.update_idletasks()
        
        width = 500
        height = 600
        
        # 相对于父窗口居中
        parent_x = self.parent.winfo_x()
        parent_y = self.parent.winfo_y()
        parent_width = self.parent.winfo_width()
        parent_height = self.parent.winfo_height()
        
        x = parent_x + (parent_width - width) // 2
        y = parent_y + (parent_height - height) // 2
        
        self.dialog.geometry(f'{width}x{height}+{x}+{y}')
    
    def _create_widgets(self):
        """创建界面组件"""
        main_frame = ttk.Frame(self.dialog, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # === 窗口信息区域 ===
        info_frame = ttk.LabelFrame(main_frame, text="检测到的窗口", padding="5")
        info_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # 窗口列表
        list_frame = ttk.Frame(info_frame)
        list_frame.pack(fill=tk.BOTH, expand=True)
        
        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.window_listbox = tk.Listbox(
            list_frame,
            yscrollcommand=scrollbar.set,
            height=8,
            font=("Microsoft YaHei UI", 9)
        )
        self.window_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.window_listbox.yview)
        
        # 刷新按钮
        ttk.Button(info_frame, text="🔄 刷新", command=self._refresh_windows, width=10).pack(pady=(5, 0))
        
        # === 快速排列区域 ===
        preset_frame = ttk.LabelFrame(main_frame, text="快速排列", padding="5")
        preset_frame.pack(fill=tk.X, pady=(0, 10))
        
        # 平铺按钮
        ttk.Button(
            preset_frame,
            text="📐 平铺（不重叠）",
            command=lambda: self._apply_layout("tile"),
            width=25
        ).pack(pady=(0, 15))
        
        # 堆叠覆盖率设置（居中）
        stack_settings_frame = ttk.Frame(preset_frame)
        stack_settings_frame.pack(pady=(0, 10))
        
        ttk.Label(stack_settings_frame, text="堆叠覆盖率:", font=("Microsoft YaHei UI", 9)).pack(side=tk.LEFT, padx=(0, 10))
        
        self.overlap_var = tk.IntVar(value=50)
        overlap_spinbox = ttk.Spinbox(
            stack_settings_frame,
            from_=10,
            to=90,
            textvariable=self.overlap_var,
            width=10
        )
        overlap_spinbox.pack(side=tk.LEFT, padx=(0, 5))
        
        ttk.Label(stack_settings_frame, text="%", font=("Microsoft YaHei UI", 9)).pack(side=tk.LEFT)
        
        # 堆叠按钮
        ttk.Button(
            preset_frame,
            text="📚 堆叠（重叠排列）",
            command=self._apply_stack_layout,
            width=25
        ).pack()
        
        # === 底部按钮 ===
        bottom_frame = ttk.Frame(main_frame)
        bottom_frame.pack(fill=tk.X, pady=(10, 0))
        
        ttk.Button(
            bottom_frame,
            text="关闭",
            command=self.dialog.destroy,
            width=15
        ).pack(side=tk.RIGHT)
    
    def _refresh_windows(self):
        """刷新窗口列表"""
        self.window_listbox.delete(0, tk.END)
        
        windows = self.arranger.find_emulator_windows()
        
        if not windows:
            self.window_listbox.insert(tk.END, "未找到模拟器窗口")
            return
        
        for i, hwnd in enumerate(windows, 1):
            info = self.arranger.get_window_info(hwnd)
            if info:
                self.window_listbox.insert(
                    tk.END,
                    f"{i}. {info['title']} ({info['width']}x{info['height']})"
                )
    
    def _apply_layout(self, layout_key: str):
        """应用预设布局（保持窗口原有尺寸）
        
        Args:
            layout_key: 布局键名
        """
        try:
            result = self.arranger.arrange_windows(
                layout_key, 
                keep_size=True, 
                minimize_others=True,
                dialog_hwnd=self.dialog_hwnd
            )
            
            if result['success']:
                message = f"✅ {result['message']}"
                if self.log_callback:
                    self.log_callback(message)
                messagebox.showinfo("成功", message, parent=self.dialog)
            else:
                message = f"❌ {result['message']}"
                if self.log_callback:
                    self.log_callback(message)
                messagebox.showerror("失败", message, parent=self.dialog)
            
            # 刷新窗口列表
            self._refresh_windows()
            
        except Exception as e:
            error_msg = f"排列窗口失败: {str(e)}"
            if self.log_callback:
                self.log_callback(f"❌ {error_msg}")
            messagebox.showerror("错误", error_msg, parent=self.dialog)
    
    def _apply_stack_layout(self):
        """应用堆叠布局（使用自定义覆盖率）"""
        try:
            overlap_percent = self.overlap_var.get()
            result = self.arranger.arrange_windows(
                "stack", 
                keep_size=True, 
                overlap_percent=overlap_percent,
                minimize_others=True,
                dialog_hwnd=self.dialog_hwnd
            )
            
            if result['success']:
                message = f"✅ {result['message']}"
                if self.log_callback:
                    self.log_callback(message)
                messagebox.showinfo("成功", message, parent=self.dialog)
            else:
                message = f"❌ {result['message']}"
                if self.log_callback:
                    self.log_callback(message)
                messagebox.showerror("失败", message, parent=self.dialog)
            
            # 刷新窗口列表
            self._refresh_windows()
            
        except Exception as e:
            error_msg = f"排列窗口失败: {str(e)}"
            if self.log_callback:
                self.log_callback(f"❌ {error_msg}")
            messagebox.showerror("错误", error_msg, parent=self.dialog)


class WorkflowControlWindow:
    """流程控制窗口"""
    
    def __init__(self, parent, gui_instance):
        """初始化流程控制窗口
        
        Args:
            parent: 父窗口
            gui_instance: AutomationGUI实例
        """
        self.parent = parent
        self.gui = gui_instance
        
        # 创建顶层窗口
        self.window = tk.Toplevel(parent)
        self.window.title("流程控制")
        self.window.geometry("500x450")
        self.window.resizable(False, False)
        
        # 窗口居中
        self._center_window()
        
        # 加载当前配置
        self._load_config()
        
        # 创建界面
        self._create_widgets()
        
        # 设置窗口关闭协议
        self.window.protocol("WM_DELETE_WINDOW", self._on_closing)
    
    def _center_window(self):
        """将窗口居中显示"""
        self.window.update_idletasks()
        width = 500
        height = 450
        screen_width = self.window.winfo_screenwidth()
        screen_height = self.window.winfo_screenheight()
        x = (screen_width // 2) - (width // 2)
        y = (screen_height // 2) - (height // 2)
        self.window.geometry(f'{width}x{height}+{x}+{y}')
    
    def _load_config(self):
        """加载当前流程配置"""
        # 从config加载，如果没有则使用默认值
        config = self.gui.config
        self.workflow_mode = getattr(config, 'workflow_mode', 'complete')
        self.enable_login = getattr(config, 'workflow_enable_login', True)
        self.enable_profile = getattr(config, 'workflow_enable_profile', True)
        self.enable_checkin = getattr(config, 'workflow_enable_checkin', True)
        self.enable_transfer = getattr(config, 'workflow_enable_transfer', True)
    
    def _create_widgets(self):
        """创建界面组件"""
        main_frame = ttk.Frame(self.window, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 标题
        title_label = ttk.Label(main_frame, text="流程控制设置", font=("Microsoft YaHei UI", 12, "bold"))
        title_label.pack(pady=(0, 15))
        
        # 说明文字
        desc_label = ttk.Label(main_frame, text="选择要执行的流程模块，自由组合工作流程", foreground="gray")
        desc_label.pack(pady=(0, 10))
        
        # === 预设模式选择 ===
        mode_frame = ttk.LabelFrame(main_frame, text="预设模式", padding="10")
        mode_frame.pack(fill=tk.X, pady=(0, 15))
        
        self.mode_var = tk.StringVar(value=self.workflow_mode)
        
        modes = [
            ("complete", "完整流程", "登录 → 获取资料 → 签到 → 获取最终余额 → 转账"),
            ("quick_checkin", "快速签到", "登录 → 签到 → 获取余额 → 转账（跳过签到前资料获取）"),
            ("login_only", "只登录", "仅登录（或验证缓存）→ 获取资料"),
            ("transfer_only", "只转账", "登录 → 获取资料 → 转账（跳过签到）"),
            ("custom", "自定义", "自由勾选需要的流程模块")
        ]
        
        for value, text, desc in modes:
            rb = ttk.Radiobutton(mode_frame, text=text, variable=self.mode_var, value=value, 
                                command=self._on_mode_changed)
            rb.pack(anchor=tk.W, pady=2)
            desc_label = ttk.Label(mode_frame, text=f"  {desc}", foreground="gray", font=("Microsoft YaHei UI", 8))
            desc_label.pack(anchor=tk.W, padx=(20, 0))
        
        # === 自定义流程模块 ===
        custom_frame = ttk.LabelFrame(main_frame, text="自定义流程模块", padding="10")
        custom_frame.pack(fill=tk.X, pady=(0, 15))
        
        self.login_var = tk.BooleanVar(value=self.enable_login)
        self.profile_var = tk.BooleanVar(value=self.enable_profile)
        self.checkin_var = tk.BooleanVar(value=self.enable_checkin)
        self.transfer_var = tk.BooleanVar(value=self.enable_transfer)
        
        ttk.Checkbutton(custom_frame, text="☑️ 登录", variable=self.login_var).pack(anchor=tk.W, pady=3)
        ttk.Label(custom_frame, text="  执行登录流程（或使用缓存验证）", foreground="gray", font=("Microsoft YaHei UI", 8)).pack(anchor=tk.W, padx=(20, 0))
        
        ttk.Checkbutton(custom_frame, text="☑️ 获取资料", variable=self.profile_var).pack(anchor=tk.W, pady=3)
        ttk.Label(custom_frame, text="  获取个人资料（昵称、ID、余额、积分等）", foreground="gray", font=("Microsoft YaHei UI", 8)).pack(anchor=tk.W, padx=(20, 0))
        
        ttk.Checkbutton(custom_frame, text="☑️ 签到", variable=self.checkin_var).pack(anchor=tk.W, pady=3)
        ttk.Label(custom_frame, text="  执行签到流程并获取签到奖励", foreground="gray", font=("Microsoft YaHei UI", 8)).pack(anchor=tk.W, padx=(20, 0))
        
        ttk.Checkbutton(custom_frame, text="☑️ 转账", variable=self.transfer_var).pack(anchor=tk.W, pady=3)
        ttk.Label(custom_frame, text="  执行转账流程（受'自动转账'开关控制）", foreground="gray", font=("Microsoft YaHei UI", 8)).pack(anchor=tk.W, padx=(20, 0))
        
        # 初始状态：如果不是自定义模式，禁用复选框
        if self.mode_var.get() != "custom":
            for var in [self.login_var, self.profile_var, self.checkin_var, self.transfer_var]:
                for widget in custom_frame.winfo_children():
                    if isinstance(widget, ttk.Checkbutton):
                        widget.configure(state=tk.DISABLED)
        
        # === 按钮区域 ===
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(10, 0))
        
        ttk.Button(button_frame, text="保存", command=self._save_config, width=12).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(button_frame, text="取消", command=self._on_closing, width=12).pack(side=tk.LEFT)
        ttk.Button(button_frame, text="恢复默认", command=self._reset_to_default, width=12).pack(side=tk.RIGHT)
    
    def _on_mode_changed(self):
        """预设模式改变时的回调"""
        mode = self.mode_var.get()
        
        # 根据模式设置复选框状态
        if mode == "complete":
            # 完整流程：全部启用
            self.login_var.set(True)
            self.profile_var.set(True)
            self.checkin_var.set(True)
            self.transfer_var.set(True)
            self._disable_checkboxes()
        elif mode == "quick_checkin":
            # 快速签到：登录 + 签到 + 转账（跳过签到前的资料获取）
            self.login_var.set(True)
            self.profile_var.set(False)  # 跳过签到前的资料获取
            self.checkin_var.set(True)
            self.transfer_var.set(True)  # 启用转账（签到后会获取用户ID）
            self._disable_checkboxes()
        elif mode == "login_only":
            # 只登录：登录 + 获取资料
            self.login_var.set(True)
            self.profile_var.set(True)
            self.checkin_var.set(False)
            self.transfer_var.set(False)
            self._disable_checkboxes()
        elif mode == "transfer_only":
            # 只转账：登录 + 获取资料 + 转账
            self.login_var.set(True)
            self.profile_var.set(True)
            self.checkin_var.set(False)
            self.transfer_var.set(True)
            self._disable_checkboxes()
        elif mode == "custom":
            # 自定义：启用复选框
            self._enable_checkboxes()
        
        # 实时保存配置
        self._auto_save_config()
    
    def _disable_checkboxes(self):
        """禁用自定义流程模块的复选框"""
        for widget in self.window.winfo_children():
            self._disable_checkboxes_recursive(widget)
    
    def _disable_checkboxes_recursive(self, widget):
        """递归禁用复选框"""
        if isinstance(widget, ttk.Checkbutton):
            widget.configure(state=tk.DISABLED)
        for child in widget.winfo_children():
            self._disable_checkboxes_recursive(child)
    
    def _enable_checkboxes(self):
        """启用自定义流程模块的复选框"""
        for widget in self.window.winfo_children():
            self._enable_checkboxes_recursive(widget)
    
    def _enable_checkboxes_recursive(self, widget):
        """递归启用复选框"""
        if isinstance(widget, ttk.Checkbutton):
            widget.configure(state=tk.NORMAL)
        for child in widget.winfo_children():
            self._enable_checkboxes_recursive(child)
    
    def _auto_save_config(self):
        """自动保存流程配置（不显示消息）"""
        # 保存到config
        self.gui.config.workflow_mode = self.mode_var.get()
        self.gui.config.workflow_enable_login = self.login_var.get()
        self.gui.config.workflow_enable_profile = self.profile_var.get()
        self.gui.config.workflow_enable_checkin = self.checkin_var.get()
        self.gui.config.workflow_enable_transfer = self.transfer_var.get()
        
        # 保存配置文件
        self.gui.config_loader.save(self.gui.config)
        
        # 在GUI日志中显示当前模式
        mode_names = {
            "complete": "完整流程",
            "quick_checkin": "快速签到",
            "login_only": "只登录",
            "transfer_only": "只转账",
            "custom": "自定义"
        }
        mode_name = mode_names.get(self.mode_var.get(), "未知")
        self.gui._log(f"✓ 流程模式已切换: {mode_name}")
    
    def _save_config(self):
        """保存流程配置（显示消息）"""
        # 保存到config
        self.gui.config.workflow_mode = self.mode_var.get()
        self.gui.config.workflow_enable_login = self.login_var.get()
        self.gui.config.workflow_enable_profile = self.profile_var.get()
        self.gui.config.workflow_enable_checkin = self.checkin_var.get()
        self.gui.config.workflow_enable_transfer = self.transfer_var.get()
        
        # 保存配置文件
        self.gui.config_loader.save(self.gui.config)
        
        # 显示保存成功消息
        mode_names = {
            "complete": "完整流程",
            "quick_checkin": "快速签到",
            "login_only": "只登录",
            "transfer_only": "只转账",
            "custom": "自定义"
        }
        mode_name = mode_names.get(self.mode_var.get(), "未知")
        
        messagebox.showinfo("成功", f"流程控制已保存\n当前模式: {mode_name}", parent=self.window)
        self.gui._log(f"✓ 流程控制已保存: {mode_name}")
        
        # 关闭窗口
        self._on_closing()
    
    def _reset_to_default(self):
        """恢复默认设置"""
        self.mode_var.set("complete")
        self.login_var.set(True)
        self.profile_var.set(True)
        self.checkin_var.set(True)
        self.transfer_var.set(True)
        self._on_mode_changed()
    
    def _on_closing(self):
        """窗口关闭时的处理"""
        # 关闭窗口时也保存配置
        self._auto_save_config()
        self.window.destroy()
    
    def winfo_exists(self):
        """检查窗口是否存在"""
        try:
            return self.window.winfo_exists()
        except:
            return False
    
    def lift(self):
        """将窗口提升到最前"""
        self.window.lift()
    
    def focus_force(self):
        """强制获取焦点"""
        self.window.focus_force()


def main(adb_bridge=None):
    """GUI主函数入口
    
    Args:
        adb_bridge: ADBBridge实例（可选），用于后台加载模型
    """
    gui = AutomationGUI(adb_bridge=adb_bridge)
    gui.root.mainloop()


if __name__ == "__main__":
    main()
